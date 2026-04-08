import sys, io
# Force UTF-8 output so emoji in print() never crash on Windows cp1252 consoles
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace', line_buffering=True)
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace', line_buffering=True)

from flask import Flask, request, jsonify, send_from_directory, send_file, Response, make_response, g
import requests as req_lib
from bs4 import BeautifulSoup
import os
import re
import json
import traceback
import subprocess
import gzip
import math
import shutil
import tempfile
from collections import Counter
from datetime import datetime
from pathlib import Path
import fitz          # PyMuPDF  — PDF → image
import pytesseract
from PIL import Image
import io
import xml_processor
import ezdxf

# ── Profile support (multi-user) ──────────────────────────────────────────────
from helpers.profiles import (
    list_profiles, get_profile, save_profile, create_profile,
    delete_profile, update_profile_field, migrate_from_config,
)

# ── SaaS auth & subscription gating ───────────────────────────────────────────
from helpers.auth import (
    create_user, find_user_by_email, get_user as get_saas_user,
    verify_password, generate_token, verify_token,
    increment_search_count, reset_monthly_counts_if_needed, public_user,
    generate_reset_token, reset_password as reset_user_password,
    add_search_history, get_search_history,
    check_login_allowed, record_failed_login, clear_failed_logins,
)
from helpers.subscription import (
    require_auth, require_pro, require_team,
    check_search_quota, has_feature, get_tier_limits,
)
from helpers.stripe_webhook import verify_and_parse as _stripe_verify, dispatch_event as _stripe_dispatch
from helpers.admin import (
    check_admin_password, list_users_summary, get_user_stats,
    admin_set_tier, admin_toggle_active, admin_reset_searches,
)
from helpers.rate_limit import rate_limit, rate_limit_ip
from helpers.backup import list_backups, restore_backup
from helpers.teams import (
    get_team_members, get_seat_count, invite_member,
    accept_invite, remove_member, leave_team,
    verify_invite_token,
)


# ── Helper modules (extracted from this file for maintainability) ─────────────
from helpers.metes_bounds import (
    parse_metes_bounds, calls_to_coords, calls_to_full_coords, _bearing_to_azimuth,
    extract_trs, detect_monuments, classify_description_type,
    shoelace_area, has_pob,
    _BEARING_PAT, _BEARING_VERBOSE, _CURVE_PAT,
    _MONUMENT_PATTERNS, _LOT_BLOCK_RE, _TRACT_RE, _POB_RE,
)
from helpers.pdf_extract import (
    extract_pdf_text as _extract_pdf_text_impl,
    setup_tesseract, ocr_plat_file as _ocr_plat_file_impl,
    _ocr_cache_path,
)
from helpers.adjoiner import (
    parse_adjoiner_names as _parse_adjoiner_names_impl,
    _ADJ_PATTERNS, _NOISE_WORDS,
)
from helpers.cabinet import (
    CABINET_FOLDERS, parse_cabinet_refs,
    extract_plat_name_tokens as _extract_plat_name_tokens,
    extract_cabinet_display_name as _extract_cabinet_display_name,
    extract_cabinet_doc_number as _extract_cabinet_doc_number,
    search_local_cabinet as _search_local_cabinet_impl,
    _warm_cabinet_caches, _init_index_path,
    _cab_scan_cache,
)
from helpers.deed_analysis import (
    analyze_deed as _analyze_deed_impl,
    isolate_legal_description as _isolate_legal_description_impl,
)
from helpers.dxf import generate_boundary_dxf as _generate_dxf_impl
from helpers.legal_similarity import search_similar_descriptions as _search_similar_descriptions
from helpers.research_analytics import (
    get_analytics as _get_research_analytics,
    score_session_completeness as _score_session_completeness,
    predict_job_complexity as _predict_job_complexity,
    scan_all_research as _scan_all_research,
)
from helpers.county_registry import search_counties, get_county, get_all_counties

# Stripe billing (optional — import gracefully so app works without stripe installed)
try:
    from helpers.stripe_billing import (
        create_checkout_session, create_customer_portal_session, handle_webhook,
        STRIPE_SECRET_KEY as _STRIPE_KEY_SET,
    )
    _STRIPE_AVAILABLE = bool(_STRIPE_KEY_SET)
except ImportError:
    _STRIPE_AVAILABLE = False
    create_checkout_session = create_customer_portal_session = handle_webhook = None

# Point pytesseract at the Tesseract binary (delegated to helpers/pdf_extract.py)
setup_tesseract()

app = Flask(__name__, static_folder='.', static_url_path='')

# ── Security configuration ────────────────────────────────────────────────────
_is_production = os.environ.get("DEED_APP_URL", "").startswith("https://")
app.config.update(
    SECRET_KEY=os.environ.get("DEED_SECRET_KEY", "dev-insecure-key-change-me"),
    SESSION_COOKIE_SECURE=_is_production,      # HTTPS-only cookies in prod
    SESSION_COOKIE_HTTPONLY=True,              # JS cannot read session cookies
    SESSION_COOKIE_SAMESITE="Lax",            # CSRF mitigation
)

# ── Register Blueprints ───────────────────────────────────────────────────────
from routes.auth import auth_bp
from routes.stripe import stripe_bp, init_stripe
from routes.admin import admin_bp
from routes.team import team_bp

app.register_blueprint(auth_bp)
app.register_blueprint(stripe_bp)
app.register_blueprint(admin_bp)
app.register_blueprint(team_bp)

# AI integration (optional — works without AI deps installed)
try:
    from ai.routes import ai_bp
    app.register_blueprint(ai_bp)
    print("[ai] AI endpoints registered at /api/ai/*", flush=True)
except ImportError:
    print("[ai] AI modules not available — install networkx, scikit-learn to enable", flush=True)


# Inject Stripe functions into the Stripe Blueprint (avoids circular import)
init_stripe(
    available=_STRIPE_AVAILABLE,
    checkout_fn=create_checkout_session,
    portal_fn=create_customer_portal_session,
    verify_fn=_stripe_verify,
    dispatch_fn=_stripe_dispatch,
)

# Default portal URL — overridden per-user via Settings → URL field.
# Each user enters the URL for their own county records portal (e.halFILE, etc.).
_DEFAULT_PORTAL_URL = "http://records.1stnmtitle.com"

def _get_portal_url() -> str:
    """Return the county records portal base URL for the current request.
    Reads from the active profile first, falls back to global config, then
    falls back to the compiled-in default.  Strips trailing slashes."""
    try:
        pid = request.cookies.get('profile_id')
    except RuntimeError:
        pid = None  # called outside request context

    url = ""
    if pid:
        p = get_profile(pid)
        if p:
            url = p.get("firstnm_url", "").strip()
    if not url:
        cfg = load_config()
        url = cfg.get("firstnm_url", "").strip()
    return (url or _DEFAULT_PORTAL_URL).rstrip("/")


# ── ArcGIS Parcel Layer — configurable per user ────────────────────────────────
# The app uses an ArcGIS REST FeatureService/MapServer layer for:
#   • Parcel geometry  (map picker + spatial adjoiner discovery)
#   • Owner name       (relevance scoring)
#   • TRS (Township/Range/Section)  (relevance scoring)
#   • Property address (address lookup)
#
# Each county organises its ArcGIS layer differently.  The user configures:
#   arcgis_url   — full query URL, e.g. .../MapServer/29/query
#   arcgis_fields — dict mapping concept → actual attribute field name
#
# ─────────────────────────────────────────────────────────────────────────────

# Built-in preset: NM OSE statewide parcel service, Taos County layer
_ARCGIS_PRESETS = {
    "taos_nm": {
        "label": "Taos County, NM (default)",
        "url": "https://gis.ose.nm.gov/server_s/rest/services/Parcels/County_Parcels_2025/MapServer/29/query",
        "fields": {
            "parcel_id":   "UPC",
            "owner":       "OwnerAll",
            "address_all": "SitusAddressAll",
            "address1":    "SitusAddress1",
            "street_no":   "SitusStreetNumber",
            "street_name": "SitusStreetName",
            "city":        "SitusCity",
            "zipcode":     "SitusZipCode",
            "legal":       "LegalDescription",
            "area":        "LandArea",
            "subdivision": "Subdivision",
            "zoning":      "ZoningDescription",
            "land_use":    "LandUseDescription",
            "township":    "Township",
            "twp_dir":     "TownshipDirection",
            "range":       "Range",
            "rng_dir":     "RangeDirection",
            "section":     "Section",
            "struct_count":"StructureCount",
            "struct_type": "StructureType",
            "owner_type":  "OwnerType",
            "mail_addr":   "MailAddressAll",
        },
    },
}

# Concepts that are optional — silently skipped if the field is not configured
_ARCGIS_OPTIONAL_FIELDS = {
    "address1", "street_no", "street_name", "city", "zipcode",
    "zoning", "land_use", "struct_count", "struct_type", "owner_type",
    "mail_addr", "twp_dir", "rng_dir",
}


def _get_arcgis_config() -> dict:
    """Return the ArcGIS layer config for the current request.

    Reads from active profile → global config → Taos NM default.
    Returns dict: { url, fields: { concept: field_name } }
    """
    try:
        pid = request.cookies.get('profile_id')
    except RuntimeError:
        pid = None

    stored = None
    if pid:
        p = get_profile(pid)
        if p and p.get('arcgis_url'):
            stored = p
    if not stored:
        cfg = load_config()
        if cfg.get('arcgis_url'):
            stored = cfg

    if stored:
        # Merge user-supplied fields over the default field map
        default_fields = dict(_ARCGIS_PRESETS['taos_nm']['fields'])
        user_fields = stored.get('arcgis_fields') or {}
        default_fields.update({k: v for k, v in user_fields.items() if v})
        return {
            'url':    stored['arcgis_url'].rstrip('/'),
            'fields': default_fields,
        }

    # No user config — use built-in Taos NM default
    preset = _ARCGIS_PRESETS['taos_nm']
    return {'url': preset['url'], 'fields': dict(preset['fields'])}


def _arcgis_field(cfg: dict, concept: str) -> str:
    """Return the actual ArcGIS attribute field name for a logical concept.
    Falls back to the Taos default if not configured."""
    return cfg['fields'].get(concept) or _ARCGIS_PRESETS['taos_nm']['fields'].get(concept, '')


def _arcgis_out_fields(cfg: dict, concepts: list) -> str:
    """Build a comma-separated outFields string for a given list of concepts."""
    fields = []
    for c in concepts:
        f = _arcgis_field(cfg, c)
        if f and f not in fields:
            fields.append(f)
    return ','.join(fields)


# ── Removable-drive detection ───────────────────────────────────────────────────
# The Survey Data folder lives on a removable drive whose letter changes
# between computers.  We scan all available drive letters at startup and
# cache the result; the /api/drive-status endpoint lets you re-scan live.

_SURVEY_RELATIVE   = os.path.join("AI DATA CENTER", "Survey Data")
_CABINET_RELATIVE  = os.path.join("AI DATA CENTER", "Survey Data",
                                   "00 COUNTY CLERK SCANS Cabs A-B- C-D - E")
_detected_drive: str | None = None   # e.g. "F"

# ── Helpers (defined early so drive detection can use load_config) ──────────
CONFIG_FILE = os.path.join(os.path.dirname(__file__), "config.json")

def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_config(data):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def detect_survey_drive(force: bool = False) -> str | None:
    """Scan all drive letters for the Survey Data folder.
    Returns the drive letter (e.g. 'F') or None if not found.
    Caches the result; pass force=True to rescan.
    """
    global _detected_drive
    # ── Dev mode: DEV_DATA_DIR env var bypasses drive scanning ──────────────
    dev_dir = os.environ.get("DEV_DATA_DIR", "").strip()
    if dev_dir and Path(dev_dir).exists():
        _detected_drive = "__dev__"
        if not force:
            return _detected_drive
        print(f"[drive] DEV MODE — using local data: {dev_dir}", flush=True)
        return _detected_drive
    if _detected_drive and not force:
        # Verify cached drive is still present
        if _detected_drive == "__dev__" or Path(f"{_detected_drive}:\\").exists():
            return _detected_drive

    # Try config override first
    cfg = load_config()
    override = cfg.get("survey_drive", "").strip().upper()
    if override and len(override) == 1 and Path(f"{override}:\\").exists():
        candidate = Path(f"{override}:\\") / _SURVEY_RELATIVE
        if candidate.exists():
            _detected_drive = override
            return _detected_drive
    # Scan preferred drives first (J = primary SSD), then remaining letters
    import string
    _PREFERRED = ['J', 'K', 'I', 'H', 'G', 'F', 'E']
    _REMAINING = [c for c in string.ascii_uppercase if c not in _PREFERRED]
    for letter in _PREFERRED + _REMAINING:
        root = Path(f"{letter}:\\")
        if not root.exists():
            continue
        candidate = root / _SURVEY_RELATIVE
        if candidate.exists():
            _detected_drive = letter
            print(f"[drive] Survey Data found on {letter}:\\", flush=True)
            return _detected_drive
    _detected_drive = None
    print("[drive] Survey Data NOT found on any drive.", flush=True)
    return None


def get_survey_data_path() -> str:
    """Return the current Survey Data path, auto-detecting the drive."""
    drive = detect_survey_drive()
    if drive == "__dev__":
        dev_dir = os.environ.get("DEV_DATA_DIR", "").strip()
        return dev_dir if dev_dir else ""
    if drive:
        return str(Path(f"{drive}:\\") / _SURVEY_RELATIVE)
    return ""  # drive not found — caller should check for empty string and warn user


def get_cabinet_path() -> str:
    """Return the current Cabinet path, auto-detecting the drive."""
    drive = detect_survey_drive()
    if drive:
        return str(Path(f"{drive}:\\") / _CABINET_RELATIVE)
    return ""  # drive not found — caller should check for empty string


# Kick off detection at startup (non-blocking — just sets module-level cache)
try:
    detect_survey_drive()
except Exception as e:
    print(f"[warn] drive detection at startup failed: {e}", flush=True)


# ── Cabinet index initialization ────────────────────────────────────────────
# Must run at import time so desktop_app.py (which imports this module)
# also gets the cabinet index loaded.  Previously this lived inside
# `if __name__ == "__main__"` and was skipped when imported.
try:
    _init_index_path(os.path.dirname(os.path.abspath(__file__)))
    _cab_path_init = get_cabinet_path()
    if _cab_path_init:
        _warm_cabinet_caches(_cab_path_init)
    else:
        print("[cabinet] No cabinet path found — cabinet search will be unavailable until drive is connected", flush=True)
except Exception as _cab_init_err:
    print(f"[cabinet] Initialization failed: {_cab_init_err}", flush=True)


# CABINET_FOLDERS — imported from helpers.cabinet
# CONFIG_FILE, load_config, save_config — defined above (before drive detection)

# Must match the <select> options in index.html
JOB_TYPES = ["BDY", "ILR", "SE", "SUB", "TIE", "TOPO", "ELEV", "ALTA", "CONS", "OTHER"]

# ── Per-user web sessions (multi-user support) ──────────────────────────────
# Each profile gets its own requests.Session so login cookies don't collide.
import threading
_user_sessions: dict[str, req_lib.Session] = {}   # profile_id -> Session
_user_sessions_lock = threading.Lock()

_DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

def _get_web_session(profile_id: str | None = None) -> req_lib.Session:
    """Return a per-user requests.Session.  Falls back to a shared one."""
    key = profile_id or "__default__"
    with _user_sessions_lock:
        if key not in _user_sessions:
            s = req_lib.Session()
            s.headers.update(_DEFAULT_HEADERS)
            _user_sessions[key] = s
        return _user_sessions[key]

# Backward-compat: callers that still reference `web_session` get the default
web_session = _get_web_session()

def _session() -> req_lib.Session:
    """Return the web session for the current request's profile.
    Reads the profile_id cookie set by the frontend and dispatches to
    the correct per-user requests.Session.  Falls back to the shared
    default session if no cookie is present."""
    try:
        pid = request.cookies.get('profile_id')
    except RuntimeError:
        pid = None  # called outside request context
    return _get_web_session(pid)


def _auto_login() -> bool:
    """Log into the 1stNMTitle portal using config.json credentials.

    Used by the auto-research background worker which has no browser
    cookie/session.  Returns True if login succeeded, False otherwise.
    The authenticated session is stored under the '__auto__' profile key.
    """
    try:
        cfg = load_config()
        username = cfg.get("firstnm_user", "")
        password = cfg.get("firstnm_pass", "")
        if not username or not password:
            print("[auto-login] No credentials in config.json — skipping portal login", flush=True)
            return False

        sess = _get_web_session("__auto__")

        # Check if already logged in
        check = sess.get(_get_portal_url() + "/scripts/hfweb.asp?Application=FNM&Database=TP", timeout=10)
        if 'CROSSNAMEFIELD' in check.text:
            print("[auto-login] Session already active", flush=True)
            return True

        # Fetch login page
        resp = sess.get(_get_portal_url() + "/", timeout=10)
        soup = BeautifulSoup(resp.text, "lxml")
        form = soup.find("form")
        if not form:
            print("[auto-login] Login form not found", flush=True)
            return False

        action = form.get("action", "/")
        if not action.startswith("http"):
            action = _get_portal_url() + "/" + action.lstrip("/")

        form_data = {}
        for inp in form.find_all('input'):
            nm = inp.get('name')
            itype = (inp.get('type') or 'text').lower()
            if nm:
                if itype == 'image':
                    form_data[nm + '.x'] = '1'
                    form_data[nm + '.y'] = '1'
                else:
                    form_data[nm] = inp.get('value', '')

        # Map credentials (e.halFILE field names)
        for inp in form.find_all('input'):
            nm = inp.get('name', '')
            if nm == 'FormUser':
                form_data['FormUser'] = username
            elif nm == 'FormPassword':
                form_data['FormPassword'] = password

        post_resp = sess.post(action, data=form_data, timeout=10)
        portal_root = _get_portal_url().lower().rstrip('/')
        landed_url  = post_resp.url.lower()
        success = ('hfweb' in landed_url or
                   'new search' in post_resp.text.lower() or
                   ('logout' in post_resp.text.lower() and
                    landed_url.rstrip('/') != portal_root))

        if success:
            print(f"[auto-login] ✓ Logged in as {username}", flush=True)
        else:
            print(f"[auto-login] ✗ Login failed for {username}", flush=True)
        return success

    except Exception as e:
        print(f"[auto-login] Error: {e}", flush=True)
        return False

# ── helpers (load_config / save_config defined above, before drive detection) ──

def _trigger_portal_pdf(doc_no: str):
    """Visit the portal document detail page to force PDF generation in WebTemp."""
    try:
        cfg = load_config()
        fuser = ""
        try:
            pid = request.cookies.get("profile_id")
        except RuntimeError:
            pid = None
        if pid:
            prof = get_profile(pid)
            if prof:
                fuser = prof.get("firstnm_user", "")
        if not fuser:
            fuser = cfg.get("firstnm_user", "")
        trigger_url = (f"{_get_portal_url()}/scripts/hfpage.asp"
                       f"?Appl=FNM&Doctype=TP&DocNo={doc_no}&FormUser={fuser}")
        # Use __auto__ session when outside request context (background thread)
        try:
            sess = _session()
        except RuntimeError:
            sess = _get_web_session("__auto__")
        sess.get(trigger_url, timeout=15)
    except Exception:
        pass  # best-effort


def _fetch_portal_pdf(doc_no: str, pdf_url: str = "", max_retries: int = 3):
    """Trigger the portal and fetch the PDF with automatic retry.

    The portal may need a moment after the trigger visit to generate
    the WebTemp PDF.  This helper retries with increasing delays
    (1s, 2s, 3s) if the first fetch returns 404.

    Returns (response, error_string).  On success error_string is None.
    """
    import time as _time
    if not pdf_url:
        pdf_url = f"{_get_portal_url()}/WebTemp/{doc_no}.pdf"

    _trigger_portal_pdf(doc_no)

    last_status = 0
    for attempt in range(max_retries):
        if attempt > 0:
            _time.sleep(attempt)  # 1s, 2s on retries
            print(f"[pdf-fetch] Retry {attempt}/{max_retries} for {doc_no}…", flush=True)
            _trigger_portal_pdf(doc_no)  # re-trigger on retry

        # Use __auto__ session when outside request context (background thread)
        try:
            _sess = _session()
        except RuntimeError:
            _sess = _get_web_session("__auto__")
        pdf_resp = _sess.get(pdf_url, stream=True, timeout=30)
        last_status = pdf_resp.status_code

        if pdf_resp.status_code == 200:
            # Verify we got an actual PDF, not an HTML error page
            ct = pdf_resp.headers.get("Content-Type", "")
            if "html" in ct.lower():
                return None, "Portal returned HTML instead of PDF — session may have expired"
            return pdf_resp, None

        if pdf_resp.status_code != 404:
            # Non-404 error — don't retry
            return None, f"PDF fetch failed: {pdf_resp.status_code}"

    return None, f"PDF fetch failed after {max_retries} attempts (last status: {last_status})"

def next_job_info():
    """Scan Survey Data to find the next job number and its range folder."""
    survey_str = get_survey_data_path()
    if not survey_str or not Path(survey_str).exists():
        # Drive not connected — return safe defaults
        return 2938, "2900-2999"
    survey = Path(survey_str)
    max_num = 0
    for child in survey.iterdir():
        if not child.is_dir():
            continue
        for job_dir in child.iterdir():
            if not job_dir.is_dir():
                continue
            m = re.match(r'^(\d{4})', job_dir.name)
            if m:
                max_num = max(max_num, int(m.group(1)))
    if max_num == 0:
        max_num = 2937  # fallback
    next_num = max_num + 1
    rstart = (next_num // 100) * 100
    return next_num, f"{rstart}-{rstart + 99}"

def create_project_folders(job_number, client_name, job_type):
    """Create the full folder tree and return the deeds path."""
    rstart = (int(job_number) // 100) * 100
    range_folder = f"{rstart}-{rstart + 99}"
    last_name = client_name.split(",")[0].strip()

    survey = Path(get_survey_data_path())
    client_path = survey / range_folder / f"{job_number} {client_name}"
    sub_path = client_path / f"{job_number}-01-{job_type} {last_name}"

    for folder in ["A Office", "B Drafting", "C Survey", "D Correspondence", "E Research", "F PROOFING"]:
        (sub_path / folder).mkdir(parents=True, exist_ok=True)
    for rf in ["A Deeds", "B Plats", "C Other"]:
        (sub_path / "E Research" / rf).mkdir(exist_ok=True)
    # Adjoiner subfolders
    (sub_path / "E Research" / "A Deeds" / "Adjoiners").mkdir(exist_ok=True)
    (sub_path / "E Research" / "B Plats" / "Adjoiners").mkdir(exist_ok=True)
    (client_path / "XXXX-00-LEGACY").mkdir(exist_ok=True)

    deeds_path = sub_path / "E Research" / "A Deeds"
    return str(client_path), str(deeds_path)

def _job_base_path(job_number, client_name: str, job_type: str) -> Path:
    """Return the canonical job subfolder path (does NOT create folders).

    Pattern: SurveyData / {range} / {job_number} {client_name} /
             {job_number}-01-{job_type} {last_name}
    """
    rstart    = (int(job_number) // 100) * 100
    last_name = client_name.split(",")[0].strip()
    return (
        Path(get_survey_data_path())
        / f"{rstart}-{rstart + 99}"
        / f"{job_number} {client_name}"
        / f"{job_number}-01-{job_type} {last_name}"
    )


def _next_ref_number(folder: Path, prefix: str = "D") -> int:
    """Return the next RTSI reference number for deeds (D1, D2...) or plats (P1, P2...).

    Scans `folder` for files whose names start with the prefix pattern
    (e.g. 'D1 ', 'D2 ', 'P1 ', 'P2 ') and returns max + 1.
    Falls back to 1 if no numbered files exist.
    """
    if not folder.is_dir():
        return 1
    max_num = 0
    pattern = re.compile(rf'^{prefix}(\d+)\s', re.IGNORECASE)
    for f in folder.iterdir():
        m = pattern.match(f.name)
        if m:
            max_num = max(max_num, int(m.group(1)))
    return max_num + 1


def _extract_pdf_text(pdf_path: str) -> tuple[str, str]:
    """Delegates to helpers.pdf_extract.extract_pdf_text."""
    return _extract_pdf_text_impl(pdf_path)



def _scrape_form_data(soup) -> dict:
    """Pull all input/select default values from the first <form> in soup.
    Returns a flat dict of {name: value} ready for a POST.
    """
    form = soup.find("form")
    if not form:
        return {}
    fd: dict = {}
    for inp in form.find_all("input"):
        nm = inp.get("name")
        if nm:
            fd[nm] = inp.get("value", "")
    for sel in form.find_all("select"):
        nm = sel.get("name")
        if nm:
            opt = sel.find("option", selected=True) or sel.find("option")
            fd[nm] = opt["value"] if opt and opt.get("value") else ""
    return fd


def _research_path(job_number, client_name, job_type) -> Path:
    return _job_base_path(job_number, client_name, job_type) / "E Research" / "research.json"


def load_research(job_number, client_name, job_type) -> dict:
    p = _research_path(job_number, client_name, job_type)
    data = None
    if p.exists():
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            pass
    if data is None:
        # Default skeleton
        data = {
            "job_number":  job_number,
            "client_name": client_name,
            "job_type":    job_type,
            "subjects": [
                {"id": "client", "type": "client", "name": client_name,
                 "deed_saved": False, "plat_saved": False,
                 "status": "pending", "notes": "",
                 "deed_path": "", "plat_path": ""}
            ]
        }
    # Migrate subjects that pre-date new fields (idempotent)
    for s in data.get("subjects", []):
        s.setdefault("status",    "pending")
        s.setdefault("notes",     "")
        s.setdefault("deed_path", "")
        s.setdefault("plat_path", "")
    return data


def save_research(job_number, client_name, job_type, data: dict):
    p = _research_path(job_number, client_name, job_type)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

# ── static ─────────────────────────────────────────────────────────────────────

@app.route("/")
def landing():
    """Public marketing/landing page. Authenticated users are redirected to /app by the JS."""
    resp = make_response(send_from_directory(".", "landing.html"))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp

@app.route("/app")
def index():
    """The main SPA. Served to authenticated users."""
    resp = make_response(send_from_directory(".", "index.html"))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp

@app.route("/app.js")
def serve_appjs():
    # No-cache headers are applied by the @after_request hook below
    return send_from_directory(".", "app.js")

@app.route("/style.css")
def serve_css():
    # No-cache headers are applied by the @after_request hook below
    return send_from_directory(".", "style.css")

@app.route("/favicon.png")
def serve_favicon():
    return send_from_directory(".", "favicon.png")


@app.route("/robots.txt")
def robots_txt():
    """Block search engine crawlers from API, auth, and admin routes."""
    content = (
        "User-agent: *\n"
        "Disallow: /api/\n"
        "Disallow: /auth/\n"
        "Disallow: /admin/\n"
        "Disallow: /api/admin/\n"
        "Disallow: /api/stripe/\n"
        "Allow: /\n"
    )
    return app.response_class(content, mimetype="text/plain")


@app.route("/.well-known/security.txt")
def security_txt():
    """Standard security.txt — tells researchers how to report vulnerabilities."""
    _app_url = os.environ.get("DEED_APP_URL", "https://deedplathelper.netlify.app")
    content = (
        f"Contact: mailto:support@deedplathelper.com\n"
        f"Expires: 2027-01-01T00:00:00.000Z\n"
        f"Preferred-Languages: en\n"
        f"Canonical: {_app_url}/.well-known/security.txt\n"
        f"Policy: Please report security vulnerabilities responsibly via email before public disclosure.\n"
    )
    return app.response_class(content, mimetype="text/plain")

@app.after_request
def add_security_headers(response):
    """Add HTTP security headers and no-cache directives to every response."""
    # ── Cache control ────────────────────────────────────────────────────────
    if "max-age" not in response.headers.get("Cache-Control", ""):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    # ── Security headers ─────────────────────────────────────────────────────
    response.headers["X-Content-Type-Options"] = "nosniff"            # Prevent MIME sniffing
    response.headers["X-Frame-Options"] = "DENY"                      # Clickjacking protection
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    if _is_production:
        response.headers["Strict-Transport-Security"] = (
            "max-age=31536000; includeSubDomains"                      # Force HTTPS for 1 year
        )
    return response

# ── SaaS Auth routes → moved to routes/auth.py Blueprint ─────────────────────



# ── Stripe billing → moved to routes/stripe.py Blueprint ─────────────────────
# ── Admin panel   → moved to routes/admin.py Blueprint ─────────────────────


# Team + backup routes -> moved to routes/team.py Blueprint

# ── profiles ──────────────────────────────────────────────────────

@app.route("/api/profiles", methods=["GET"])
def api_profiles_list():
    """Return all user profiles."""
    profiles = list_profiles()
    # If no profiles exist, auto-migrate from config.json
    if not profiles:
        cfg = load_config()
        p = migrate_from_config(cfg)
        profiles = [p]
    return jsonify({"success": True, "profiles": profiles})


@app.route("/api/profiles", methods=["POST"])
def api_profiles_create():
    """Create a new user profile.  Body: { display_name }"""
    data = request.get_json(silent=True) or {}
    name = data.get("display_name", "").strip()
    if not name:
        return jsonify({"success": False, "error": "display_name is required"})
    p = create_profile(name)
    # Optionally copy shared credentials + portal URL into the new profile
    cfg = load_config()
    if cfg.get("firstnm_user"):
        p["firstnm_user"] = cfg["firstnm_user"]
        p["firstnm_pass"] = cfg.get("firstnm_pass", "")
        p["firstnm_url"]  = cfg.get("firstnm_url", "")
        save_profile(p)
    return jsonify({"success": True, "profile": p})


@app.route("/api/profiles/<profile_id>", methods=["GET"])
def api_profile_get(profile_id):
    p = get_profile(profile_id)
    if p is None:
        return jsonify({"success": False, "error": "Profile not found"}), 404
    return jsonify({"success": True, "profile": p})


@app.route("/api/profiles/<profile_id>", methods=["PUT"])
def api_profile_update(profile_id):
    """Update fields on a profile.  Body: { field: value, ... }"""
    p = get_profile(profile_id)
    if p is None:
        return jsonify({"success": False, "error": "Profile not found"}), 404
    data = request.get_json(silent=True) or {}
    for k, v in data.items():
        if k != "id":  # id is immutable
            p[k] = v
    save_profile(p)
    return jsonify({"success": True, "profile": p})


@app.route("/api/profiles/<profile_id>", methods=["DELETE"])
def api_profile_delete(profile_id):
    ok = delete_profile(profile_id)
    return jsonify({"success": ok})


def _get_request_profile_id() -> str | None:
    """Extract profile_id from request cookie or query param."""
    return request.cookies.get("profile_id") or request.args.get("profile_id")


def _get_request_session() -> req_lib.Session:
    """Return the web session for the current request's profile."""
    return _get_web_session(_get_request_profile_id())

# ── config ─────────────────────────────────────────────────────────────────────

@app.route("/api/config", methods=["GET", "POST"])
def api_config():
    profile_id = _get_request_profile_id()
    profile = get_profile(profile_id) if profile_id else None

    if request.method == "GET":
        cfg = load_config()
        # Prefer profile-level credentials, fall back to server config
        if profile:
            user = profile.get("firstnm_user") or cfg.get("firstnm_user", "")
            pwd  = profile.get("firstnm_pass") or cfg.get("firstnm_pass", "")
            sess = profile.get("last_session") or cfg.get("last_session")
        else:
            user = cfg.get("firstnm_user") or cfg.get("username", "")
            pwd  = cfg.get("firstnm_pass") or cfg.get("password", "")
            sess = cfg.get("last_session")
        # ArcGIS layer — prefer profile, fall back to global config, then default
        arcgis_cfg = _get_arcgis_config()
        return jsonify({
            "success": True,
            "config": {
                "firstnm_user": user,
                "firstnm_pass": pwd,
                "firstnm_url":  cfg.get("firstnm_url", ""),
                "last_session": sess,
                # ArcGIS layer config
                "arcgis_url":    arcgis_cfg["url"],
                "arcgis_fields": arcgis_cfg["fields"],
                "arcgis_is_default": not bool(
                    (profile and profile.get("arcgis_url")) or cfg.get("arcgis_url")
                ),
                # Expose presets so the frontend can offer a selector
                "arcgis_presets": [
                    {"id": k, "label": v["label"], "url": v["url"],
                     "fields": v["fields"]}
                    for k, v in _ARCGIS_PRESETS.items()
                ],
            }
        })
    data = request.get_json()
    # Save user-specific fields to profile if available, otherwise config
    if profile:
        if "firstnm_user" in data or "username" in data:
            profile["firstnm_user"] = data.get("firstnm_user", data.get("username", ""))
        if "firstnm_pass" in data or "password" in data:
            profile["firstnm_pass"] = data.get("firstnm_pass", data.get("password", ""))
        if "last_session" in data:
            profile["last_session"] = data["last_session"]
        if "arcgis_url" in data:
            new_url = (data["arcgis_url"] or "").strip()
            if new_url:  # Only overwrite if user actually supplied a URL
                profile["arcgis_url"] = new_url
        if "arcgis_fields" in data and isinstance(data["arcgis_fields"], dict) and profile.get("arcgis_url"):
            profile["arcgis_fields"] = data["arcgis_fields"]
        if "firstnm_url" in data:
            profile["firstnm_url"] = data["firstnm_url"]
        save_profile(profile)
    else:
        cfg = load_config()
        if "firstnm_user" in data:
            cfg["firstnm_user"] = data["firstnm_user"]
        elif "username" in data:
            cfg["firstnm_user"] = data["username"]
        if "firstnm_pass" in data:
            cfg["firstnm_pass"] = data["firstnm_pass"]
        elif "password" in data:
            cfg["firstnm_pass"] = data["password"]
        if "firstnm_url" in data:
            cfg["firstnm_url"] = data["firstnm_url"]
        if "arcgis_url" in data:
            new_url = (data["arcgis_url"] or "").strip()
            if new_url:  # Only overwrite if user actually supplied a URL
                cfg["arcgis_url"] = new_url
        if "arcgis_fields" in data and isinstance(data["arcgis_fields"], dict) and cfg.get("arcgis_url"):
            cfg["arcgis_fields"] = data["arcgis_fields"]
        if "last_session" in data:
            cfg["last_session"] = data["last_session"]
        cfg.pop("username", None)
        cfg.pop("password", None)
        save_config(cfg)
    # Invalidate ArcGIS address cache so new settings take effect immediately
    _address_cache.clear()
    return jsonify({"success": True})


# ── ArcGIS discover & test ──────────────────────────────────────────────────────

@app.route("/api/arcgis-discover", methods=["POST"])
@require_auth
@require_pro
def api_arcgis_discover():
    """Probe an ArcGIS REST layer URL and return all available field names.

    Body: { "url": "https://...query" }  (or just the base layer URL)
    Returns: { success, fields: [{name, type, alias}], layer_info: {...} }
    """
    try:
        data = request.get_json() or {}
        raw_url = (data.get("url") or "").strip().rstrip("/")
        if not raw_url:
            return jsonify({"success": False, "error": "No URL provided"})

        # Strip /query suffix if present so we hit the layer metadata endpoint
        base_url = raw_url[:-6] if raw_url.lower().endswith("/query") else raw_url

        resp = req_lib.get(
            base_url,
            params={"f": "json"},
            headers={"User-Agent": "DeedPlatHelper/1.0"},
            timeout=15,
        )
        if resp.status_code != 200:
            return jsonify({"success": False, "error": f"HTTP {resp.status_code} from ArcGIS server"})

        info = resp.json()
        if "error" in info:
            msg = info["error"].get("message", str(info["error"]))
            return jsonify({"success": False, "error": f"ArcGIS error: {msg}"})

        raw_fields = info.get("fields", [])
        if not raw_fields:
            return jsonify({"success": False,
                           "error": "No fields found. Make sure the URL points to a specific layer (ending in /0, /1, /29, etc.) not the service root."})

        fields = [
            {
                "name":  f.get("name", ""),
                "alias": f.get("alias") or f.get("name", ""),
                "type":  f.get("type", ""),
            }
            for f in raw_fields
            if f.get("name")
        ]

        return jsonify({
            "success": True,
            "fields":  fields,
            "layer_info": {
                "name":        info.get("name", ""),
                "description": info.get("description", ""),
                "geometry_type": info.get("geometryType", ""),
                "feature_count": info.get("maxRecordCount"),
            },
            "query_url": base_url + "/query",
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/arcgis-test", methods=["POST"])
def api_arcgis_test():
    """Run a sample parcel lookup with the provided ArcGIS config.

    Body: {
        "url":    "https://.../query",
        "fields": { concept: field_name, ... },   // field mapping to test
        "sample_id": "ABC123",                     // optional parcel ID to look up
    }
    Returns: { success, sample_result: {...}, matched_fields: [...] }
    """
    try:
        data = request.get_json() or {}
        url    = (data.get("url") or "").strip().rstrip("/")
        fields = data.get("fields") or {}
        sample_id = (data.get("sample_id") or "").strip()

        if not url:
            return jsonify({"success": False, "error": "No URL provided"})

        # Ensure URL ends with /query
        query_url = url if url.lower().endswith("/query") else url + "/query"

        # Build a test config
        test_cfg = {
            "url":    query_url,
            "fields": {
                **_ARCGIS_PRESETS["taos_nm"]["fields"],  # defaults
                **{k: v for k, v in fields.items() if v},   # user overrides
            },
        }

        pid_field = _arcgis_field(test_cfg, "parcel_id")

        if sample_id:
            # Try to look up the provided parcel ID
            result = _arcgis_lookup_upc(sample_id, arcgis_cfg=test_cfg)
            if result:
                return jsonify({"success": True, "sample_result": result,
                               "tested_with": sample_id})
            return jsonify({"success": False,
                           "error": f"Parcel '{sample_id}' not found using field '{pid_field}'. Try a different sample ID."})

        # No sample ID — fetch the first record from the layer to verify connectivity
        out_fields = _arcgis_out_fields(test_cfg, list(test_cfg["fields"].keys()))
        resp = req_lib.get(
            query_url,
            params={
                "where":             "1=1",
                "outFields":         out_fields,
                "returnGeometry":    "false",
                "resultRecordCount": "1",
                "f":                 "json",
            },
            headers={"User-Agent": "DeedPlatHelper/1.0"},
            timeout=15,
        )
        if resp.status_code != 200:
            return jsonify({"success": False, "error": f"HTTP {resp.status_code} from ArcGIS"})

        resp_data = resp.json()
        if "error" in resp_data:
            return jsonify({"success": False,
                           "error": resp_data["error"].get("message", "ArcGIS query error")})

        features = resp_data.get("features", [])
        if not features:
            return jsonify({"success": False, "error": "Layer returned no features. Check the URL and ensure the layer has data."})

        sample_attrs = features[0].get("attributes", {})
        # Show which configured fields actually have data
        matched_fields = [
            {"concept": c, "field": f, "value": str(sample_attrs.get(f, "(not found)"))[:80]}
            for c, f in test_cfg["fields"].items()
            if f
        ]
        return jsonify({
            "success":        True,
            "sample_attrs":   {k: str(v)[:80] for k, v in sample_attrs.items()},
            "matched_fields": matched_fields,
            "record_count":   resp_data.get("exceededTransferLimit", False),
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})

# County Registry -> moved to routes/admin.py Blueprint

# ── test-connection (onboarding wizard) ────────────────────────────────────────

@app.route("/api/test-connection", methods=["POST"])
def api_test_connection():
    """Test portal login with a caller-supplied URL + credentials.

    Body: { url, username, password }
    Returns: { success, error? }
    """
    try:
        data = request.get_json(silent=True) or {}
        portal_url = (data.get("url") or "").strip().rstrip("/")
        username   = data.get("username", "")
        password   = data.get("password", "")

        if not portal_url:
            return jsonify({"success": False, "error": "No portal URL provided"})

        # Use a temporary session so we don't pollute the main one
        tmp_sess = req_lib.Session()
        tmp_sess.headers.update({"User-Agent": "DeedPlatHelper/1.0"})

        resp = tmp_sess.get(portal_url + "/", timeout=10)
        soup = BeautifulSoup(resp.text, "lxml")
        form = soup.find("form")
        if not form:
            return jsonify({"success": False, "error": "Login form not found at that URL"})

        action = form.get("action", "/")
        if not action.startswith("http"):
            action = portal_url + "/" + action.lstrip("/")

        form_data = {}
        for inp in form.find_all("input"):
            nm = inp.get("name")
            itype = (inp.get("type") or "text").lower()
            if nm:
                if itype == "image":
                    form_data[nm + ".x"] = "1"
                    form_data[nm + ".y"] = "1"
                else:
                    form_data[nm] = inp.get("value", "")

        # Map credentials
        mapped = False
        for inp in form.find_all("input"):
            nm = inp.get("name", "")
            if nm == "FormUser":
                form_data["FormUser"] = username; mapped = True
            elif nm == "FormPassword":
                form_data["FormPassword"] = password; mapped = True
        if not mapped:
            for inp in form.find_all("input"):
                itype = (inp.get("type") or "text").lower()
                iname = (inp.get("name") or "").lower()
                if itype == "password":
                    form_data[inp["name"]] = password
                elif itype == "text" and ("user" in iname or "login" in iname):
                    form_data[inp["name"]] = username

        post_resp = tmp_sess.post(action, data=form_data, timeout=10)
        landed_url = post_resp.url.lower()
        portal_root = portal_url.lower().rstrip("/")
        success = ("hfweb" in landed_url
                   or "new search" in post_resp.text.lower()
                   or ("logout" in post_resp.text.lower()
                       and landed_url.rstrip("/") != portal_root))

        if success:
            return jsonify({"success": True})
        else:
            return jsonify({"success": False, "error": "Invalid credentials or login failed"})
    except req_lib.exceptions.Timeout:
        return jsonify({"success": False, "error": "Connection timed out — check the URL"})
    except req_lib.exceptions.ConnectionError:
        return jsonify({"success": False, "error": "Cannot reach that URL — check the address"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

# ── login ──────────────────────────────────────────────────────

@app.route("/api/login", methods=["POST"])
def api_login():
    try:
        data = request.get_json()
        username = data.get("username", "")
        password = data.get("password", "")
        remember = data.get("remember", False)

        # Fetch login page to discover form
        sess = _session()
        resp = sess.get(_get_portal_url() + "/", timeout=8)
        soup = BeautifulSoup(resp.text, "lxml")
        form = soup.find("form")
        if not form:
            return jsonify({"success": False, "error": "Login form not found"})

        action = form.get("action", "/")
        if not action.startswith("http"):
            action = _get_portal_url() + "/" + action.lstrip("/")

        form_data = {}
        for inp in form.find_all('input'):
            nm = inp.get('name')
            itype = (inp.get('type') or 'text').lower()
            if nm:
                if itype == 'image':
                    # Image-type submit buttons need .x/.y coords
                    form_data[nm + '.x'] = '1'
                    form_data[nm + '.y'] = '1'
                else:
                    form_data[nm] = inp.get('value', '')

        # Map credentials using known e.halFILE field names first,
        # then fall back to heuristic detection
        mapped = False
        for inp in form.find_all('input'):
            nm  = inp.get('name', '')
            itype = (inp.get('type') or 'text').lower()
            if nm == 'FormUser':
                form_data['FormUser'] = username
                mapped = True
            elif nm == 'FormPassword':
                form_data['FormPassword'] = password
                mapped = True
        if not mapped:
            for inp in form.find_all('input'):
                itype = (inp.get('type') or 'text').lower()
                iname = (inp.get('name') or '').lower()
                if itype == 'password':
                    form_data[inp['name']] = password
                elif itype == 'text' and ('user' in iname or 'login' in iname):
                    form_data[inp['name']] = username

        post_resp = sess.post(action, data=form_data, timeout=8)
        # Success = we landed on the search/welcome page, NOT back on login
        portal_root = _get_portal_url().lower().rstrip('/')
        landed_url = post_resp.url.lower()
        success = ('hfweb' in landed_url or 'new search' in post_resp.text.lower() or
                   ('logout' in post_resp.text.lower() and landed_url.rstrip('/') != portal_root))

        if success:
            if remember:
                cfg = load_config()
                cfg["firstnm_user"] = username
                cfg["firstnm_pass"] = password
                cfg.pop("username", None)  # clean up legacy keys
                cfg.pop("password", None)
                save_config(cfg)
            return jsonify({"success": True, "username": username})
        else:
            return jsonify({"success": False, "error": "Invalid credentials or login failed"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

# ── logout ─────────────────────────────────────────────────────────────────────

@app.route("/api/logout", methods=["POST"])
def api_logout():
    _session().cookies.clear()
    return jsonify({"success": True})

# extract_trs — imported from helpers.metes_bounds


# ── search ─────────────────────────────────────────────────────────────────────

@app.route("/api/search", methods=["POST"])
@require_auth
@rate_limit(requests=30, window=60)   # hard IP throttle: 30 req/min (quota gate is the main limit)
def api_search():
    # ── Quota gate ──────────────────────────────────────────────────────────
    allowed, quota_msg = check_search_quota(g.current_user)
    if not allowed:
        return jsonify({
            "success":        False,
            "error":          quota_msg,
            "upgrade_required": True,
            "current_tier":   g.current_user.get("tier", "free"),
        }), 403
    try:
        data = request.get_json()
        name      = data.get("name", "").strip()
        address   = data.get("address", "").strip()
        name_type = data.get("name_type", "grantor")  # "grantor" | "grantee"
        # Map UI operator labels to site's actual option values
        op_map = {"contains": "contains", "begins with": "begin", "exact match": "exact", "equals": "exact"}
        operator = op_map.get(data.get("operator", "contains"), "contains")

        search_url = f"{_get_portal_url()}/scripts/hfweb.asp?Application=FNM&Database=TP"
        sess = _session()
        resp = sess.get(search_url, timeout=15)

        # Detect redirect back to login page
        landed = resp.url.lower().rstrip('/')
        if landed == _get_portal_url().lower().rstrip('/') or 'login' in landed:
            return jsonify({"success": False, "error": "Session expired — please log in again."})

        # The site has malformed HTML (form appears after </html>),
        # so use html.parser which tolerates this. Also detect auth via raw text.
        if 'CROSSNAMEFIELD' not in resp.text and 'FIELD14' not in resp.text:
            return jsonify({"success": False, "error": "Session expired — please log in again."})

        soup = BeautifulSoup(resp.text, "html.parser")

        # Action is set via JS; we know it's always hflook.asp
        action = _get_portal_url() + "/scripts/hflook.asp"


        form_data = _scrape_form_data(soup)
        if not form_data:
            return jsonify({"success": False, "error": "Search form not found"})

        # Apply search criteria
        if name:
            if name_type == "grantee":
                # Grantee index: FIELD20 is typically the grantee cross-reference field
                form_data["CROSSNAMEFIELD"] = name
                form_data["CROSSNAMETYPE"]  = operator
                form_data["CROSSTYPE"]       = "GE"   # GE = grantee, GR = grantor
            else:
                form_data["CROSSNAMEFIELD"] = name
                form_data["CROSSNAMETYPE"]  = operator
                form_data["CROSSTYPE"]       = "GR"
        if address:
            form_data["FIELD19"] = address
            form_data["SEARCHTYPE19"] = operator

        post_resp = sess.post(action, data=form_data, timeout=20)
        soup2 = BeautifulSoup(post_resp.text, "html.parser")

        # Parse results table
        results = []
        # Find count
        count_text = ""
        for tag in soup2.find_all(string=re.compile(r'\d+ records? found', re.I)):
            count_text = tag.strip()
            break

        rows = soup2.find_all("tr")
        for row in rows:
            cells = row.find_all("td")
            if len(cells) < 9:
                continue
            doc_link = cells[1].find("a") if len(cells) > 1 else None
            if not doc_link:
                continue
            doc_no = doc_link.text.strip()
            if not doc_no or not re.match(r'^[A-Z0-9]+$', doc_no):
                continue
            results.append({
                "doc_no":           doc_no,
                "location":         cells[2].text.strip() if len(cells) > 2 else "",
                "document_code":    cells[3].text.strip() if len(cells) > 3 else "",
                "gf_number":        cells[4].text.strip() if len(cells) > 4 else "",
                "instrument_type":  cells[5].text.strip() if len(cells) > 5 else "",
                "document_no":      cells[6].text.strip() if len(cells) > 6 else "",
                "recorded_date":    cells[7].text.strip() if len(cells) > 7 else "",
                "instrument_date":  cells[8].text.strip() if len(cells) > 8 else "",
                "grantor":          cells[9].text.strip() if len(cells) > 9 else "",
                "grantee":          cells[10].text.strip() if len(cells) > 10 else "",
            })

        # Increment monthly search counter for tier tracking + save to history
        increment_search_count(g.current_user)
        _query = name or address
        if _query:
            add_search_history(g.current_user["id"], _query, len(results))
        return jsonify({"success": True, "results": results, "count": len(results), "count_text": count_text})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ── relevance scoring ─────────────────────────────────────────────────────────

# Instrument types most relevant to boundary survey research
_DEED_TYPES = {'deed', 'warranty', 'quitclaim', 'grant', 'conveyance', 'bargain'}
_PLAT_TYPES = {'plat', 'survey', 'partition', 'subdivision'}
_LOW_TYPES  = {'mortgage', 'lien', 'release', 'assignment', 'satisfaction', 'ucc'}


def _score_search_result(
    result: dict,
    client_trs: str = "",
    client_name: str = "",
    adjoiner_names: list = None,
    client_subdivision: str = "",
) -> dict:
    """Score a single 1stNMTitle search result for relevance to the client property.

    Modifies the result dict in-place, adding:
      - relevance_score (int 0-100)
      - relevance_tags  (list of tag strings)

    Returns the modified result.
    """
    score = 0
    tags = []
    adjoiner_set = {n.upper().split(",")[0].strip() for n in (adjoiner_names or []) if n}
    client_last = client_name.split(",")[0].strip().upper() if client_name else ""

    inst_type = (result.get("instrument_type") or "").lower()
    grantor   = (result.get("grantor") or "").upper()
    grantee   = (result.get("grantee") or "").upper()
    location  = (result.get("location") or "").upper()

    # ── TRS match (highest value — same section as client) ─────────────
    if client_trs:
        trs_up = client_trs.upper()
        # Extract TRS components from location field
        # Location often contains book-page like "M568-482" but sometimes has section refs
        # Also check if the deed's grantor/grantee names match KML parcels in same TRS
        # (this is a heuristic — full TRS matching happens via the KML index)
        # For now, check if the location field directly mentions the same section
        trs_parts = re.findall(r'SEC(?:TION)?\s*(\d+)', location, re.I)
        client_sec = re.search(r'SEC\s*(\d+)', trs_up)
        if client_sec and trs_parts:
            if client_sec.group(1) in trs_parts:
                score += 40
                tags.append("trs_match")

    # ── Subdivision match ─────────────────────────────────────────────
    if client_subdivision:
        subdiv_up = client_subdivision.upper()
        if subdiv_up in location or subdiv_up in grantor or subdiv_up in grantee:
            score += 25
            tags.append("same_subdivision")

    # ── Name match (client or adjoiner) ───────────────────────────────
    if client_last and len(client_last) >= 2:
        if client_last in grantor or client_last in grantee:
            score += 30
            tags.append("client_name")

    if adjoiner_set:
        for adj_last in adjoiner_set:
            if adj_last and len(adj_last) >= 3:
                if adj_last in grantor or adj_last in grantee:
                    score += 20
                    tags.append("adjoiner")
                    break

    # ── Instrument type priority ──────────────────────────────────────
    inst_words = set(inst_type.split())
    if _DEED_TYPES & inst_words:
        score += 15
        tags.append("deed")
    elif _PLAT_TYPES & inst_words:
        score += 12
        tags.append("plat")
    elif _LOW_TYPES & inst_words:
        score -= 5  # deprioritize mortgages/liens

    # ── Recency bonus ─────────────────────────────────────────────────
    rec_date = result.get("recorded_date") or result.get("instrument_date") or ""
    if rec_date:
        try:
            year = int(rec_date.split("-")[0]) if "-" in rec_date else int(rec_date[-4:])
            if year >= 2015:
                score += 5
            elif year >= 2000:
                score += 3
        except (ValueError, IndexError):
            pass

    result["relevance_score"] = max(0, min(100, score))
    result["relevance_tags"]  = tags
    return result


@app.route("/api/search-enriched", methods=["POST"])
@require_auth
@require_pro
def api_search_enriched():
    """Enriched search: standard 1stNMTitle search + ArcGIS relevance scoring.

    Body: {
        name, operator, name_type, address,   (standard search params)
        client_upc,                           (triggers ArcGIS enrichment)
        client_name, adjoiner_names,          (for name matching)
        sort_by                               (relevance|date|type|original)
    }
    """
    try:
        data = request.get_json() or {}
        client_upc  = (data.get("client_upc") or "").strip()
        client_name = (data.get("client_name") or "").strip()
        adj_names   = data.get("adjoiner_names") or []
        sort_by     = data.get("sort_by", "relevance")

        # Step 1: Perform the standard search by delegating to api_search internals
        # (We replicate the call rather than HTTP-fetching ourselves)
        name      = (data.get("name") or "").strip()
        address   = (data.get("address") or "").strip()
        name_type = data.get("name_type", "grantor")
        op_map = {"contains": "contains", "begins with": "begin",
                  "exact match": "exact", "equals": "exact"}
        operator = op_map.get(data.get("operator", "contains"), "contains")

        if not name and not address:
            return jsonify({"success": False, "error": "No search criteria provided"})

        sess = _session()
        search_url = f"{_get_portal_url()}/scripts/hfweb.asp?Application=FNM&Database=TP"
        resp = sess.get(search_url, timeout=15)

        landed = resp.url.lower().rstrip('/')
        if landed == _get_portal_url().lower().rstrip('/') or 'login' in landed:
            return jsonify({"success": False, "error": "Session expired — please log in again."})

        if 'CROSSNAMEFIELD' not in resp.text and 'FIELD14' not in resp.text:
            return jsonify({"success": False, "error": "Session expired — please log in again."})

        soup = BeautifulSoup(resp.text, "html.parser")
        action = _get_portal_url() + "/scripts/hflook.asp"
        form_data = _scrape_form_data(soup)
        if not form_data:
            return jsonify({"success": False, "error": "Search form not found"})

        if name:
            form_data["CROSSNAMEFIELD"] = name
            form_data["CROSSNAMETYPE"]  = operator
            form_data["CROSSTYPE"]      = "GE" if name_type == "grantee" else "GR"
        if address:
            form_data["FIELD19"] = address
            form_data["SEARCHTYPE19"] = operator

        post_resp = sess.post(action, data=form_data, timeout=20)
        soup2 = BeautifulSoup(post_resp.text, "html.parser")

        # Parse results
        results = []
        count_text = ""
        for tag in soup2.find_all(string=re.compile(r'\d+ records? found', re.I)):
            count_text = tag.strip()
            break

        rows = soup2.find_all("tr")
        for row in rows:
            cells = row.find_all("td")
            if len(cells) < 9:
                continue
            doc_link = cells[1].find("a") if len(cells) > 1 else None
            if not doc_link:
                continue
            doc_no = doc_link.text.strip()
            if not doc_no or not re.match(r'^[A-Z0-9]+$', doc_no):
                continue
            results.append({
                "doc_no":           doc_no,
                "location":         cells[2].text.strip() if len(cells) > 2 else "",
                "document_code":    cells[3].text.strip() if len(cells) > 3 else "",
                "gf_number":        cells[4].text.strip() if len(cells) > 4 else "",
                "instrument_type":  cells[5].text.strip() if len(cells) > 5 else "",
                "document_no":      cells[6].text.strip() if len(cells) > 6 else "",
                "recorded_date":    cells[7].text.strip() if len(cells) > 7 else "",
                "instrument_date":  cells[8].text.strip() if len(cells) > 8 else "",
                "grantor":          cells[9].text.strip() if len(cells) > 9 else "",
                "grantee":          cells[10].text.strip() if len(cells) > 10 else "",
            })

        # Step 2: Get client's ArcGIS context for scoring
        client_trs = ""
        client_subdivision = ""
        if client_upc:
            arc = _arcgis_lookup_upc(client_upc)
            if arc and arc.get("success"):
                client_trs = arc.get("trs", "")
                client_subdivision = arc.get("subdivision", "")

        # Also try from KML index if no ArcGIS TRS
        if not client_trs and client_upc:
            survey = get_survey_data_path()
            idx = xml_processor.load_index(survey)
            if idx:
                for p in idx.get("parcels", []):
                    if p.get("upc") == client_upc:
                        client_trs = p.get("trs", "") or (
                            p.get("arcgis", {}).get("trs", "") if p.get("arcgis") else ""
                        )
                        client_subdivision = (
                            p.get("arcgis", {}).get("subdivision", "") if p.get("arcgis") else ""
                        )
                        break

        # Step 3: Score each result
        for r in results:
            _score_search_result(
                r,
                client_trs=client_trs,
                client_name=client_name,
                adjoiner_names=adj_names,
                client_subdivision=client_subdivision,
            )

        # Step 4: Sort by requested order
        if sort_by == "relevance":
            results.sort(key=lambda r: r.get("relevance_score", 0), reverse=True)
        elif sort_by == "date":
            results.sort(
                key=lambda r: r.get("recorded_date") or r.get("instrument_date") or "",
                reverse=True
            )
        elif sort_by == "type":
            type_order = {"deed": 0, "plat": 1, "survey": 2}
            results.sort(key=lambda r: type_order.get(
                (r.get("instrument_type") or "").lower().split()[0] if r.get("instrument_type") else "",
                99
            ))

        return jsonify({
            "success":    True,
            "results":    results,
            "count":      len(results),
            "count_text": count_text,
            "sort_by":    sort_by,
            "client_trs":        client_trs,
            "client_subdivision": client_subdivision,
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ── chain-of-title search ─────────────────────────────────────────────────────

@app.route("/api/chain-search", methods=["POST"])
@require_auth
@require_pro
def api_chain_search():
    """
    Trace ownership backward by recursively searching the grantor as grantee.
    Returns a chain of {doc_no, grantor, grantee, location, date, has_plat_ref}.
    Stop when: no results, cycle detected, plat reference found, or max_hops reached.
    """
    try:
        data = request.get_json()
        start_grantor = data.get("start_grantor", "").strip()
        max_hops = min(int(data.get("max_hops", 10)), 20)  # cap at 20

        if not start_grantor:
            return jsonify({"success": False, "error": "No starting grantor provided"})

        chain = []
        seen_docs = set()
        current_name = start_grantor
        stop_reason = ""

        # Plat reference patterns
        plat_re = re.compile(r'(?:plat|cabinet|cab\.?|survey|plat\s+book)', re.I)

        for hop in range(max_hops):
            # Search for current_name as GRANTEE (find deed where they received property)
            search_url = f"{_get_portal_url()}/scripts/hfweb.asp?Application=FNM&Database=TP"
            try:
                resp = _session().get(search_url, timeout=15)
            except Exception:
                stop_reason = "Network error during search"
                break

            if 'CROSSNAMEFIELD' not in resp.text:
                stop_reason = "Session expired — could not continue chain search"
                break


            soup = BeautifulSoup(resp.text, "html.parser")
            form_data = _scrape_form_data(soup)
            if not form_data:
                stop_reason = "Search form not found"
                break

            # Search as grantee
            form_data["CROSSNAMEFIELD"] = current_name
            form_data["CROSSNAMETYPE"] = "begin"
            form_data["CROSSTYPE"] = "GE"  # GE = grantee

            action = _get_portal_url() + "/scripts/hflook.asp"
            try:
                post_resp = _session().post(action, data=form_data, timeout=20)
            except Exception:
                stop_reason = f"Network error searching for {current_name}"
                break

            soup2 = BeautifulSoup(post_resp.text, "html.parser")
            rows = soup2.find_all("tr")

            # Find deed results (prioritize warranty deeds)
            candidates = []
            for row in rows:
                cells = row.find_all("td")
                if len(cells) < 9:
                    continue
                doc_link = cells[1].find("a") if len(cells) > 1 else None
                if not doc_link:
                    continue
                doc_no = doc_link.text.strip()
                if not doc_no or not re.match(r'^[A-Z0-9]+$', doc_no):
                    continue
                if doc_no in seen_docs:
                    continue

                inst_type = cells[5].text.strip() if len(cells) > 5 else ""
                grantor = cells[9].text.strip() if len(cells) > 9 else ""
                grantee = cells[10].text.strip() if len(cells) > 10 else ""
                location = cells[2].text.strip() if len(cells) > 2 else ""
                rec_date = cells[7].text.strip() if len(cells) > 7 else ""

                # Prefer deeds over mortgages/liens
                is_deed = any(kw in inst_type.lower() for kw in ['deed', 'warranty', 'quitclaim', 'grant', 'convey'])
                candidates.append({
                    "doc_no": doc_no,
                    "grantor": grantor,
                    "grantee": grantee,
                    "location": location,
                    "date": rec_date,
                    "instrument_type": inst_type,
                    "is_deed": is_deed,
                    "has_plat_ref": bool(plat_re.search(location + " " + grantor)),
                })

            if not candidates:
                stop_reason = f"No prior deeds found for {current_name}"
                break

            # Sort: deeds first, then by date descending (most recent first)
            candidates.sort(key=lambda c: (not c["is_deed"], c["date"]), reverse=False)
            best = candidates[0]

            seen_docs.add(best["doc_no"])
            chain.append(best)

            if best["has_plat_ref"]:
                stop_reason = f"Plat reference found in deed {best['doc_no']}"
                break

            # Continue chain with the grantor of this deed
            next_name = best["grantor"]
            if not next_name or next_name.lower() == current_name.lower():
                stop_reason = f"Grantor same as previous — chain ends at {best['doc_no']}"
                break

            current_name = next_name

        if not stop_reason and len(chain) >= max_hops:
            stop_reason = f"Reached maximum depth ({max_hops} hops)"

        return jsonify({
            "success": True,
            "chain": chain,
            "hops": len(chain),
            "stop_reason": stop_reason,
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ── document detail ────────────────────────────────────────────────────────────

@app.route("/api/document/<doc_no>", methods=["GET", "POST"])
@require_auth
def api_document(doc_no):
    try:
        cfg = load_config()
        username = request.args.get("username") or cfg.get("firstnm_user", "")

        # Accept an optional search_result passthrough from the frontend
        # so we can merge rich search row data with scrape results.
        search_row = {}
        if request.method == "POST":
            body = request.get_json(silent=True) or {}
            search_row = body.get("search_result", {})

        url = f"{_get_portal_url()}/scripts/hfpage.asp?Appl=FNM&Doctype=TP&DocNo={doc_no}&FormUser={username}"
        resp = _session().get(url, timeout=15)
        soup = BeautifulSoup(resp.text, "lxml")

        detail = {"doc_no": doc_no}

        # ── Strategy 1: 2-column <td> tables (standard e.halFILE layout) ─────
        tables = soup.find_all("table")
        for table in tables:
            for row in table.find_all("tr"):
                cells = row.find_all("td")
                if len(cells) == 2:
                    label = cells[0].text.strip().rstrip(":")
                    value = cells[1].text.strip()
                    if label and value:
                        detail[label] = value

        # ── Strategy 2: <b> or <strong> label followed by text ───────────────
        # Some pages use patterns like: <b>Grantor:</b> SMITH, JOHN
        if len(detail) <= 1:   # only doc_no found so far
            for b in soup.find_all(["b", "strong"]):
                label = b.text.strip().rstrip(":")
                if not label or len(label) > 40:
                    continue
                # Get the next sibling text
                nxt = b.next_sibling
                if nxt and isinstance(nxt, str) and nxt.strip():
                    detail[label] = nxt.strip()

        # ── Strategy 3: pull all visible text as a raw dump for TRS ──────────
        page_text = soup.get_text(" ", strip=True)

        # ── Merge search_row data (always trust it for known fields) ──────────
        field_map = {
            "location":        "Location",
            "grantor":         "Grantor",
            "grantee":         "Grantee",
            "instrument_type": "Instrument Type",
            "recorded_date":   "Recorded Date",
            "instrument_date": "Instrument Date",
            "document_no":     "Document Number",
            "gf_number":       "GF Number",
            "document_code":   "Document Code",
        }
        for sr_key, detail_key in field_map.items():
            if search_row.get(sr_key) and not detail.get(detail_key):
                detail[detail_key] = search_row[sr_key]

        # ── PDF URL ───────────────────────────────────────────────────────────
        pdf_link = soup.find("a", string=re.compile(r"pdf all pages", re.I))
        if pdf_link:
            href = pdf_link.get("href", "")
            detail["pdf_url"] = (_get_portal_url() + "/" + href.lstrip("/")
                                 if not href.startswith("http") else href)
        else:
            detail["pdf_url"] = f"{_get_portal_url()}/WebTemp/{doc_no}.pdf"

        # ── TRS extraction ────────────────────────────────────────────────────
        all_text = page_text + " " + " ".join(str(v) for v in detail.values())
        trs_refs = extract_trs(all_text)
        if trs_refs:
            detail["_trs"] = trs_refs

        return jsonify({"success": True, "detail": detail})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})

# ── find adjoiners ─────────────────────────────────────────────────────────────

# _ADJ_PATTERNS, _NOISE_WORDS, parse_adjoiner_names — imported from helpers.adjoiner

def parse_adjoiner_names(detail: dict) -> list[dict]:
    """Delegates to helpers.adjoiner.parse_adjoiner_names."""
    return _parse_adjoiner_names_impl(detail)



def find_adjoiners_online(location: str, grantor: str, sess: req_lib.Session | None = None) -> list[dict]:
    """
    Search online records for deeds in the same location book/page range.
    Location format: M568-482 → book 568, search 5 pages around 482.
    Returns list of {name, doc_no, instrument_type, location, source}.
    """
    results = []
    if not location:
        return results

    # Parse book & page: handles "M568-482", "568-482", "L1053-BC" etc.
    m = re.match(r'^[A-Za-z]?(\d+)-', location.strip())
    if not m:
        return results
    book = m.group(1)

    # Search online by location prefix (book number)
    try:
        search_url = f"{_get_portal_url()}/scripts/hfweb.asp?Application=FNM&Database=TP"
        _s = sess or _session()
        resp = _s.get(search_url, timeout=12)
        if "FIELD14" not in resp.text and "CROSSNAMEFIELD" not in resp.text:
            return results  # not logged in


        soup = BeautifulSoup(resp.text, "html.parser")
        fd = _scrape_form_data(soup)
        if not fd:
            return results

        # Search by location book prefix
        # FIELD14 is typically the "Location" search field on this site
        fd["FIELD14"]    = book + "-"
        fd["FIELD14TYPE"] = "begin"

        post = _s.post(f"{_get_portal_url()}/scripts/hflook.asp", data=fd, timeout=20)
        soup2 = BeautifulSoup(post.text, "html.parser")

        # Collect unique grantors/grantees from nearby records (skip our own grantor)
        own_last = grantor.split(",")[0].strip().upper()
        seen_names = set()
        if own_last:
            seen_names.add(own_last)

        for row in soup2.find_all("tr"):
            cells = row.find_all("td")
            if len(cells) < 10:
                continue
            doc_link = cells[1].find("a") if len(cells) > 1 else None
            if not doc_link:
                continue
            itype   = cells[5].text.strip()  if len(cells) > 5  else ""
            # Only deeds / conveyances
            if not re.search(r'deed|warranty|quitclaim|convey|grant', itype, re.I):
                continue
            g_name  = cells[9].text.strip()   if len(cells) > 9  else ""
            ee_name = cells[10].text.strip()  if len(cells) > 10 else ""
            for name in [g_name, ee_name]:
                if not name:
                    continue
                last = name.split(",")[0].strip().upper()
                if last in seen_names or len(last) < 2:
                    continue
                seen_names.add(last)
                display = name.title()
                results.append({
                    "name":            display,
                    "doc_no":          doc_link.text.strip(),
                    "instrument_type": itype,
                    "location":        cells[2].text.strip() if len(cells) > 2 else "",
                    "source":          "online_range",
                })
                if len(results) >= 8:
                    break
            if len(results) >= 8:
                break
    except Exception:
        pass  # online search is best-effort

    return results

# _ocr_cache_path — imported from helpers.pdf_extract

def ocr_plat_file(pdf_path: str) -> list[str]:
    """Delegates to helpers.pdf_extract.ocr_plat_file."""
    return _ocr_plat_file_impl(pdf_path)



@app.route("/api/find-adjoiners", methods=["POST"])
@require_auth
@require_pro
def api_find_adjoiners():
    """
    Automatically discover adjoiner names from:
      1. OCR the local plat file referenced in the deed (primary — plat maps show adjoiners)
      2. Online records search for nearby location/book entries (fallback)
    """
    try:
        data   = request.get_json()
        detail = data.get("detail", {})
        # Guard: detail may arrive as a JSON string or some non-dict — normalize to dict
        if isinstance(detail, str):
            try:
                detail = json.loads(detail)
            except Exception:
                detail = {}
        if not isinstance(detail, dict):
            detail = {}
        # Extract key fields from detail when not sent as top-level keys
        grantor   = data.get("grantor", "") or detail.get("Grantor", "") or detail.get("grantor", "")
        location  = data.get("location", "") or detail.get("Location", "") or detail.get("location", "")
        doc_no    = data.get("doc_no", "")   or detail.get("doc_no", "")
        deed_path = data.get("deed_path", "")  # saved deed PDF path from session
        client_upc = data.get("client_upc", "").strip()  # from map picker selection

        results  = []
        seen_names: set = set()
        plat_file_used = None
        raw_ocr_text = ""

        # ── Strategy 1: scan deed legal-description text for 'Lands of' names ──
        legal_adj = parse_adjoiner_names(detail)
        for item in legal_adj:
            key = item["name"].lower()
            if key not in seen_names:
                seen_names.add(key)
                results.append(item)

        # ── Strategy 2: OCR the local plat referenced by this deed ─────────────
        #   Use ocr_plat_file() which has a disk cache — avoids re-rendering pages
        #   on repeat calls. The raw text is stored in the cache file alongside.
        cab_refs = parse_cabinet_refs(detail)
        print(f"[adjoiners] cabinet refs found: {cab_refs}", flush=True)
        for ref in cab_refs:
            hits = search_local_cabinet(ref["cabinet"], ref["doc"])
            if hits:
                plat_path = hits[0]["path"]
                plat_file_used = hits[0]["file"]
                # ocr_plat_file() handles caching — only renders pages once
                ocr_names = ocr_plat_file(plat_path)
                # Capture raw OCR text from cache file for the response
                try:
                    cache = _ocr_cache_path(plat_path)
                    if cache.exists():
                        cached = json.loads(cache.read_text(encoding="utf-8"))
                        raw_ocr_text = " ".join(cached.get("names", []))
                except Exception:
                    pass
                for name in ocr_names:
                    key = name.lower()
                    if key not in seen_names:
                        seen_names.add(key)
                        results.append({
                            "name":   name,
                            "raw":    "",
                            "field":  plat_file_used,
                            "source": "plat_ocr",
                            "plat":   plat_file_used,
                        })
                break  # only OCR the first matched plat

        # ── Strategy 2b: extract text from saved deed PDF (text layer first, OCR fallback) ──
        if deed_path and os.path.isfile(deed_path):
            print(f"[adjoiners] reading deed text from: {deed_path}", flush=True)
            try:
                deed_text, deed_src = _extract_pdf_text(deed_path)
                print(f"[adjoiners] deed text source={deed_src} chars={len(deed_text.strip())}", flush=True)
                print(f"[adjoiners] deed text sample: {deed_text[:300]!r}", flush=True)
                # Parse for adjoiner names
                deed_detail_fake = {"Legal": deed_text}
                for item in parse_adjoiner_names(deed_detail_fake):
                    key = item["name"].lower()
                    if key not in seen_names:
                        seen_names.add(key)
                        item["source"] = "deed_text"
                        results.append(item)
            except Exception as e2:
                print(f"[adjoiners] deed text error: {e2}", flush=True)


        # ── Strategy 3: KML/XML parcel index — neighboring parcels ─────────────
        #   Uses geometry-based adjacency (polygon edge proximity) as primary,
        #   with centroid-distance fallbacks.
        MAX_KML_PER_SUBSTRATEGY = 8  # cap each sub-strategy
        kml_upc_count = 0
        kml_geom_count = 0
        kml_prox_count = 0
        # Filter out non-person owner names (roads, easements, government, numeric-only)
        _SKIP_OWNER_PATS = re.compile(
            r'^(?:\d+|upc\s*\d|road|street|highway|hwy|county|state|'
            r'new\s*mexico|nm\s*dot|pueblo|blm|usfs|forest\s*service|'
            r'right.?of.?way|easement|vacant|unknown|none)$',
            re.I
        )
        def _clean_owner_name(name: str) -> str:
            """Strip trailing conjunctions (&, Or, And) from owner names."""
            if not name:
                return ''
            n = name.strip()
            for _ in range(3):
                n = re.sub(r'[,\s]+(?:&|AND|OR)\s*$', '', n, flags=re.I).strip()
            n = re.sub(r',\s*$', '', n).strip()
            return n

        def _is_valid_owner(name: str) -> bool:
            """Return False for garbage / non-person owner entries."""
            if not name or len(name.strip()) < 3:
                return False
            clean = name.strip()
            # All digits or mostly digits = bad (e.g. "9617", "20870")
            if re.fullmatch(r'[\d\s\-]+', clean):
                return False
            # Digits followed by a word = UPC concat garbage (e.g. "1073172Road")
            if re.fullmatch(r'\d{4,}[A-Za-z]+', clean):
                return False
            # Starts with UPC-like prefix
            if re.match(r'^upc\s*\d', clean, re.I):
                return False
            if _SKIP_OWNER_PATS.search(clean):
                return False
            return True
        try:
            survey_path = get_survey_data_path()
            kml_idx = xml_processor.load_index(survey_path)
            if kml_idx:
                parcels = kml_idx.get("parcels", [])

                # Step A: find the client parcel — prefer client_upc (from map
                #   picker), fall back to grantor name + book/page heuristic.
                client_parcel = None

                # Priority 1: exact UPC match from map selection
                if client_upc:
                    for p in parcels:
                        if p.get("upc", "") == client_upc:
                            client_parcel = p
                            print(f"[adjoiners][kml] client parcel (UPC match): {p.get('owner')} UPC={client_upc} centroid={p.get('centroid')}", flush=True)
                            break

                # Priority 2: grantor name + book/page fallback
                if not client_parcel:
                    grantor_last  = grantor.split(",")[0].strip().upper() if grantor else ""
                    loc_m = re.match(r'^[A-Za-z]?(\d+)-(\d+)', location.strip()) if location else None
                    book_num = loc_m.group(1) if loc_m else ""
                    page_num = loc_m.group(2) if loc_m else ""

                    for p in parcels:
                        p_owner = p.get("owner", "").upper()
                        owner_match = grantor_last and grantor_last in p_owner
                        book_match  = book_num and p.get("book", "") == book_num and p.get("page", "") == page_num
                        if owner_match or book_match:
                            client_parcel = p
                            print(f"[adjoiners][kml] client parcel (name/book fallback): {p.get('owner')} UPC={p.get('upc')} centroid={p.get('centroid')}", flush=True)
                            break

                if client_parcel:
                    client_upc     = client_parcel.get("upc", "")
                    client_centroid = client_parcel.get("centroid")  # [lng, lat]

                    # Step B: Geometry-based adjacency (most accurate)
                    #   Uses polygon edge proximity — finds parcels that actually share
                    #   a boundary or are within ~33m of the client parcel's edges.
                    if client_upc:
                        try:
                            adj_parcels = xml_processor.find_adjacent_parcels(
                                survey_path, client_upc,
                                max_results=MAX_KML_PER_SUBSTRATEGY,
                                edge_threshold_deg=0.0001  # ~11m — actual shared boundaries only
                            )
                            for p in adj_parcels:
                                if kml_geom_count >= MAX_KML_PER_SUBSTRATEGY:
                                    break
                                name = _clean_owner_name(p.get("owner", "").title())
                                if not _is_valid_owner(name):
                                    continue
                                key = name.lower()
                                if key not in seen_names:
                                    seen_names.add(key)
                                    adj_type = p.get("_adjacency_type", "edge")
                                    results.append({
                                        "name":   name,
                                        "raw":    f"edge dist: {p.get('_adjacency_dist', 0):.5f}°",
                                        "field":  f"KML {adj_type} adjacency",
                                        "source": "kml_geometry",
                                        "upc":    p.get("upc", ""),
                                        "plat":   p.get("plat", ""),
                                    })
                                    kml_geom_count += 1
                            print(f"[adjoiners][kml] geometry adjacency: {kml_geom_count} found", flush=True)
                        except Exception as geom_err:
                            print(f"[adjoiners][kml] geometry adjacency error: {geom_err}", flush=True)

                    # Step C: UPC-prefix neighbors — DISABLED
                    # Sequential UPC numbers don't reliably indicate physical adjacency.
                    # Geometry adjacency (Step B) is far more accurate.

                    # Step D: centroid proximity — catches across-the-street parcels
                    #   that don't share a physical edge but are still adjoiners.
                    if client_centroid:
                        clng, clat = client_centroid
                        RADIUS_DEG = 0.0011  # ~120m — catches parcels across a road
                        MAX_PROX = 6         # hard cap on proximity results
                        BOX = RADIUS_DEG * 1.5
                        for p in parcels:
                            if kml_prox_count >= MAX_PROX:
                                break
                            pc = p.get("centroid")
                            if not pc or p.get("upc") == client_upc:
                                continue
                            if abs(pc[0] - clng) > BOX or abs(pc[1] - clat) > BOX:
                                continue
                            dlng = abs(pc[0] - clng)
                            dlat = abs(pc[1] - clat)
                            if dlng < RADIUS_DEG and dlat < RADIUS_DEG:
                                name = _clean_owner_name(p.get("owner", "").title())
                                if not _is_valid_owner(name):
                                    continue
                                key  = name.lower()
                                if key not in seen_names:
                                    seen_names.add(key)
                                    results.append({
                                        "name":   name,
                                        "raw":    f"{pc[1]:.5f},{pc[0]:.5f}",
                                        "field":  "KML proximity",
                                        "source": "kml_proximity",
                                        "upc":    p.get("upc", ""),
                                        "plat":   p.get("plat", ""),
                                    })
                                    kml_prox_count += 1
                else:
                    print(f"[adjoiners][kml] no client parcel found for grantor={grantor_last!r} book={book_num!r}", flush=True)
            else:
                print("[adjoiners][kml] index not loaded (build it first)", flush=True)
        except Exception as kml_err:
            print(f"[adjoiners][kml] error: {kml_err}", flush=True)


        # ── Strategy 4: online location-range search (supplement) ───────────────
        #   Only add a few — these are same-book deeds, NOT geographic neighbors.
        #   Cap at 8 and only run if we have fewer than 10 results so far.
        MAX_ONLINE = 8
        online_count = 0
        if len(results) < 10:
            print(f"[adjoiners] online search: location={location!r} grantor={grantor!r}", flush=True)
            online = find_adjoiners_online(location, grantor, sess=_session())
            for om in online:
                if online_count >= MAX_ONLINE:
                    break
                if om["name"].lower() not in seen_names:
                    results.append(om)
                    seen_names.add(om["name"].lower())
                    online_count += 1
        else:
            print(f"[adjoiners] skipping online search — already have {len(results)} results", flush=True)

        # ── Hard cap: keep best results, drop lowest-priority sources ──────────
        #   Priority: legal_desc > deed_text > plat_ocr > kml_upc > kml_proximity > online_range
        MAX_ADJOINERS = 15
        if len(results) > MAX_ADJOINERS:
            SOURCE_PRIORITY = {
                "legal_desc": 0, "deed_text": 1, "plat_ocr": 2,
                "kml_geometry": 3, "kml_upc": 4, "kml_proximity": 5, "online_range": 6,
            }
            results.sort(key=lambda r: SOURCE_PRIORITY.get(r.get("source", ""), 9))
            results = results[:MAX_ADJOINERS]

        # Log breakdown by source for debugging
        src_counts = Counter(r.get("source", "?") for r in results)
        print(f"[adjoiners] total: {len(results)}  breakdown: {dict(src_counts)}", flush=True)

        return jsonify({
            "success":      True,
            "doc_no":       doc_no,
            "adjoiners":    results,
            "count":        len(results),
            "plat_used":    plat_file_used,
            "ocr_raw_text": raw_ocr_text[:8000] if raw_ocr_text else "",  # cap size
        })
    except Exception as e:

        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})

# _extract_plat_name_tokens — imported from helpers.cabinet
# parse_cabinet_refs — imported from helpers.cabinet




def _extract_target_cabinets(detail: dict, kml_matches: list = None) -> tuple[list[str], str]:
    """
    Determine which cabinet letter(s) to search for a plat, using:

    Priority 1 — KML matches already returned by /find-plat-kml.
                  Extract cab_refs from those hits (most reliable since
                  the KML PLAT field is tied to the parcel, not the owner name).

    Priority 2 — Deed Location (book/page) → query KML index directly.
                  The recording book/page is a stable reference that doesn't
                  change with ownership, unlike owner names on deeds/KML.

    Priority 3 — Cabinet refs embedded in the deed text itself (e.g. "Cabinet C-191A").

    Fallback    — Search all cabinets if none of the above yield a cabinet letter.

    Returns (cabinet_letters, reason_string) so callers can log / display why.

    NOTE: Name-based KML targeting is intentionally skipped — deeds and KML may
    reflect different eras of ownership, making name matches unreliable for targeting.
    """
    all_cabs = list(CABINET_FOLDERS.keys())

    # ── Priority 1: cab_refs from KML matches already resolved ─────────────────
    if kml_matches:
        letters = set()
        for hit in kml_matches:
            for cr in hit.get("cab_refs", []):
                letter = cr.split("-")[0].upper()
                if letter in CABINET_FOLDERS:
                    letters.add(letter)
        if letters:
            return sorted(letters), f"KML parcel cab_refs → Cabinet(s) {', '.join(sorted(letters))}"

    # ── Priority 2: deed Location book/page → KML index lookup ─────────────────
    location = detail.get("Location", "")
    if location:
        m = re.match(r'^[A-Za-z]?(\d+)-(\d+)', location.strip())
        if m:
            book_num = m.group(1)
            page_num = m.group(2)
            idx = xml_processor._cached_index
            if idx is None:
                try:
                    survey = get_survey_data_path()
                    idx = xml_processor.load_index(survey)
                except Exception:
                    idx = None
            if idx:
                hits = xml_processor.search_parcels_in_index(idx, book=book_num, page=page_num, limit=10)
                letters = set()
                for h in hits:
                    for cr in h.get("cab_refs", []):
                        letter = cr.split("-")[0].upper()
                        if letter in CABINET_FOLDERS:
                            letters.add(letter)
                if letters:
                    return sorted(letters), f"Deed location {book_num}-{page_num} → KML → Cabinet(s) {', '.join(sorted(letters))}"

    # ── Priority 3: cabinet refs directly in deed text ──────────────────────────
    deed_refs = parse_cabinet_refs(detail)
    if deed_refs:
        letters = list(dict.fromkeys(r["cabinet"] for r in deed_refs if r["cabinet"] in CABINET_FOLDERS))
        if letters:
            return letters, f"Deed text ref(s) → Cabinet(s) {', '.join(letters)}"

    # ── Fallback: search all ─────────────────────────────────────────────────────
    return all_cabs, "No cabinet target found — searching all cabinets"



# _extract_cabinet_display_name, _extract_cabinet_doc_number — imported from helpers.cabinet
# _cab_scan_cache — imported from helpers.cabinet

def search_local_cabinet(cabinet: str, doc_num: str,
                          grantor: str = "", grantee: str = "") -> list[dict]:
    """Delegates to helpers.cabinet.search_local_cabinet, injecting the cabinet path."""
    return _search_local_cabinet_impl(
        cabinet, doc_num, cabinet_path=get_cabinet_path(),
        grantor=grantor, grantee=grantee
    )



@app.route("/api/find-plat", methods=["POST"])
@require_auth
def api_find_plat():
    """
    INSTANT: Returns parsed cabinet refs from deed text — zero I/O.
    All slow I/O work is split into separate async endpoints:
      /api/find-plat-kml   — KML parcel index lookup
      /api/find-plat-local — local cabinet folder scan (name-based primary)
      /api/find-plat-online— online survey record search
    """
    try:
        data   = request.get_json() or {}
        detail = data.get("detail", {})
        cab_refs_list = parse_cabinet_refs(detail)
        return jsonify({
            "success":      True,
            "cabinet_refs": cab_refs_list,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/find-plat-kml", methods=["POST"])
@require_auth
def api_find_plat_kml():
    """
    KML parcel index cross-reference.
    Requires the index to be built first (via /api/xml/build-index).

    When deed detail is empty but client_name is provided, performs a
    name-only owner search so Step 3 can show KML results without a deed.
    """
    try:
        data        = request.get_json() or {}
        detail      = data.get("detail", {})
        client_name = data.get("client_name", "").strip()
        client_upc  = data.get("client_upc", "").strip()
        kml_matches = []
        idx = xml_processor._cached_index
        if idx is None:
            survey = get_survey_data_path()
            idx    = xml_processor.load_index(survey)
        if idx:
            survey = get_survey_data_path()
            # If we have a deed, use the full cross-reference (grantor/grantee/book/page/cab)
            if detail and (detail.get("Grantor") or detail.get("Grantee") or detail.get("Location")):
                kml_results = xml_processor.cross_reference_deed(survey, detail, client_upc=client_upc)
            elif client_name:
                # No deed yet — search by client name (owner contains last name)
                last_name = client_name.split(",")[0].strip().upper()
                if len(last_name) >= 2:
                    raw = xml_processor.search_parcels_in_index(idx, owner=last_name, operator="contains", limit=15)
                    for p in raw:
                        p["_match_reason"] = f"Client name: {client_name}"
                    # If we have a client_upc, boost that parcel to the front
                    if client_upc:
                        upc_hit = xml_processor.search_parcels_in_index(idx, upc=client_upc, limit=1)
                        if upc_hit:
                            upc_hit[0]["_match_reason"] = "Map selection"
                            # Remove if already in raw results, then prepend
                            raw = [p for p in raw if p.get("upc") != client_upc]
                            raw = upc_hit + raw
                    kml_results = raw
                else:
                    kml_results = []
            else:
                kml_results = []
            for p in kml_results:
                kml_matches.append({
                    "owner":        p.get("owner", ""),
                    "upc":          p.get("upc", ""),
                    "plat":         p.get("plat", ""),
                    "book":         p.get("book", ""),
                    "page":         p.get("page", ""),
                    "cab_refs":     p.get("cab_refs", []),
                    "cab_refs_str": ", ".join(p.get("cab_refs", [])) or p.get("plat", ""),
                    "centroid":     p.get("centroid"),
                    "match_reason": p.get("_match_reason", ""),
                    "local_files":  [],
                    "source":       "kml",
                })
        return jsonify({"success": True, "kml_matches": kml_matches})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e), "kml_matches": []})


@app.route("/api/find-plat-local", methods=["POST"])
@require_auth
def api_find_plat_local():
    """
    Targeted cabinet scan — locates plat PDFs by finding the correct cabinet
    first via KML parcel data or deed book/page, then scanning only there.

    Search priority:
      1. KML kml_matches.cab_refs  → target specific cabinet letter(s)
      2. Deed Location book/page   → KML index lookup → cabinet letter(s)
      3. Deed text cabinet refs    → cabinet letter(s) from deed itself
      4. Fallback                  → all cabinets (same as before)

    Within targeted cabinet(s), files are matched by:
      a. Exact cab_ref string  (e.g. "C-191A" in filename)   — most precise
      b. Client name token     (job subject name in filename) — owner-aware
      c. Deed grantor/grantee name in filename               — secondary
    """
    try:
        data           = request.get_json() or {}
        detail         = data.get("detail", {})
        grantor        = data.get("grantor", "") or detail.get("Grantor", "")
        grantee        = data.get("grantee", "") or detail.get("Grantee", "")
        client_name    = data.get("client_name", "")
        prior_owners   = data.get("prior_owners", [])   # grantor names from deed hint
        kml_matches    = data.get("kml_matches", [])   # may be populated if frontend chains requests
        forced_cabinets = [c.upper() for c in data.get("forced_cabinets", []) if c]
        # When the user manually typed a name in the 🔍 filter panel, suppress
        # client_name token matching so it doesn't produce false positives.
        name_override  = bool(data.get("name_override", False))

        # ── Determine which cabinet(s) to search ────────────────────────────────
        if forced_cabinets:
            # User explicitly selected a cabinet from the dropdown — override everything
            target_cabs    = forced_cabinets
            targeting_reason = f"Manual selection: Cabinet{'s' if len(forced_cabinets) > 1 else ''} {', '.join(forced_cabinets)}"
        else:
            target_cabs, targeting_reason = _extract_target_cabinets(detail, kml_matches)

        # Detect whether we have real deed detail (Location/book-page) or just a name.
        # Name-only mode = adjoiner plat search from Step 5 where no deed has been
        # looked up yet. In this mode we need to be MORE restrictive with matches.
        has_deed_detail = bool(
            detail.get("Location") or detail.get("Reference") or
            detail.get("_cab_refs")
        )

        print(f"[local] {targeting_reason}  (deed_detail={'yes' if has_deed_detail else 'NO — name-only mode'})", flush=True)


        local_hits       = []
        seen_local_paths = set()

        # ── Build cab_ref → doc map from KML hits for precise filename matching ─
        # e.g. kml says "C-191A" → look for "191A" in Cabinet C filenames
        # Also handles letter-only refs like "C" (used by the filter panel for targeting)
        kml_cab_refs = {}   # {"C": ["191A", "84"], ...}
        for hit in kml_matches:
            for cr in hit.get("cab_refs", []):
                parts = cr.split("-", 1)
                if len(parts) == 2:
                    letter, doc = parts[0].upper(), parts[1]
                    kml_cab_refs.setdefault(letter, [])
                    if doc not in kml_cab_refs[letter]:
                        kml_cab_refs[letter].append(doc)
                elif len(parts) == 1 and parts[0].upper() in CABINET_FOLDERS:
                    # Bare letter ref (e.g. "C") — used for cabinet targeting only
                    kml_cab_refs.setdefault(parts[0].upper(), [])

        # ── Also pull cab_refs from deed text for the targeted cabinets ─────────
        deed_refs = parse_cabinet_refs(detail)
        deed_cab_refs = {}  # {"C": ["191A"], ...}
        for r in deed_refs:
            deed_cab_refs.setdefault(r["cabinet"], [])
            if r["doc"] not in deed_cab_refs[r["cabinet"]]:
                deed_cab_refs[r["cabinet"]].append(r["doc"])

        # ── Build client name tokens for cabinet filename matching ───────────────
        # The client IS the current property owner, so their name appears on cabinet
        # files (e.g. "Rael Adela.pdf" or "Adela Rael.pdf").
        # We build tokens from client_name AND the deed grantee (same person).
        #
        #  client_name format examples: "Rael, Adela"  "ADELA RAEL"  "rael adela"
        #  Cabinet file name examples:  "Adela Rael.pdf"  "Rael Adela.pdf"
        #    → We need both "Rael" and "Adela" to reliably find the file.
        client_tokens: list[str] = []
        for raw_name in [client_name, grantee]:
            if not raw_name or not raw_name.strip():
                continue
            raw = raw_name.strip()
            # Add the full string (handles "Adela Rael" matching "Adela Rael.pdf")
            if len(raw) >= 3 and raw.lower() not in [t.lower() for t in client_tokens]:
                client_tokens.append(raw)
            # Swap "Last, First" → "First Last" (handles "Rael, Adela" → "Adela Rael")
            if ',' in raw:
                parts = [p.strip() for p in raw.split(',', 1) if p.strip()]
                if len(parts) == 2:
                    swapped = f"{parts[1]} {parts[0]}"
                    if swapped.lower() not in [t.lower() for t in client_tokens]:
                        client_tokens.append(swapped)
                # Also add just the last name (before comma)
                last = parts[0]
                if len(last) >= 3 and last.lower() not in [t.lower() for t in client_tokens]:
                    client_tokens.append(last)
            # Add individual words >= 4 chars
            for w in re.split(r'[\s,]+', raw):
                if len(w) >= 4 and w.lower() not in [t.lower() for t in client_tokens]:
                    client_tokens.append(w)

        # ── In name-only mode (no deed detail), restrict tokens to multi-word ─────
        # Single words like "Rael" (4 chars) would match "Israel", "Rafael", etc.
        # and cause hundreds of false positives across all cabinets.
        # Require at least "first last" (multi-word) or "last, first" (comma) tokens.
        if not has_deed_detail and client_tokens:
            multi_word_tokens = [t for t in client_tokens if ' ' in t or ',' in t]
            if multi_word_tokens:
                client_tokens = multi_word_tokens
                print(f"[local] name-only mode: restricted to multi-word tokens: {client_tokens}", flush=True)
            else:
                # If we only have single-word tokens (e.g. just a last name), keep them
                # but log a warning — results may be broad.
                print(f"[local] name-only mode: only single-word tokens available: {client_tokens}", flush=True)

        print(f"[local] client_tokens for cabinet search: {client_tokens}", flush=True)

        # ── Scan each target cabinet ─────────────────────────────────────────────
        for cab_letter in target_cabs:
            if not CABINET_FOLDERS.get(cab_letter):
                continue

            # a) Exact cab_ref match from KML or deed text
            #    NOTE: Most cabinet files do NOT embed the cabinet ref in their
            #    filename (e.g. files are named "Rael Adela.pdf", not "C-191A.pdf").
            #    kml_cab_refs are used exclusively for cabinet *targeting* above;
            #    name-based strategies below are what actually find the file.
            for doc in kml_cab_refs.get(cab_letter, []):
                try:
                    for h in search_local_cabinet(cab_letter, doc, "", ""):
                        if h["path"] not in seen_local_paths:
                            seen_local_paths.add(h["path"])
                            h["source"]   = "local"
                            h["ref"]      = f"{cab_letter}-{doc}"
                            h["strategy"] = "kml_cab_ref"
                            h["_tok_len"] = 200
                            local_hits.append(h)
                except Exception:
                    pass

            for doc in deed_cab_refs.get(cab_letter, []):
                try:
                    for h in search_local_cabinet(cab_letter, doc, "", ""):
                        if h["path"] not in seen_local_paths:
                            seen_local_paths.add(h["path"])
                            h["source"]   = "local"
                            h["ref"]      = f"{cab_letter}-{doc}"
                            h["strategy"] = "deed_cab_ref"
                            h["_tok_len"] = 150
                            local_hits.append(h)
                except Exception:
                    pass

            # b) KML PLAT field name tokens — PRIMARY name-based strategy.
            #     Cabinet files are named after the ORIGINAL plat filer / subdivider
            #     (e.g. "Adela Rael.pdf"), NOT the current owner. The KML PLAT
            #     field contains that original name, so it's the most reliable
            #     way to find the correct cabinet file by name.
            #     Example: "C-191-A ADELA RAEL" → tokens ["ADELA RAEL", "ADELA", "RAEL"]
            for hit in kml_matches:
                plat_tokens = _extract_plat_name_tokens(hit.get("plat", ""))
                for tok in plat_tokens:
                    try:
                        for h in search_local_cabinet(cab_letter, "", tok, ""):
                            if h["path"] not in seen_local_paths:
                                seen_local_paths.add(h["path"])
                                h["source"]   = "local"
                                h["ref"]      = "kml_plat_name"
                                h["strategy"] = "kml_plat_name"  # distinct strategy for sorting/display
                                h["_tok_len"] = len(tok) + 200   # highest KML name score
                                local_hits.append(h)
                    except Exception:
                        pass

            # b2) KML owner name — SECONDARY name-based strategy.
            #    The current parcel owner (e.g. GARZA, VERONICA) may or may not
            #    match the cabinet file name. Ownership changes over time but
            #    cabinet files don't get renamed, so this is a fallback.
            for hit in kml_matches:
                owner = hit.get("owner", "").strip()
                if not owner:
                    continue
                # Try full owner string, then each word >= 4 chars
                owner_tokens = [owner]
                for w in re.split(r'[\s,]+', owner):
                    if len(w) >= 4 and w not in owner_tokens:
                        owner_tokens.append(w)
                for tok in owner_tokens:
                    try:
                        for h in search_local_cabinet(cab_letter, "", tok, ""):
                            if h["path"] not in seen_local_paths:
                                seen_local_paths.add(h["path"])
                                h["source"]   = "local"
                                h["ref"]      = "kml_owner"
                                h["strategy"] = "kml_cab_ref"   # secondary KML match
                                h["_tok_len"] = len(tok) + 180
                                local_hits.append(h)
                    except Exception:
                        pass

            # c) Client name tokens — HIGH priority for Step 3 (client plat).
            #    The client IS the current owner, so their name IS on the cabinet file
            #    (e.g. "Rael, Adela" → finds "Adela Rael.pdf" or "Rael Adela.pdf").
            #    This also covers the grantee since they were pre-merged into client_tokens.
            #    SKIPPED when name_override is set (user typed a specific plat name)
            #    to avoid flooding results with owner-name false positives.
            if not name_override:
                for tok in client_tokens:
                    try:
                        for h in search_local_cabinet(cab_letter, "", tok, ""):
                            if h["path"] not in seen_local_paths:
                                seen_local_paths.add(h["path"])
                                h["source"]   = "local"
                                h["ref"]      = "client_name"
                                h["strategy"] = "client_name"
                                h["_tok_len"] = len(tok) + 100
                                local_hits.append(h)
                    except Exception:
                        pass

            # d) Prior owner name tokens — the plat may be filed under a previous owner.
            #    Build tokens the same way as client_tokens: full name + individual words.
            #    Strategy = 'prior_owner' ranks below client hits but above generic grantor.
            for raw_name in prior_owners:
                if not raw_name or not raw_name.strip():
                    continue
                raw = raw_name.strip()
                prior_tokens = [raw]
                if ',' in raw:
                    parts = [p.strip() for p in raw.split(',', 1) if p.strip()]
                    if len(parts) == 2:
                        prior_tokens.append(f"{parts[1]} {parts[0]}")
                    prior_tokens.append(parts[0])  # last name only
                for w in re.split(r'[\s,]+', raw):
                    if len(w) >= 4 and w not in prior_tokens:
                        prior_tokens.append(w)

                for tok in prior_tokens:
                    try:
                        for h in search_local_cabinet(cab_letter, "", tok, ""):
                            if h["path"] not in seen_local_paths:
                                seen_local_paths.add(h["path"])
                                h["source"]   = "local"
                                h["ref"]      = "prior_owner"
                                h["strategy"] = "prior_owner"
                                h["_tok_len"] = len(tok) + 60
                                local_hits.append(h)
                    except Exception:
                        pass

            # e) Grantor name from deed — lowest priority, may not match
            #    if ownership has changed since the deed was recorded.
            if grantor and grantor not in prior_owners:
                try:
                    for h in search_local_cabinet(cab_letter, "", grantor, ""):
                        if h["path"] not in seen_local_paths:
                            seen_local_paths.add(h["path"])
                            h["source"] = "local"
                            h["ref"]    = "grantor_search"
                            local_hits.append(h)
                except Exception:
                    pass

        # ── Auto-broaden: if 0 results and we targeted specific cabinets, retry ALL ──
        # DISABLED in name-only mode (no deed detail) to prevent flooding with
        # hundreds of false-positive name matches across all 6 cabinets.
        all_cabs = list(CABINET_FOLDERS.keys())
        if (not local_hits and set(target_cabs) != set(all_cabs) and client_tokens
                and has_deed_detail):   # ← name-only mode skips broadening
            print(f"[local] 0 results in target cabinet(s) {target_cabs} — broadening to all cabinets", flush=True)
            broadened_reason = f"Broadened search: no results in {', '.join(target_cabs)} — scanning all cabinets"
            for cab_letter in all_cabs:
                if cab_letter in target_cabs:
                    continue  # already searched
                if not CABINET_FOLDERS.get(cab_letter):
                    continue
                for tok in client_tokens:
                    try:
                        for h in search_local_cabinet(cab_letter, "", tok, ""):
                            if h["path"] not in seen_local_paths:
                                seen_local_paths.add(h["path"])
                                h["source"]   = "local"
                                h["ref"]      = "broadened_search"
                                h["strategy"] = "client_name"
                                h["_tok_len"] = len(tok) + 50  # lower than direct matches
                                local_hits.append(h)
                    except Exception:
                        pass
            if local_hits:
                targeting_reason = broadened_reason
        elif not local_hits and not has_deed_detail:
            print("[local] 0 results (name-only mode — auto-broaden disabled)", flush=True)

        # ── Diagnostic: log what was searched when 0 results ─────────────────────
        if not local_hits:
            kml_plat_tokens_searched = []
            for hit in kml_matches:
                plat_tokens = _extract_plat_name_tokens(hit.get("plat", ""))
                kml_plat_tokens_searched.extend(plat_tokens)
            print(
                f"[local] ZERO RESULTS — cabinet(s)={target_cabs!r} "
                f"kml_plat_tokens={kml_plat_tokens_searched!r} "
                f"client_tokens={client_tokens!r} "
                f"kml_cab_refs={dict(kml_cab_refs)!r} "
                f"deed_cab_refs={dict(deed_cab_refs)!r}",
                flush=True
            )
            # List first 20 files in each target cabinet so we can spot name mismatches
            cab_path = get_cabinet_path()
            for cl in target_cabs:
                folder = CABINET_FOLDERS.get(cl)
                if not folder:
                    continue
                cab_dir = Path(cab_path) / folder
                if not cab_dir.exists():
                    print(f"[local] Cabinet {cl} folder NOT FOUND: {cab_dir}", flush=True)
                    continue
                files_sample = [f.name for f in sorted(cab_dir.iterdir()) if f.suffix.lower() == '.pdf'][:20]
                print(f"[local] Cabinet {cl} first 20 PDFs: {files_sample}", flush=True)

        # ── Sort: doc_number → kml_plat_name → kml_cab_ref → deed_cab_ref → client_name → prior_owner → name_match ─
        # doc_number = exact plat doc number match from file's leading number (highest confidence)
        # kml_plat_name = PLAT field name match (original filer — most reliable name match)
        strategy_order = {"doc_number": 0, "kml_plat_name": 1, "kml_cab_ref": 2, "deed_cab_ref": 3,
                          "client_name": 4, "prior_owner": 5, "name_match": 6, "page_ref": 7}
        local_hits.sort(key=lambda r: (
            strategy_order.get(r.get("strategy", ""), 9),
            -(r.get("_tok_len") or 0)
        ))

        # ── Cap total results to prevent UI flooding ──────────────────────────────
        MAX_LOCAL_HITS = 25
        if len(local_hits) > MAX_LOCAL_HITS:
            print(f"[local] Capping {len(local_hits)} results to {MAX_LOCAL_HITS}", flush=True)
            local_hits = local_hits[:MAX_LOCAL_HITS]

        return jsonify({
            "success":          True,
            "local":            local_hits,
            "target_cabinets":  target_cabs,
            "targeting_reason": targeting_reason,
            "kml_matches":      [],   # KML enrichment now handled by find-plat-kml
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e), "local": [], "kml_matches": []})




@app.route("/api/find-plat-online", methods=["POST"])
@require_auth
def api_find_plat_online():
    """
    Slow online survey search (separate endpoint so it doesn't block the UI).
    Searches by client name (current owner), grantee, and deed grantor — in
    that order — since the survey plat lists the current owner as the grantor.
    """
    try:
        data        = request.get_json() or {}
        detail      = data.get("detail", {})
        grantor     = data.get("grantor", "") or detail.get("Grantor", "")
        grantee     = data.get("grantee", "") or detail.get("Grantee", "")
        client_name = data.get("client_name", "")
        hits      = []
        seen_docs = set()

        def _do_online_search(search_name, label):
            """Run one name search (full 'LAST, FIRST' format) and append any survey hits found."""
            if not search_name or len(search_name) < 2:
                return
            try:
                search_url = f"{_get_portal_url()}/scripts/hfweb.asp?Application=FNM&Database=TP"
                resp = _session().get(search_url, timeout=8)
                if "CROSSNAMEFIELD" not in resp.text and "FIELD14" not in resp.text:
                    return  # not logged in
                soup = BeautifulSoup(resp.text, "html.parser")
                fd = _scrape_form_data(soup)
                if not fd:
                    return
                fd["CROSSNAMEFIELD"] = search_name
                fd["CROSSNAMETYPE"]  = "begin"
                # NOTE: Do NOT set FIELD7="SUR" here — it conflicts with the form
                # and may suppress all results. Instrument-type filtering is done
                # by the regex below after results are returned.
                post  = _session().post(f"{_get_portal_url()}/scripts/hflook.asp", data=fd, timeout=10)
                soup2 = BeautifulSoup(post.text, "html.parser")
                for row in soup2.find_all("tr"):
                    cells = row.find_all("td")
                    if len(cells) < 9:
                        continue
                    doc_link = cells[1].find("a") if len(cells) > 1 else None
                    if not doc_link:
                        continue
                    doc_no = doc_link.text.strip()
                    if doc_no in seen_docs:
                        continue
                    itype = cells[5].text.strip() if len(cells) > 5 else ""
                    if not re.search(r"survey|plat|map|lot.?line|replat|subdiv", itype, re.I):
                        continue
                    seen_docs.add(doc_no)
                    hits.append({
                        "doc_no":          doc_no,
                        "instrument_type": itype,
                        "location":        cells[2].text.strip() if len(cells) > 2 else "",
                        "recorded_date":   cells[7].text.strip() if len(cells) > 7 else "",
                        "grantor":         cells[9].text.strip() if len(cells) > 9 else "",
                        "grantee":         cells[10].text.strip() if len(cells) > 10 else "",
                        "pdf_url":         f"{_get_portal_url()}/WebTemp/{doc_no}.pdf",
                        "source":          "online",
                        "search_label":    label,
                    })
            except Exception:
                pass

        # Build ordered list of full names to search (most specific first).
        # Previously this only sent the last name (e.g. "SMITH"), which was
        # far too broad.  Now we send the full "LAST, FIRST" just like the
        # Step 2 deed search does — the portal supports this natively.
        names_to_search = []   # list of (full_name, label) tuples
        seen_names = set()

        def _add_name(raw, label):
            if not raw:
                return
            # Normalise to "LAST, FIRST" — the portal's CROSSNAMEFIELD
            # accepts this format and matches much more precisely.
            full = raw.strip().upper()
            # Deduplicate on the normalised full name
            if full and len(full) >= 2 and full not in seen_names:
                seen_names.add(full)
                names_to_search.append((full, label))

        # Priority 1: client_name (current owner — most likely survey grantor)
        _add_name(client_name, "client")
        # Priority 2: grantee from deed (same person as client in Step 3)
        _add_name(grantee, "grantee")
        # Priority 3: grantor from deed (the seller — less likely to have a survey here)
        _add_name(grantor, "grantor")

        for full_name, label in names_to_search:
            _do_online_search(full_name, label)

        return jsonify({"success": True, "online": hits})
    except Exception as e:
        return jsonify({"success": False, "error": str(e), "online": []})


# ── save plat to project folder ────────────────────────────────────────────────

@app.route("/api/save-plat", methods=["POST"])
def api_save_plat():
    """Copy a local cabinet file OR download an online plat PDF into the project's B Plats folder."""
    # shutil imported at module top level
    try:
        data        = request.get_json()
        source      = data.get("source")         # "local" or "online"
        file_path   = data.get("file_path", "")  # for local
        doc_no      = data.get("doc_no", "")     # for online
        job_number  = data.get("job_number")
        client_name = data.get("client_name", "")
        job_type    = data.get("job_type", "BDY")
        filename    = data.get("filename", "")
        is_adjoiner = data.get("is_adjoiner", False)
        subject_id  = data.get("subject_id", "client")

        if not job_number or not client_name:
            return jsonify({"success": False, "error": "job_number and client_name required"})

        plats_root = (_job_base_path(job_number, client_name, job_type)
                      / "E Research" / "B Plats")
        plats_root.mkdir(parents=True, exist_ok=True)
        (plats_root / "Adjoiners").mkdir(exist_ok=True)

        dest_dir = plats_root / "Adjoiners" if is_adjoiner else plats_root

        # Build RTSI-convention filename: P1 filename.pdf / P2 SW Adj Name.pdf
        ref_num = _next_ref_number(dest_dir, prefix="P")
        adj_direction = data.get("adjoiner_direction", "").strip()

        base_name = doc_no or Path(file_path).stem
        if is_adjoiner and adj_direction:
            raw_filename = f"P{ref_num} {adj_direction} {base_name}.pdf"
        else:
            raw_filename = f"P{ref_num} {base_name}.pdf"

        if filename:
            # Caller provided explicit name — still prepend P# ref
            raw_filename = f"P{ref_num} {filename}"

        raw_filename = re.sub(r'[<>:"/\\|?*]', '', raw_filename).strip()
        if not raw_filename.endswith(".pdf"):
            raw_filename += ".pdf"

        dest = dest_dir / raw_filename


        # ── Duplicate check ───────────────────────────────────────────────────
        if dest.exists():
            return jsonify({
                "success":    True,
                "skipped":    True,
                "reason":     "File already exists in destination folder",
                "saved_to":   str(dest),
                "filename":   filename,
                "subject_id": subject_id,
            })

        if source == "local":
            shutil.copy2(file_path, dest)
        elif source == "online":
            pdf_url  = data.get("pdf_url", f"{_get_portal_url()}/WebTemp/{doc_no}.pdf")

            # Fetch with retry (portal may need a moment to generate PDF)
            pdf_resp, pdf_err = _fetch_portal_pdf(doc_no, pdf_url)
            if pdf_err:
                return jsonify({"success": False, "error": pdf_err})
            with open(dest, "wb") as f:
                for chunk in pdf_resp.iter_content(8192):
                    f.write(chunk)
        else:
            return jsonify({"success": False, "error": "source must be 'local' or 'online'"})

        # ── Mark plat saved in research session ───────────────────────────────
        try:
            rs = load_research(job_number, client_name, job_type)
            for subj in rs.get("subjects", []):
                if subj["id"] == subject_id:
                    subj["plat_saved"] = True
                    break
            save_research(job_number, client_name, job_type, rs)
        except Exception:
            pass

        return jsonify({"success": True, "skipped": False, "saved_to": str(dest), "filename": filename, "subject_id": subject_id})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ── PDF preview for plats ──────────────────────────────────────────────────────

@app.route("/api/preview-pdf", methods=["GET", "POST"])
def api_preview_pdf():
    """
    Render the first page of a local PDF as a JPEG image for in-app preview.
    Accepts either GET ?path=... or POST { path: "..." }.
    Security: validates the file is within the survey drive directory.
    """
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        pdf_path = data.get("path", "")
    else:
        pdf_path = request.args.get("path", "")

    if not pdf_path:
        return jsonify({"success": False, "error": "No path provided"}), 400

    p = Path(pdf_path)
    if not p.exists() or not p.is_file():
        return jsonify({"success": False, "error": "File not found"}), 404

    # Security: only allow files within the survey drive or project directories
    try:
        resolved = p.resolve()
    except Exception:
        return jsonify({"success": False, "error": "Invalid path"}), 400

    drive = detect_survey_drive()
    project_dir = Path(__file__).resolve().parent
    allowed = False
    if drive and str(resolved).upper().startswith(f"{drive}:\\"):
        allowed = True
    elif str(resolved).startswith(str(project_dir)):
        allowed = True
    if not allowed:
        return jsonify({"success": False, "error": "Path not within allowed directories"}), 403

    try:
        doc = fitz.open(str(p))
        if doc.page_count == 0:
            doc.close()
            return jsonify({"success": False, "error": "PDF has no pages"}), 400

        # Get requested page (default: first page)
        page_num = 0
        if request.method == "POST":
            page_num = min(int((request.get_json(silent=True) or {}).get("page", 0)), doc.page_count - 1)
        else:
            page_num = min(int(request.args.get("page", 0)), doc.page_count - 1)

        page = doc[page_num]
        total_pages = doc.page_count  # capture before closing
        # Render at 200 DPI for good quality without being too slow
        pix = page.get_pixmap(dpi=200)
        img_bytes = pix.tobytes("jpeg", jpg_quality=85)
        doc.close()

        response = make_response(img_bytes)
        response.headers["Content-Type"] = "image/jpeg"
        response.headers["Cache-Control"] = "public, max-age=3600"
        response.headers["X-Page-Count"] = str(total_pages)
        return response

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": f"Failed to render PDF: {str(e)}"}), 500


# ── next job number ────────────────────────────────────────────────────────────

@app.route("/api/next-job-number")
def api_next_job():
    num, range_folder = next_job_info()
    return jsonify({"next_job_number": num, "range_folder": range_folder})

# ── create project ─────────────────────────────────────────────────────────────

@app.route("/api/create-project", methods=["POST"])
@require_auth
def api_create_project():
    try:
        data = request.get_json()
        job_number = data.get("job_number")
        client_name = data.get("client_name", "")
        job_type = data.get("job_type", "BDY")

        if not job_number:
            job_number, _ = next_job_info()

        project_path, deeds_path = create_project_folders(job_number, client_name, job_type)
        return jsonify({"success": True, "job_number": job_number,
                        "project_path": project_path, "deeds_path": deeds_path})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

# ── download deed PDF ──────────────────────────────────────────────────────────

@app.route("/api/download", methods=["POST"])
def api_download():
    try:
        data        = request.get_json()
        doc_no      = data.get("doc_no", "")
        grantor     = data.get("grantor", "")
        grantee     = data.get("grantee", "")
        location    = data.get("location", "")
        job_number  = data.get("job_number")
        client_name = data.get("client_name", "")
        job_type    = data.get("job_type", "BDY")
        create_new  = data.get("create_project", False)
        is_adjoiner = data.get("is_adjoiner", False)
        adj_name    = data.get("adjoiner_name", "").strip()
        subject_id  = data.get("subject_id", "client")

        if not job_number:
            job_number, _ = next_job_info()

        if create_new:
            _, deeds_path = create_project_folders(job_number, client_name, job_type)
        else:
            deeds_path = str(_job_base_path(job_number, client_name, job_type)
                             / "E Research" / "A Deeds")
            Path(deeds_path).mkdir(parents=True, exist_ok=True)
            (Path(deeds_path) / "Adjoiners").mkdir(exist_ok=True)

        # Route to Adjoiners subfolder if needed
        dest_dir = Path(deeds_path) / "Adjoiners" if is_adjoiner else Path(deeds_path)

        # Build RTSI-convention filename: D1 Grantor to Grantee.pdf  / D2 SW Adj Name.pdf
        loc_clean = re.sub(r'^[A-Z]', '', location.strip())
        def clean_name(n):
            parts = n.split(",")
            return parts[0].strip().title() if parts else n.title()

        ref_num = _next_ref_number(dest_dir, prefix="D")
        adj_direction = data.get("adjoiner_direction", "").strip()  # e.g. "SW Adj"

        if is_adjoiner and adj_name:
            dir_label = f" {adj_direction}" if adj_direction else ""
            filename = f"D{ref_num}{dir_label} {clean_name(adj_name)} {doc_no}.pdf"
        else:
            grantor_short = clean_name(grantor)
            grantee_short = clean_name(grantee)
            filename = f"D{ref_num} {loc_clean} {grantor_short} to {grantee_short}.pdf"
        filename = re.sub(r'[<>:"/\\|?*]', '', filename).strip()
        if not filename.endswith(".pdf"):
            filename += ".pdf"

        save_path = dest_dir / filename


        # ── Duplicate check ───────────────────────────────────────────────────
        if save_path.exists():
            return jsonify({
                "success":      True,
                "skipped":      True,
                "reason":       "File already exists in destination folder",
                "saved_to":     str(save_path),
                "filename":     filename,
                "job_number":   job_number,
                "subject_id":   subject_id,
            })

        # Download — trigger portal to generate PDF, then fetch with retry
        pdf_url  = f"{_get_portal_url()}/WebTemp/{doc_no}.pdf"
        pdf_resp, pdf_err = _fetch_portal_pdf(doc_no, pdf_url)
        if pdf_err:
            return jsonify({"success": False, "error": pdf_err})

        with open(save_path, "wb") as f:
            for chunk in pdf_resp.iter_content(8192):
                f.write(chunk)

        # ── Mark deed saved in research session ───────────────────────────────
        try:
            rs = load_research(job_number, client_name, job_type)
            for subj in rs.get("subjects", []):
                if subj["id"] == subject_id:
                    subj["deed_saved"] = True
                    subj["deed_path"]  = str(save_path)
                    break
            save_research(job_number, client_name, job_type, rs)
        except Exception:
            pass  # non-fatal

        return jsonify({
            "success":    True,
            "skipped":    False,
            "saved_to":   str(save_path),
            "saved_path": str(save_path),
            "filename":   filename,
            "job_number": job_number,
            "subject_id": subject_id,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ── download deed PDF → user's browser ────────────────────────────────────────

@app.route("/api/download-to-browser", methods=["POST"])
def api_download_to_browser():
    """Stream a deed/plat PDF from the county portal directly to the user's browser.

    Unlike /api/download which saves to the server filesystem, this endpoint
    streams the PDF bytes with Content-Disposition: attachment so the user's
    browser downloads the file to their own computer.

    Body: { doc_no: str, grantor: str, grantee: str, location: str, filename: str (optional) }
    """
    try:
        data     = request.get_json() or {}
        doc_no   = data.get("doc_no", "")
        grantor  = data.get("grantor", "")
        grantee  = data.get("grantee", "")
        location = data.get("location", "")
        filename = data.get("filename", "")

        if not doc_no:
            return jsonify({"success": False, "error": "doc_no is required"}), 400

        # Build filename if not provided
        if not filename:
            loc_clean = re.sub(r'^[A-Z]', '', location.strip())
            def _cn(n):
                parts = n.split(",")
                return parts[0].strip().title() if parts else n.title()
            filename = f"{loc_clean} {_cn(grantor)} to {_cn(grantee)}.pdf"
            filename = re.sub(r'[<>:"/\\|?*]', '', filename).strip()
            if not filename.endswith(".pdf"):
                filename += ".pdf"

        # Trigger portal to generate PDF, then fetch with retry
        pdf_url  = f"{_get_portal_url()}/WebTemp/{doc_no}.pdf"
        pdf_resp, pdf_err = _fetch_portal_pdf(doc_no, pdf_url)
        if pdf_err:
            return jsonify({"success": False, "error": pdf_err}), 502

        # Stream the response to the browser
        def generate():
            for chunk in pdf_resp.iter_content(8192):
                yield chunk

        response = Response(generate(), content_type="application/pdf")
        response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
        response.headers["Cache-Control"] = "no-cache"
        return response

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# ── serve a local file → user's browser ───────────────────────────────────────

@app.route("/api/serve-local-file", methods=["POST"])
def api_serve_local_file():
    """Serve a local file (cabinet plat, saved deed, DXF, etc.) for browser download.

    Security: only allows files within the survey drive or project directories.

    Body: { path: str, filename: str (optional — defaults to file's basename) }
    """
    try:
        data      = request.get_json() or {}
        file_path = data.get("path", "").strip()
        filename  = data.get("filename", "").strip()

        if not file_path:
            return jsonify({"success": False, "error": "path is required"}), 400

        p = Path(file_path)
        if not p.exists() or not p.is_file():
            return jsonify({"success": False, "error": "File not found"}), 404

        # Security: restrict to survey drive and project directories
        try:
            resolved = p.resolve()
        except Exception:
            return jsonify({"success": False, "error": "Invalid path"}), 400

        drive = detect_survey_drive()
        project_dir = Path(__file__).resolve().parent
        allowed = False
        if drive and str(resolved).upper().startswith(f"{drive}:\\"):
            allowed = True
        elif str(resolved).startswith(str(project_dir)):
            allowed = True
        if not allowed:
            return jsonify({"success": False, "error": "Path not within allowed directories"}), 403

        if not filename:
            filename = p.name

        return send_file(
            str(resolved),
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# ── extract deed info (lightweight — for Step 5 adjoiner plat targeting) ────────


@app.route("/api/extract-deed-info", methods=["POST"])
def api_extract_deed_info():
    """
    Extract basic deed metadata from a saved PDF for plat search targeting.

    This is a lightweight alternative to full deed analysis — it extracts
    Grantor, Grantee, Location (book/page) from the filename pattern, and
    cabinet references from the PDF text.

    Filename pattern:  {book-page} {Grantor} to {Grantee}.pdf
    Example:           568-482 Martinez to Rael.pdf

    Body:  { pdf_path: str }
    Returns: { success, detail: { Grantor, Grantee, Location, ... } }
    """
    try:
        data     = request.get_json() or {}
        pdf_path = data.get("pdf_path", "").strip()
        if not pdf_path or not os.path.isfile(pdf_path):
            return jsonify({"success": False, "error": "File not found"})

        fname = Path(pdf_path).stem  #  e.g. "568-482 Martinez to Rael"
        detail: dict = {}

        # ── Parse filename for Location / Grantor / Grantee ──────────────────
        # Pattern:  "{book-page} {Grantor} to {Grantee}"
        m = re.match(
            r'^(\d+-\d+)\s+(.+?)\s+to\s+(.+)$', fname, re.I
        )
        if m:
            detail["Location"] = m.group(1).strip()
            detail["Grantor"]  = m.group(2).strip()
            detail["Grantee"]  = m.group(3).strip()
        else:
            # Fall back: just split on " to " if present
            if " to " in fname.lower():
                parts = re.split(r'\s+to\s+', fname, maxsplit=1, flags=re.I)
                if len(parts) == 2:
                    detail["Grantor"] = parts[0].strip()
                    detail["Grantee"] = parts[1].strip()

        # ── Scan PDF text for cabinet references (first 2 pages) ─────────────
        try:
            text, _ = _extract_pdf_text(pdf_path)
            if text:
                cab_refs = parse_cabinet_refs({"text": text[:5000]})
                if cab_refs:
                    detail["_cab_refs"] = cab_refs
                    # Also inject as text so parse_cabinet_refs works downstream
                    detail["Reference"] = " ".join(r["raw"] for r in cab_refs)
        except Exception:
            pass

        print(f"[extract-deed-info] {Path(pdf_path).name} → {detail}", flush=True)
        return jsonify({"success": True, "detail": detail})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ── drive status ────────────────────────────────────────────────────────────────

@app.route("/api/drive-status")
def api_drive_status():
    """Return detected removable drive info.  Add ?rescan=1 to force re-scan."""
    try:
        force  = request.args.get("rescan", "0") == "1"
        drive  = detect_survey_drive(force=force)
        survey = get_survey_data_path()
        cabinet= get_cabinet_path()
        return jsonify({
            "success":      True,
            "drive":        drive,          # None if not found
            "survey_path":  survey,
            "cabinet_path": cabinet,
            "drive_ok":     drive is not None and Path(survey).exists(),
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/drive-override", methods=["POST"])
def api_drive_override():
    """Manually pin a drive letter (saved to config.json as 'survey_drive')."""
    try:
        data   = request.get_json()
        letter = (data.get("drive") or "").strip().upper()
        cfg    = load_config()
        if letter:
            cfg["survey_drive"] = letter
        else:
            cfg.pop("survey_drive", None)
        save_config(cfg)
        drive = detect_survey_drive(force=True)
        return jsonify({"success": True, "drive": drive, "drive_ok": drive is not None})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ── research session ───────────────────────────────────────────────────────────

@app.route("/api/research-session", methods=["GET"])
def api_research_get():
    try:
        job_number  = request.args.get("job_number", "")
        client_name = request.args.get("client_name", "")
        job_type    = request.args.get("job_type", "BDY")
        if not job_number or not client_name:
            return jsonify({"success": False, "error": "job_number + client_name required"})
        data = load_research(job_number, client_name, job_type)
        # (field migration is now handled inside load_research)

        # Progress summary
        subjects = data.get("subjects", [])
        total    = len(subjects)
        deeds    = sum(1 for s in subjects if s.get("deed_saved"))
        plats    = sum(1 for s in subjects if s.get("plat_saved"))
        done     = sum(1 for s in subjects if s.get("deed_saved") and s.get("plat_saved"))
        data["progress"] = {"total": total, "deeds": deeds, "plats": plats, "done": done}

        return jsonify({"success": True, "session": data})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/research-session", methods=["POST"])
def api_research_post():
    try:
        body        = request.get_json()
        job_number  = body.get("job_number")
        client_name = body.get("client_name", "")
        job_type    = body.get("job_type", "BDY")
        session     = body.get("session", {})
        if not job_number or not client_name:
            return jsonify({"success": False, "error": "job_number + client_name required"})
        save_research(job_number, client_name, job_type, session)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

# ── research analytics (historical patterns + predictions) ─────────────────

@app.route("/api/research-analytics")
def api_research_analytics():
    """Return aggregate statistics from all historical research sessions.

    Scans all research.json files across job folders and computes:
      - Total jobs, subjects, deeds, plats
      - Average/median adjoiner counts
      - Cabinet letter distribution
      - Job type breakdown
      - Monthly activity (last 12 months)
      - Default complexity prediction (BDY jobs)

    Query params: refresh=1 to force re-scan (otherwise cached for 5 min)
    """
    try:
        from helpers.research_analytics import get_analytics
        survey = get_survey_data_path()
        if not survey:
            return jsonify({"success": False, "error": "Survey drive not connected"})
        force = request.args.get("refresh", "0") == "1"
        result = get_analytics(survey, force_refresh=force)
        return jsonify({"success": True, **result})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/research-analytics/predict", methods=["POST"])
def api_research_predict():
    """Predict complexity for a specific job based on historical patterns.

    Body: { job_type: str, trs: str (optional) }
    Returns: { predicted_adjoiners, predicted_complexity, likely_cabinets, confidence }
    """
    try:
        from helpers.research_analytics import scan_all_research, predict_job_complexity
        survey = get_survey_data_path()
        if not survey:
            return jsonify({"success": False, "error": "Survey drive not connected"})

        data = request.get_json() or {}
        job_type = data.get("job_type", "BDY")
        trs = data.get("trs", "")

        sessions = scan_all_research(survey)
        prediction = predict_job_complexity(sessions, job_type=job_type, trs=trs)
        return jsonify({"success": True, **prediction})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/session-completeness", methods=["POST"])
def api_session_completeness():
    """Score a research session's completeness.

    Body: { job_number: str, client_name: str, job_type: str (default "BDY") }
    Returns: { success, overall_score, deed_score, plat_score, status, missing_items }
    """
    try:
        data = request.get_json() or {}
        job_number  = data.get("job_number", "")
        client_name = data.get("client_name", "")
        job_type    = data.get("job_type", "BDY")

        if not job_number or not client_name:
            return jsonify({"success": False, "error": "job_number and client_name required"})

        session_data = load_research(job_number, client_name, job_type)
        result = _score_session_completeness(session_data)
        return jsonify({"success": True, **result})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ── AI Surveyor Bridge endpoints ─────────────────────────────────────────────

@app.route("/api/session/status", methods=["GET"])
def api_session_status():
    """Return the current research session status for the AI Surveyor bridge.

    Query: ?job_number=<str>&client_name=<str>&job_type=<str>
    Returns: {
        success, session, step, progress,
        property_description, calls_count, closure_error,
        subjects_summary: [{name, type, deed_saved, plat_saved}]
    }
    """
    try:
        job_number  = request.args.get("job_number", "")
        client_name = request.args.get("client_name", "")
        job_type    = request.args.get("job_type", "BDY")

        if not job_number or not client_name:
            return jsonify({"success": False, "error": "job_number + client_name required"})

        data = load_research(job_number, client_name, job_type)
        subjects = data.get("subjects", [])
        client = next((s for s in subjects if s.get("type") == "client"), None)

        # Determine current step based on what data exists
        step = 1  # Job Setup
        if client:
            if client.get("deed_saved"):
                step = 3  # Plat search
            elif client.get("name"):
                step = 2  # Deed search
            if any(s.get("type") == "adjoiner" for s in subjects):
                step = max(step, 4)  # Adjoiner board
            if all(s.get("deed_saved") and s.get("plat_saved") for s in subjects if s.get("type") == "adjoiner"):
                step = 6  # Export

        # Property description summary
        prop_desc = client.get("property_description", "") if client else ""
        calls_count = client.get("calls_count", 0) if client else 0

        total = len(subjects)
        deeds = sum(1 for s in subjects if s.get("deed_saved"))
        plats = sum(1 for s in subjects if s.get("plat_saved"))

        return jsonify({
            "success": True,
            "job_number": job_number,
            "client_name": client_name,
            "step": step,
            "progress": {"total": total, "deeds": deeds, "plats": plats},
            "has_property_description": bool(prop_desc),
            "calls_count": calls_count,
            "desc_type": client.get("desc_type", "") if client else "",
            "trs_refs": client.get("trs_refs", []) if client else [],
            "subjects_summary": [
                {
                    "name": s.get("name", ""),
                    "type": s.get("type", ""),
                    "deed_saved": bool(s.get("deed_saved")),
                    "plat_saved": bool(s.get("plat_saved")),
                    "upc": s.get("upc", ""),
                }
                for s in subjects
            ],
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/session/advance", methods=["POST"])
def api_session_advance():
    """Allow the AI Surveyor bridge to advance the research session.

    Body: {
        job_number: str, client_name: str, job_type: str,
        action: str (one of: "add_adjoiner", "mark_deed", "mark_plat", "set_description"),
        subject_name: str (for adjoiner actions),
        upc: str (optional),
        description: str (for set_description),
        desc_type: str (for set_description),
    }
    Returns: { success, message }
    """
    try:
        body = request.get_json() or {}
        job_number  = body.get("job_number", "")
        client_name = body.get("client_name", "")
        job_type    = body.get("job_type", "BDY")
        action      = body.get("action", "")

        if not job_number or not client_name:
            return jsonify({"success": False, "error": "job_number + client_name required"})
        if not action:
            return jsonify({"success": False, "error": "action required"})

        data = load_research(job_number, client_name, job_type)
        subjects = data.get("subjects", [])

        if action == "add_adjoiner":
            name = body.get("subject_name", "").strip()
            upc  = body.get("upc", "").strip()
            if not name and not upc:
                return jsonify({"success": False, "error": "subject_name or upc required"})
            # Check for duplicates
            existing = any(
                s.get("name", "").lower() == name.lower() or
                (upc and s.get("upc", "") == upc)
                for s in subjects if s.get("type") == "adjoiner"
            )
            if existing:
                return jsonify({"success": True, "message": f"Adjoiner '{name}' already exists"})
            subjects.append({
                "type": "adjoiner",
                "name": name or f"UPC {upc}",
                "upc": upc,
                "deed_saved": False,
                "plat_saved": False,
            })
            data["subjects"] = subjects
            save_research(job_number, client_name, data, job_type)
            return jsonify({"success": True, "message": f"Added adjoiner '{name or upc}'"})

        elif action == "mark_deed":
            target = body.get("subject_name", "").strip()
            for s in subjects:
                if s.get("name", "").lower() == target.lower():
                    s["deed_saved"] = True
                    break
            else:
                return jsonify({"success": False, "error": f"Subject '{target}' not found"})
            save_research(job_number, client_name, data, job_type)
            return jsonify({"success": True, "message": f"Marked deed saved for '{target}'"})

        elif action == "mark_plat":
            target = body.get("subject_name", "").strip()
            for s in subjects:
                if s.get("name", "").lower() == target.lower():
                    s["plat_saved"] = True
                    break
            else:
                return jsonify({"success": False, "error": f"Subject '{target}' not found"})
            save_research(job_number, client_name, data, job_type)
            return jsonify({"success": True, "message": f"Marked plat saved for '{target}'"})

        elif action == "set_description":
            desc_text = body.get("description", "").strip()
            desc_type = body.get("desc_type", "unknown")
            client = next((s for s in subjects if s.get("type") == "client"), None)
            if not client:
                return jsonify({"success": False, "error": "No client subject in session"})
            client["property_description"] = desc_text
            client["desc_type"] = desc_type
            save_research(job_number, client_name, data, job_type)
            return jsonify({"success": True, "message": "Property description updated"})

        else:
            return jsonify({"success": False, "error": f"Unknown action: {action}"})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ── legal description similarity search ───────────────────────────────────────

@app.route("/api/similar-descriptions", methods=["POST"])
def api_similar_descriptions():
    """Find parcels with legal descriptions similar to the given text.

    Uses multi-signal similarity scoring: TRS fingerprint, text tokens,
    cabinet refs, name overlap, and lot/block/tract matching.

    Body: { text: str, min_score: float (default 20), limit: int (default 20) }
    Returns: { success, results: [{ upc, owner, similarity: { score, ... } }], count }
    """
    try:
        from helpers.legal_similarity import search_similar_descriptions

        data      = request.get_json() or {}
        text      = data.get("text", "").strip()
        min_score = float(data.get("min_score", 20.0))
        limit     = int(data.get("limit", 20))

        if not text or len(text) < 10:
            return jsonify({"success": True, "results": [], "count": 0,
                            "hint": "Provide at least 10 characters of legal description text"})

        survey = get_survey_data_path()
        idx = xml_processor.load_index(survey)
        if not idx:
            return jsonify({"success": False, "error": "Parcel index not built yet",
                            "results": [], "count": 0})

        parcels = idx.get("parcels", [])
        results = search_similar_descriptions(
            text, parcels, min_score=min_score, limit=limit
        )

        return jsonify({"success": True, "results": results, "count": len(results)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e), "results": [], "count": 0})


# ── open folder in Explorer ────────────────────────────────────────────────────


def _is_safe_path(path: str) -> bool:
    """Validate that a path is within the survey drive or project directory."""
    if not path:
        return False
    try:
        resolved = Path(path).resolve()
        # Allow paths on the survey drive
        drive = detect_survey_drive()
        if drive and str(resolved).upper().startswith(f"{drive}:\\"):
            return True
        # Allow paths within the project directory itself
        project_dir = Path(__file__).resolve().parent
        if str(resolved).startswith(str(project_dir)):
            return True
        return False
    except Exception:
        return False


@app.route("/api/open-folder", methods=["POST"])
def api_open_folder():
    try:
        path = request.get_json().get("path", "")
        if not _is_safe_path(path):
            return jsonify({"success": False, "error": "Path not within allowed directories"})
        if os.path.exists(path):
            subprocess.Popen(["explorer", path])
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/open-file", methods=["POST"])
def api_open_file():
    """Open a specific file with the default Windows application."""
    try:
        path = request.get_json().get("path", "")
        if not _is_safe_path(path):
            return jsonify({"success": False, "error": "Path not within allowed directories"})
        if os.path.exists(path):
            os.startfile(path)
            return jsonify({"success": True})
        return jsonify({"success": False, "error": "File not found"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/recent-jobs")
def api_recent_jobs():
    """Scan Survey Data for the 10 most recently modified job folders."""
    try:
        survey_str = get_survey_data_path()
        if not survey_str or not Path(survey_str).exists():
            return jsonify({"success": True, "jobs": []})
        survey = Path(survey_str)
        jobs   = []
        for range_dir in survey.iterdir():
            if not range_dir.is_dir() or range_dir.name.startswith("00"):
                continue
            for job_dir in range_dir.iterdir():
                if not job_dir.is_dir():
                    continue
                m = re.match(r'^(\d{4})\s+(.*)', job_dir.name)
                if not m:
                    continue
                job_num    = int(m.group(1))
                client     = m.group(2).strip()
                job_type   = "BDY"
                for sub in job_dir.iterdir():
                    mt = re.match(r'^\d+-01-([A-Z]+)\s', sub.name)
                    if mt:
                        job_type = mt.group(1)
                        break
                jobs.append({
                    "job_number":  job_num,
                    "client_name": client,
                    "job_type":    job_type,
                    "modified":    job_dir.stat().st_mtime,
                })
        jobs.sort(key=lambda j: j["modified"], reverse=True)
        for j in jobs:
            del j["modified"]
        return jsonify({"success": True, "jobs": jobs[:10]})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/export-session", methods=["POST"])
@require_auth
def api_export_session():
    """Generate a plain-text research summary from the current session."""
    try:
        body = request.get_json()
        rs   = body.get("session", {})
        if not rs:
            return jsonify({"success": False, "error": "No session provided"})

        lines = [
            "DEED & PLAT RESEARCH SUMMARY",
            f"{'='*50}",
            f"Job #:    {rs.get('job_number')}",
            f"Client:   {rs.get('client_name')}",
            f"Type:     {rs.get('job_type')}",
            "",
        ]
        subjects = rs.get("subjects", [])
        prog     = rs.get("progress", {})
        if prog:
            lines.append(f"Progress: {prog.get('done',0)}/{prog.get('total',0)} complete  "
                         f"| Deeds: {prog.get('deeds',0)}  Plats: {prog.get('plats',0)}")
            lines.append("")

        lines.append(f"SUBJECTS ({len(subjects)}):")
        lines.append("-"*40)
        for s in subjects:
            deed_mark = "[✓]" if s.get("deed_saved") else "[ ]"
            plat_mark = "[✓]" if s.get("plat_saved") else "[ ]"
            status    = s.get("status", "pending").upper()
            lines.append(f"  {s['type'].upper():<10} {s['name']}")
            lines.append(f"    Deed {deed_mark}  Plat {plat_mark}  Status: {status}")
            if s.get("notes"):
                lines.append(f"    Notes: {s['notes']}")
            lines.append("")

        text = "\n".join(lines)
        return jsonify({"success": True, "text": text})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# (api_chain_search removed — callers now do `grantor.split(',')[0].strip()` locally)


# ══════════════════════════════════════════════════════════════════════════════
# BOUNDARY LINES & DXF GENERATION
# ══════════════════════════════════════════════════════════════════════════════

# All metes-and-bounds parsing functions are imported from helpers.metes_bounds





# ── DXF generator ─────────────────────────────────────────────────────────────

def _ensure_dwg_folder(job_number, client_name, job_type) -> Path:
    """Return (and create) the B Drafting/dwg folder for the job."""
    dwg_dir = _job_base_path(job_number, client_name, job_type) / "B Drafting" / "dwg"
    dwg_dir.mkdir(parents=True, exist_ok=True)
    return dwg_dir


def generate_boundary_dxf(
    parcels:     list[dict],
    job_number,
    client_name: str,
    job_type:    str,
    options:     dict = None,
) -> tuple[str, list[dict]]:
    """Delegates to helpers.dxf.generate_boundary_dxf, resolving output_dir."""
    dwg_dir = _ensure_dwg_folder(job_number, client_name, job_type)
    return _generate_dxf_impl(parcels, dwg_dir, job_number, client_name, job_type, options)



# ── /api/parse-calls ──────────────────────────────────────────────────────────

@app.route("/api/parse-calls", methods=["POST"])
def api_parse_calls():
    """
    Parse metes-and-bounds calls from arbitrary text.
    Body: { text: str, fields: [str]  (optional list of deed fields to scan) }
    Returns: { success, calls: [{bearing_label, azimuth, distance, bearing_raw}] }
    """
    try:
        data   = request.get_json()
        text   = data.get("text", "")
        detail = data.get("detail", {})  # full deed detail dict (optional bonus scan)

        # Combine: explicit text + any provided deed detail fields
        combined = text
        for field in ["Other_Legal", "Subdivision_Legal", "Comments",
                      "Legal Description", "Legal", "Reference", "Description"]:
            val = detail.get(field, "")
            if val and isinstance(val, str):
                combined += "\n" + val

        calls = parse_metes_bounds(combined)
        pts   = calls_to_coords(calls) if calls else []

        # Closure stats
        closure_err = 0.0
        if len(pts) >= 2:
            closure_err = round(math.hypot(pts[-1][0] - pts[0][0], pts[-1][1] - pts[0][1]), 4)

        return jsonify({
            "success":     True,
            "calls":       calls,
            "count":       len(calls),
            "closure_err": closure_err,
            "coords":      pts,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})



# ── /api/extract-calls-from-pdf ───────────────────────────────────────────────

@app.route("/api/extract-calls-from-pdf", methods=["POST"])
def api_extract_calls_from_pdf():
    """
    Given a path to a saved deed or plat PDF, extract metes-and-bounds calls.

    Strategy:
      1. Try fitz text extraction first (fast, works on text-based PDFs).
      2. Fall back to Tesseract OCR if no text is found (scanned image PDFs).
      3. Run parse_metes_bounds() on the combined text.

    Body:  { pdf_path: str }
    Returns: { success, calls, count, closure_err, coords, filename, source }
    """
    try:
        data     = request.get_json()
        pdf_path = data.get("pdf_path", "").strip()
        if not pdf_path or not os.path.exists(pdf_path):
            return jsonify({"success": False, "error": "File not found"})

        filename     = Path(pdf_path).name
        text, source = _extract_pdf_text(pdf_path)  # text layer first, OCR fallback


        calls = parse_metes_bounds(text)
        full  = calls_to_full_coords(calls) if calls else {}
        pts   = full.get("boundary_coords", []) if full else []
        tie_pts = full.get("tie_coords", []) if full else []
        has_tie = full.get("has_tie", False) if full else False
        closure_err = 0.0
        if len(pts) >= 2:
            closure_err = round(math.hypot(pts[-1][0] - pts[0][0], pts[-1][1] - pts[0][1]), 4)

        return jsonify({
            "success":     True,
            "calls":       calls,
            "count":       len(calls),
            "closure_err": closure_err,
            "coords":      pts,
            "tie_coords":  tie_pts,
            "has_tie":     has_tie,
            "filename":    filename,
            "source":      source,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ── /api/generate-dxf ─────────────────────────────────────────────────────────

@app.route("/api/generate-dxf", methods=["POST"])
@require_auth
@require_pro
def api_generate_dxf():
    """
    Generate a DXF file from one or more parcel call-lists.

    Body:
    {
      job_number:  int,
      client_name: str,
      job_type:    str,
      parcels: [
        {
          label:   str,          // e.g. "Client" or "Rael Adjoiner"
          layer:   str,          // "CLIENT" | "ADJOINERS" (optional, default CLIENT)
          start_x: float,        // optional, default 0
          start_y: float,        // optional, default 0
          calls: [
            { bearing_label, azimuth, distance }
          ]
        }
      ],
      options: {
        draw_boundary:   bool,
        draw_labels:     bool,
        draw_endpoints:  bool,
        label_size:      float,
        close_tolerance: float,
      }
    }
    Returns: { success, saved_to, filename, closure_errors: [{label, error}] }
    """
    try:
        data        = request.get_json()
        job_number  = data.get("job_number")
        client_name = data.get("client_name", "")
        job_type    = data.get("job_type", "BDY")
        parcels     = data.get("parcels", [])
        options     = data.get("options", {})

        if not job_number or not client_name:
            return jsonify({"success": False, "error": "job_number and client_name are required"})
        if not parcels:
            return jsonify({"success": False, "error": "No parcels provided"})

        saved_path, closure_errs = generate_boundary_dxf(
            parcels, job_number, client_name, job_type, options)
        filename = Path(saved_path).name

        return jsonify({
            "success":        True,
            "saved_to":       saved_path,
            "filename":       filename,
            "closure_errors": closure_errs,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ── cabinet browser ────────────────────────────────────────────────────────────

@app.route("/api/cabinet-browse")
def api_cabinet_browse():
    """
    List files in a cabinet folder, with optional name filter.
    Query params: cabinet (A-F), filter (substring, case-insensitive), page, per_page
    """
    try:
        cabinet  = (request.args.get("cabinet") or "").upper().strip()
        filt     = (request.args.get("filter") or "").lower().strip()
        page     = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 50))

        if not cabinet:
            # Return list of available cabinets
            return jsonify({"success": True, "cabinets": list(CABINET_FOLDERS.keys())})

        folder_name = CABINET_FOLDERS.get(cabinet)
        if not folder_name:
            return jsonify({"success": False, "error": f"Unknown cabinet '{cabinet}'"})

        cab_dir = Path(get_cabinet_path()) / folder_name
        if not cab_dir.exists():
            return jsonify({"success": False, "error": "Cabinet folder not found on disk"})

        files = []
        for f in sorted(cab_dir.iterdir()):
            if not f.is_file() or f.suffix.lower() != '.pdf':
                continue
            if filt and filt not in f.name.lower():
                continue
            files.append({
                "file":         f.name,
                "path":         str(f),
                "display_name": _extract_cabinet_display_name(f.name),
                "doc_number":   _extract_cabinet_doc_number(f.name),
                "size_kb":      round(f.stat().st_size / 1024),
            })

        total = len(files)
        start = (page - 1) * per_page
        paged = files[start:start + per_page]

        return jsonify({
            "success":  True,
            "cabinet":  cabinet,
            "total":    total,
            "page":     page,
            "per_page": per_page,
            "files":    paged,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ══════════════════════════════════════════════════════════════════════════════
# XML / KML PARCEL DATA
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/xml/status")
def api_xml_status():
    """Return status of the parcel index (exists, record count, age, source files)."""
    try:
        survey = get_survey_data_path()
        status = xml_processor.index_status(survey)
        return jsonify({"success": True, **status})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/index-health")
def api_index_health():
    """Return comprehensive health metrics about the parcel index.

    Reports completeness percentages (UPC, polygon, ArcGIS enrichment, owner names),
    index freshness (age, stale warning, newer XML files), and source breakdown.
    """
    try:
        survey = get_survey_data_path()
        health = xml_processor.compute_index_health(survey)
        return jsonify({"success": True, **health})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/data-conflicts")
def api_data_conflicts():
    """Detect cross-source anomalies between KML parcel data and ArcGIS enrichment.

    Flags owner mismatches, area discrepancies, TRS conflicts, and parcels
    missing ArcGIS data. Returns a summary and up to 200 individual conflicts.

    Query params: max_conflicts (int, default 200)
    """
    try:
        survey = get_survey_data_path()
        max_c  = int(request.args.get("max_conflicts", 200))
        result = xml_processor.detect_data_conflicts(survey, max_conflicts=max_c)
        return jsonify({"success": True, **result})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ══════════════════════════════════════════════════════════════════════════════
# MAP OVERLAY LAYERS — Water Rights & Geodetic Control
# ══════════════════════════════════════════════════════════════════════════════

_POD_BASE = (
    "https://gis.ose.nm.gov/server_s/rest/services"
    "/PODS/watersPods_currentFGDB/MapServer/0/query"
)

@app.route("/api/map-layers/water-rights")
def api_water_rights():
    """Proxy NM OSE Points of Diversion within a bounding box.

    Query params: minLat, maxLat, minLon, maxLon
    Returns GeoJSON-like array of water-right point features.
    """
    try:
        min_lat = float(request.args.get("minLat", 36.3))
        max_lat = float(request.args.get("maxLat", 36.5))
        min_lon = float(request.args.get("minLon", -105.7))
        max_lon = float(request.args.get("maxLon", -105.4))

        # Clamp bbox to reasonable size to avoid huge queries
        lat_span = max_lat - min_lat
        lon_span = max_lon - min_lon
        if lat_span > 0.3 or lon_span > 0.3:
            return jsonify({
                "success": True, "features": [], "count": 0,
                "message": "Zoom in further to see water rights (max 0.3° span)"
            })

        # ArcGIS envelope query — use inSR=4326 for WGS84 input coords
        envelope = f"{min_lon},{min_lat},{max_lon},{max_lat}"
        params = {
            "geometry":      envelope,
            "geometryType":  "esriGeometryEnvelope",
            "inSR":          "4326",
            "spatialRel":    "esriSpatialRelIntersects",
            "outFields":     "pod_file,pod_name,pod_status,use_of_well,"
                             "depth_well,own_lname,own_fname,ditch_name,"
                             "county,tws,rng,sec,pod_basin,use",
            "outSR":         "4326",
            "returnGeometry": "true",
            "f":             "json",
            "resultRecordCount": "500",
        }

        resp = req_lib.get(_POD_BASE, params=params, timeout=10)
        resp.raise_for_status()
        data = resp.json()

        features = []
        for feat in data.get("features", []):
            geom = feat.get("geometry", {})
            attr = feat.get("attributes", {})
            if not geom:
                continue

            # Build owner name from parts
            lname = (attr.get("own_lname") or "").strip()
            fname = (attr.get("own_fname") or "").strip()
            owner = f"{lname}, {fname}".strip(", ") if lname else fname

            features.append({
                "lat":         geom.get("y", 0),
                "lon":         geom.get("x", 0),
                "pod_file":    (attr.get("pod_file") or "").strip(),
                "name":        (attr.get("pod_name") or "").strip(),
                "status":      (attr.get("pod_status") or "").strip(),
                "use":         (attr.get("use_of_well") or attr.get("use") or "").strip(),
                "depth":       attr.get("depth_well") or 0,
                "owner":       owner,
                "ditch":       (attr.get("ditch_name") or "").strip(),
                "trs":         _pod_trs(attr),
                "basin":       (attr.get("pod_basin") or "").strip(),
            })

        return jsonify({"success": True, "features": features, "count": len(features)})

    except Exception as e:
        print(f"[water-rights] Error: {e}", flush=True)
        return jsonify({"success": False, "features": [], "error": str(e)})


def _pod_trs(attr: dict) -> str:
    """Build a TRS string from POD attributes."""
    twp = (attr.get("tws") or "").strip()
    rng = (attr.get("rng") or "").strip()
    sec = (attr.get("sec") or "").strip()
    if twp and rng:
        s = f"T{twp} R{rng}"
        if sec:
            s += f" Sec {sec}"
        return s
    return ""


_NGS_RADIAL = "https://geodesy.noaa.gov/api/nde/radial"

@app.route("/api/map-layers/survey-marks")
def api_survey_marks():
    """Proxy NGS geodetic control stations near a given point.

    Query params: lat, lon, radius (miles, default 3)
    Returns array of simplified survey mark objects.
    """
    try:
        lat    = float(request.args.get("lat", 36.4))
        lon    = float(request.args.get("lon", -105.6))
        radius = float(request.args.get("radius", 3))
        radius = min(radius, 10)  # cap at 10 miles

        resp = req_lib.get(_NGS_RADIAL, params={
            "lat": lat, "lon": lon,
            "radius": radius, "units": "MILE",
        }, timeout=12)
        resp.raise_for_status()
        marks_raw = resp.json()

        marks = []
        for m in marks_raw:
            pid = (m.get("pid") or "").strip()
            if not pid:
                continue
            lat_v = m.get("lat", "").strip()
            lon_v = m.get("lon", "").strip()
            if not lat_v or not lon_v:
                continue

            marks.append({
                "pid":           pid,
                "name":          (m.get("name") or "").strip(),
                "lat":           float(lat_v),
                "lon":           float(lon_v),
                "county":        (m.get("stCounty") or "").strip(),
                "orthoHt":       (m.get("orthoHt") or "").strip(),
                "vertDatum":     (m.get("vertDatum") or "").strip(),
                "monumentType":  (m.get("monumentType") or "").strip(),
                "stamping":      (m.get("stamping") or "").strip(),
                "setting":       (m.get("setting") or "").strip(),
                "condition":     (m.get("condition") or "").strip(),
                "lastRecovered": (m.get("lastRecovered") or "").strip(),
                "stability":     (m.get("stability") or "").strip(),
                "posDatum":      (m.get("posDatum") or "").strip(),
                "satUse":        (m.get("satUse") or "").strip() == "Y",
                "datasheet_url": f"https://geodesy.noaa.gov/cgi-bin/ds_mark.prl?PidBox={pid}",
            })

        return jsonify({"success": True, "marks": marks, "count": len(marks)})

    except Exception as e:
        print(f"[survey-marks] Error: {e}", flush=True)
        return jsonify({"success": False, "marks": [], "error": str(e)})


@app.route("/api/xml/build-index", methods=["POST"])
def api_xml_build_index():
    """Parse all KML/KMZ files in the XML folder and build/rebuild the parcel index."""
    try:
        survey = get_survey_data_path()
        result = xml_processor.build_index(survey)
        if "error" in result:
            return jsonify({"success": False, "error": result["error"]})
        return jsonify({"success": True, **result})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/xml/enrich-index", methods=["POST"])
def api_xml_enrich_index():
    """Trigger ArcGIS enrichment of the existing parcel index.

    Enriches parcels with TRS, legal description, subdivision, zoning, etc.
    without requiring a full KML/KMZ re-parse. The enriched data is saved
    back to the index JSON file.
    """
    try:
        survey = get_survey_data_path()
        idx = xml_processor.load_index(survey, force=True)
        if not idx:
            return jsonify({"success": False, "error": "No parcel index found. Build the index first."})

        stats = xml_processor.enrich_index_with_arcgis(idx)

        # Save updated index back to disk
        idx_path = xml_processor._index_path(survey)
        idx_path.write_text(json.dumps(idx, ensure_ascii=False), encoding="utf-8")

        # Force cache refresh
        xml_processor._cached_index = idx
        xml_processor._cached_index_mtime = idx_path.stat().st_mtime

        return jsonify({"success": True, **stats})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/xml/search", methods=["POST"])
def api_xml_search():
    """
    Search parcel index by owner name, UPC, book/page, cabinet ref, TRS, or subdivision.

    Body: { owner, upc, book, page, cabinet_ref, trs, subdivision, operator, limit }
    """
    try:
        survey = get_survey_data_path()
        data   = request.get_json()
        results = xml_processor.search_parcels(
            survey,
            owner=data.get("owner", ""),
            upc=data.get("upc", ""),
            book=data.get("book", ""),
            page=data.get("page", ""),
            cabinet_ref=data.get("cabinet_ref", ""),
            trs=data.get("trs", ""),
            subdivision=data.get("subdivision", ""),
            operator=data.get("operator", "contains"),
            limit=int(data.get("limit", 50)),
        )
        return jsonify({"success": True, "results": results, "count": len(results)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/xml/parcel/<upc>")
def api_xml_parcel(upc):
    """Get full parcel detail including polygon coordinates for a given UPC."""
    try:
        survey = get_survey_data_path()

        # Search index for the record
        results = xml_processor.search_parcels(survey, upc=upc, limit=1)
        if not results:
            return jsonify({"success": False, "error": "Parcel not found"})

        parcel = results[0]

        # Extract full polygon coordinates (re-scans KML — slower but accurate)
        polygon = xml_processor.extract_parcel_polygon(survey, upc)
        parcel["polygon"] = polygon

        return jsonify({"success": True, "parcel": parcel})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/xml/cross-reference", methods=["POST"])
def api_xml_cross_reference():
    """
    Given a deed detail dict, find matching parcels via name, book/page, cabinet refs.
    Body: { detail: {...deed detail...} }
    """
    try:
        survey = get_survey_data_path()
        data   = request.get_json()
        detail = data.get("detail", {})
        if not detail:
            return jsonify({"success": False, "error": "No deed detail provided"})

        results = xml_processor.cross_reference_deed(survey, detail)
        return jsonify({"success": True, "results": results, "count": len(results)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ── /api/parcel-search  (Step 1 Property Picker) ─────────────────────────────

@app.route("/api/parcel-search", methods=["POST"])
def api_parcel_search():
    """
    Search KML parcel index by owner name or UPC for the Step 1 property picker.

    Body: { query: str, operator: str ("contains"|"begins"|"exact"), limit: int }
    Returns: { success, results: [{owner, upc, book, page, plat, cab_refs, centroid, polygon}], count }
    """
    try:
        data     = request.get_json() or {}
        query    = data.get("query", "").strip()
        operator = data.get("operator", "contains")
        limit    = int(data.get("limit", 30))

        if not query or len(query) < 2:
            return jsonify({"success": True, "results": [], "count": 0,
                            "hint": "Enter at least 2 characters to search"})

        survey = get_survey_data_path()
        idx = xml_processor._cached_index
        if idx is None:
            idx = xml_processor.load_index(survey)
        if not idx:
            return jsonify({"success": False, "error": "Parcel index not built yet. Use the KML Index button to build it.",
                            "results": [], "count": 0})

        # Search by owner name
        results = xml_processor.search_parcels_in_index(
            idx, owner=query, operator=operator, limit=limit
        )

        # If no name hits and query looks like a UPC (all digits), try UPC search
        if not results and re.match(r'^\d+$', query):
            results = xml_processor.search_parcels_in_index(
                idx, upc=query, limit=limit
            )

        # Return minimal fields (strip heavy polygon data for list view)
        out = []
        for p in results:
            out.append({
                "owner":    p.get("owner", ""),
                "upc":      p.get("upc", ""),
                "book":     p.get("book", ""),
                "page":     p.get("page", ""),
                "plat":     p.get("plat", ""),
                "cab_refs": p.get("cab_refs", []),
                "centroid": p.get("centroid"),
            })

        return jsonify({"success": True, "results": out, "count": len(out)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e), "results": [], "count": 0})


# ── /api/xml/map-geojson  (Leaflet map data) ─────────────────────────────────

@app.route("/api/xml/map-geojson", methods=["POST"])
def api_xml_map_geojson():
    """
    Return a GeoJSON FeatureCollection for Leaflet rendering.
    Body: { highlight_upcs: [str], max_features: int }
    Response is gzip-compressed when the client accepts it (browsers always do).

    Auto-builds the index if KML/KMZ files exist but the index hasn't been created yet.
    """
    try:
        data           = request.get_json() or {}
        highlight_upcs = data.get("highlight_upcs", [])
        max_features   = int(data.get("max_features", 100000))
        source_filter  = data.get("source_filter", "")

        survey  = get_survey_data_path()

        # Auto-build index if it doesn't exist but KML/KMZ files are available
        idx = xml_processor.load_index(survey)
        if not idx:
            xml_files = xml_processor.discover_xml_files(survey)
            if xml_files:
                print(f"[map-geojson] No index found — auto-building from {len(xml_files)} XML/KML/KMZ files...", flush=True)
                build_result = xml_processor.build_index(survey)
                print(f"[map-geojson] Auto-build complete: {build_result.get('total', 0)} parcels in {build_result.get('elapsed_sec', '?')}s", flush=True)
            else:
                print("[map-geojson] No index and no KML/KMZ files found in XML folder", flush=True)

        geojson = xml_processor.get_map_geojson(
            survey, highlight_upcs, max_features, source_filter=source_filter
        )
        total   = len(geojson.get("features", []))
        sources = geojson.pop("sources", [])  # separate from FeatureCollection

        payload = json.dumps({"success": True, "geojson": geojson, "total": total, "sources": sources},
                             separators=(",", ":"))  # compact JSON — no spaces

        # Compress if client supports it (all modern browsers do)
        accept_enc = request.headers.get("Accept-Encoding", "")
        if "gzip" in accept_enc:
            compressed = gzip.compress(payload.encode("utf-8"), compresslevel=6)
            return Response(
                compressed,
                status=200,
                mimetype="application/json",
                headers={
                    "Content-Encoding": "gzip",
                    "Content-Length":   str(len(compressed)),
                    "Vary":             "Accept-Encoding",
                },
            )
        return Response(payload, status=200, mimetype="application/json")

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e), "total": 0})



# ══════════════════════════════════════════════════════════════════════════════
# PROPERTY ADDRESS LOOKUP  (ArcGIS + Nominatim fallback)
# ══════════════════════════════════════════════════════════════════════════════

# In-memory cache: { "upc:12345" or "ll:lat,lon" : { address dict } }
_address_cache: dict = {}
_nominatim_last_call: float = 0.0   # monotonic timestamp of last Nominatim call

# ArcGIS config is now fully dynamic — see _get_arcgis_config() above.
# All queries use _get_arcgis_config() at request time so each user/profile
# can point to their own county's ArcGIS REST layer.

# ── Nominatim (fallback) ─────────────────────────────────────────────────────
NOMINATIM_URL = "https://nominatim.openstreetmap.org/reverse"
NOMINATIM_HEADERS = {
    "User-Agent": "DeedPlatHelper/1.0 (land-survey-research-tool)",
    "Accept": "application/json",
}


def _arcgis_lookup_upc(upc: str, arcgis_cfg: dict = None) -> dict:
    """Query the configured ArcGIS parcel layer by parcel ID.

    Uses the dynamic ArcGIS config (URL + field mapping) so any county's
    layer works, not just Taos NM.
    Returns a normalised dict with address fields, or None on failure.
    """
    if not upc:
        return None

    cache_key = f"upc:{upc}"
    if cache_key in _address_cache:
        return _address_cache[cache_key]

    cfg = arcgis_cfg or _get_arcgis_config()
    fld = lambda c: _arcgis_field(cfg, c)  # shorthand

    pid_field = fld('parcel_id')
    out_concepts = [
        'parcel_id', 'owner',
        'address_all', 'address1', 'street_no', 'street_name', 'city', 'zipcode',
        'legal', 'area', 'subdivision', 'zoning', 'land_use',
        'township', 'twp_dir', 'range', 'rng_dir', 'section',
        'struct_count', 'struct_type', 'owner_type', 'mail_addr',
    ]
    out_fields = _arcgis_out_fields(cfg, out_concepts)

    try:
        resp = req_lib.get(
            cfg['url'],
            params={
                "where":          f"{pid_field}='{upc}'",
                "outFields":      out_fields,
                "returnGeometry": "false",
                "f":              "json",
            },
            headers={"User-Agent": "DeedPlatHelper/1.0"},
            timeout=15,
        )
        if resp.status_code != 200:
            print(f"[address] ArcGIS returned {resp.status_code} for {pid_field}={upc}", flush=True)
            return None

        data = resp.json()
        features = data.get("features", [])
        if not features:
            print(f"[address] ArcGIS: no features for {pid_field}={upc}", flush=True)
            return None

        attrs = features[0].get("attributes", {})
        g = lambda c: (attrs.get(fld(c)) or "").strip() if isinstance(attrs.get(fld(c)), str) else (attrs.get(fld(c)) or "")

        # Build short address — prefer address1, then street_no+name, then address_all
        situs_all = g('address_all')
        situs1    = g('address1')
        street_no = g('street_no')
        street_nm = g('street_name')
        city      = g('city')
        zipcode   = g('zipcode')

        if situs1:
            short_addr = situs1 + (f", {city}" if city else "")
        elif street_no and street_nm:
            short_addr = f"{street_no} {street_nm}" + (f", {city}" if city else "")
        elif situs_all and situs_all != zipcode:
            short_addr = situs_all
        else:
            short_addr = ""

        # Build TRS string
        twp     = str(g('township'))
        twp_dir = str(g('twp_dir'))
        rng     = str(g('range'))
        rng_dir = str(g('rng_dir'))
        sec     = str(g('section'))
        trs_str = ""
        if twp and rng:
            trs_str = f"T{twp}{twp_dir} R{rng}{rng_dir}"
            if sec:
                trs_str += f" Sec {sec}"

        # land_area may be numeric
        area_raw = attrs.get(fld('area'))
        area_val = area_raw if area_raw is not None else ""

        result = {
            "success":           True,
            "source":            "arcgis",
            "short_address":     short_addr or "(no street address on file)",
            "situs_full":        situs_all,
            "situs_address1":    situs1,
            "street_number":     street_no,
            "street_name":       street_nm,
            "city":              city,
            "zipcode":           zipcode,
            "owner_official":    g('owner'),
            "legal_description": g('legal'),
            "land_area":         area_val,
            "upc":               upc,
            "has_street_address": bool(situs1 or (street_no and street_nm)),
            "subdivision":       g('subdivision'),
            "zoning":            g('zoning'),
            "land_use":          g('land_use'),
            "trs":               trs_str,
            "structure_count":   attrs.get(fld('struct_count')) or 0,
            "structure_type":    g('struct_type'),
            "owner_type":        g('owner_type'),
            "mail_address":      g('mail_addr'),
            # Pass through the config so callers know what layer was used
            "arcgis_url":        cfg['url'],
        }

        _address_cache[cache_key] = result
        print(f"[address] ArcGIS {pid_field}={upc} → {result['short_address']}", flush=True)
        return result

    except Exception as e:
        print(f"[address] ArcGIS error for {upc}: {e}", flush=True)
        return None


def _nominatim_reverse(lat: float, lon: float) -> dict:
    """Call Nominatim reverse-geocode for a single lat/lon pair.

    Returns a dict with address fields, or an error dict.
    Enforces 1-second rate limiting and caches results.
    """
    import time as _time
    global _nominatim_last_call

    cache_key = f"ll:{round(lat, 5)},{round(lon, 5)}"
    if cache_key in _address_cache:
        return _address_cache[cache_key]

    # Rate-limit: min 1 second between Nominatim calls
    now = _time.monotonic()
    wait = 1.05 - (now - _nominatim_last_call)
    if wait > 0:
        _time.sleep(wait)

    try:
        resp = req_lib.get(
            NOMINATIM_URL,
            params={
                "format": "json",
                "lat": lat,
                "lon": lon,
                "addressdetails": 1,
                "zoom": 18,
            },
            headers=NOMINATIM_HEADERS,
            timeout=10,
        )
        _nominatim_last_call = _time.monotonic()

        if resp.status_code != 200:
            return {"success": False, "source": "nominatim", "error": f"HTTP {resp.status_code}"}

        data = resp.json()
        addr = data.get("address", {})

        # Build short human-readable address
        parts = []
        if addr.get("house_number"):
            parts.append(addr["house_number"])
        if addr.get("road"):
            parts.append(addr["road"])
        if not parts and addr.get("hamlet"):
            parts.append(addr["hamlet"])
        if not parts and addr.get("village"):
            parts.append(addr["village"])
        locality = addr.get("town") or addr.get("city") or addr.get("village") or addr.get("hamlet") or ""
        if locality and locality not in parts:
            parts.append(locality)

        result = {
            "success":           True,
            "source":            "nominatim",
            "short_address":     ", ".join(parts) if parts else "(no address found)",
            "road":              addr.get("road", ""),
            "house_number":      addr.get("house_number", ""),
            "hamlet":            addr.get("hamlet", ""),
            "village":           addr.get("village", ""),
            "town":              addr.get("town", ""),
            "city":              addr.get("city", ""),
            "county":            addr.get("county", ""),
            "state":             addr.get("state", ""),
            "postcode":          addr.get("postcode", ""),
            "has_street_address": bool(addr.get("house_number") and addr.get("road")),
            "lat":               lat,
            "lon":               lon,
        }

        _address_cache[cache_key] = result
        print(f"[address] Nominatim {lat},{lon} → {result['short_address']}", flush=True)
        return result

    except Exception as e:
        _nominatim_last_call = _time.monotonic()
        print(f"[address] Nominatim error for {lat},{lon}: {e}", flush=True)
        return {"success": False, "source": "nominatim", "error": str(e)}


@app.route("/api/property-address", methods=["POST"])
@require_auth
def api_property_address():
    """Look up property address info.

    Dual strategy:
      1. If UPC provided → query NM ArcGIS parcel database (official situs address)
      2. Fallback → Nominatim reverse geocode from lat/lon centroid

    Body: { "upc": str, "lat": float, "lon": float }
    Returns: { success, short_address, source, ... }
    """
    try:
        data = request.get_json() or {}
        upc  = (data.get("upc") or "").strip()
        lat  = float(data.get("lat", 0))
        lon  = float(data.get("lon", 0))

        # Strategy 1: ArcGIS by UPC (preferred — official govt data, no rate limit)
        arcgis_result = None
        if upc:
            arcgis_result = _arcgis_lookup_upc(upc)
            if arcgis_result and arcgis_result.get("success") and arcgis_result.get("has_street_address"):
                # ArcGIS has a real situs address — use it directly
                return jsonify(arcgis_result)

        # Strategy 2: Nominatim reverse geocode from coordinates (fallback)
        # Also triggers when ArcGIS returned data but no street address
        if lat != 0 or lon != 0:
            result = _nominatim_reverse(lat, lon)
            # Merge supplemental ArcGIS data (owner, legal desc) into Nominatim result
            if arcgis_result and arcgis_result.get("success"):
                for key in ("owner_official", "legal_description", "land_area", "upc"):
                    if arcgis_result.get(key) and not result.get(key):
                        result[key] = arcgis_result[key]
            return jsonify(result)

        # No coordinates either — return ArcGIS result as-is (even without street addr)
        if arcgis_result and arcgis_result.get("success"):
            return jsonify(arcgis_result)

        return jsonify({"success": False, "error": "No UPC or coordinates provided"})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/batch-property-address", methods=["POST"])
@require_auth
def api_batch_property_address():
    """Look up addresses for up to 10 parcels.

    Body: { "parcels": [ { "upc": str, "lat": float, "lon": float }, ... ] }
    Returns: { success, results: [ { short_address, source, ... }, ... ] }
    """
    try:
        data    = request.get_json() or {}
        parcels = data.get("parcels", [])[:10]

        if not parcels:
            return jsonify({"success": False, "error": "No parcels provided"})

        results = []
        for p in parcels:
            upc = (p.get("upc") or "").strip()
            lat = float(p.get("lat", 0))
            lon = float(p.get("lon", 0))

            result = None
            if upc:
                result = _arcgis_lookup_upc(upc)
            if not result or not result.get("success"):
                if lat != 0 or lon != 0:
                    result = _nominatim_reverse(lat, lon)
            if not result:
                result = {"success": False, "error": "No UPC or coordinates"}
            results.append(result)

        return jsonify({"success": True, "results": results})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ══════════════════════════════════════════════════════════════════════════════
# ARCGIS SPATIAL ADJOINER DISCOVERY
# ══════════════════════════════════════════════════════════════════════════════

def _arcgis_get_parcel_geometry(upc: str, arcgis_cfg: dict = None) -> dict | None:
    """Fetch the polygon geometry for a parcel from ArcGIS by parcel ID.
    Uses the dynamic ArcGIS config so any county layer works.
    Returns { "rings": [...], "spatialReference": {...} } or None.
    """
    if not upc:
        return None
    cfg = arcgis_cfg or _get_arcgis_config()
    pid_field = _arcgis_field(cfg, 'parcel_id')
    owner_field = _arcgis_field(cfg, 'owner')
    try:
        resp = req_lib.get(
            cfg['url'],
            params={
                "where":          f"{pid_field}='{upc}'",
                "outFields":      f"{pid_field},{owner_field}",
                "returnGeometry": "true",
                "outSR":          "4326",
                "f":              "json",
            },
            headers={"User-Agent": "DeedPlatHelper/1.0"},
            timeout=15,
        )
        if resp.status_code != 200:
            return None
        features = resp.json().get("features", [])
        if not features:
            return None
        return features[0].get("geometry")
    except Exception as e:
        print(f"[arcgis-adj] Geometry fetch error for {upc}: {e}", flush=True)
        return None


def _arcgis_find_touching_parcels(geometry: dict, arcgis_cfg: dict = None) -> list:
    """Query ArcGIS for all parcels that spatially touch the given polygon.

    Precision strategy:
      1. esriSpatialRelTouches — strict shared boundary. If ≥1 result, use it.
      2. esriSpatialRelIntersects — only if Touches returns zero results.
    Distance filter: 200m centroid-to-centroid to exclude road-gap parcels.
    Corner-only touches filtered by requiring the adjoiner centroid to be
    within ±90° of any boundary edge direction (not purely diagonal).
    """
    cfg = arcgis_cfg or _get_arcgis_config()
    fld = lambda c: _arcgis_field(cfg, c)
    adj_concepts = [
        'parcel_id', 'owner', 'area', 'subdivision', 'legal',
        'address_all', 'township', 'twp_dir', 'range', 'rng_dir', 'section',
    ]
    out_fields = _arcgis_out_fields(cfg, adj_concepts)

    # Compute client parcel centroid + bounding box for distance filtering
    client_cx, client_cy = 0.0, 0.0
    client_bbox_m = 300.0  # default if we can't compute
    try:
        rings = geometry.get("rings", [])
        if rings and rings[0]:
            xs = [p[0] for p in rings[0]]
            ys = [p[1] for p in rings[0]]
            client_cx = sum(xs) / len(xs)
            client_cy = sum(ys) / len(ys)
            # Estimate bbox diagonal in metres (used to scale distance threshold)
            w_m = (max(xs) - min(xs)) * 90000
            h_m = (max(ys) - min(ys)) * 111000
            client_bbox_m = (w_m**2 + h_m**2) ** 0.5
    except Exception:
        pass

    # Dynamic distance cap: larger parcels get a bigger tolerance, but cap at 400m
    # Typical Taos rural parcel diagonal is 200-600m; actual adjoiners share a boundary
    # so their centroid will be within roughly 1 parcel-width away.
    max_dist_m = min(max(client_bbox_m * 0.75, 150.0), 400.0)
    print(f"[arcgis-adj] Client parcel bbox_diag={client_bbox_m:.0f}m, distance cap={max_dist_m:.0f}m", flush=True)

    def _query_spatial(spatial_rel: str) -> list:
        """Run one ArcGIS spatial query and return raw features list."""
        try:
            resp = req_lib.get(
                cfg['url'],
                params={
                    "geometry":          json.dumps(geometry),
                    "geometryType":      "esriGeometryPolygon",
                    "spatialRel":        spatial_rel,
                    "inSR":              "4326",
                    "outSR":             "4326",
                    "outFields":         out_fields,
                    "returnGeometry":    "true",
                    "resultRecordCount": "20",
                    "f":                 "json",
                },
                headers={"User-Agent": "DeedPlatHelper/1.0"},
                timeout=20,
            )
            if resp.status_code != 200:
                return []
            return resp.json().get("features", [])
        except Exception as e:
            print(f"[arcgis-adj] Spatial query error ({spatial_rel}): {e}", flush=True)
            return []

    # Run Touches first; only fall back to Intersects if Touches returns nothing
    features = _query_spatial("esriSpatialRelTouches")
    spatial_rel_used = "esriSpatialRelTouches"
    if not features:
        print("[arcgis-adj] Touches returned 0 — falling back to Intersects", flush=True)
        features = _query_spatial("esriSpatialRelIntersects")
        spatial_rel_used = "esriSpatialRelIntersects"

    results = []
    for feat in features:
        a = feat.get("attributes", {})
        g = lambda c, _a=a: (_a.get(fld(c)) or "").strip() if isinstance(_a.get(fld(c)), str) else str(_a.get(fld(c)) or "")
        twp = g('township'); twp_dir = g('twp_dir')
        rng = g('range');    rng_dir = g('rng_dir')
        sec = g('section')
        trs = f"T{twp}{twp_dir} R{rng}{rng_dir}" if twp and rng else ""
        if trs and sec:
            trs += f" Sec {sec}"
        legal_raw = a.get(fld('legal')) or ""
        owner = g('owner')

        # ── Centroid distance filter ─────────────────────────────────────────
        adj_geom = feat.get("geometry", {})
        adj_rings = adj_geom.get("rings", [])
        if client_cx and adj_rings and adj_rings[0]:
            axs = [p[0] for p in adj_rings[0]]
            ays = [p[1] for p in adj_rings[0]]
            adj_cx = sum(axs) / len(axs)
            adj_cy = sum(ays) / len(ays)
            dlng = abs(adj_cx - client_cx) * 90000   # metres at ~36°N latitude
            dlat = abs(adj_cy - client_cy) * 111000
            dist_m = (dlng**2 + dlat**2) ** 0.5
            if dist_m > max_dist_m:
                print(f"[arcgis-adj]   SKIP {owner!r} — centroid {dist_m:.0f}m away (cap {max_dist_m:.0f}m)", flush=True)
                continue
        # ── Skip road/corridor parcels — very elongated geometry ──────────────
        if adj_rings and adj_rings[0] and len(adj_rings[0]) > 3:
            axs = [p[0] for p in adj_rings[0]]
            ays = [p[1] for p in adj_rings[0]]
            w_m = (max(axs) - min(axs)) * 90000
            h_m = (max(ays) - min(ays)) * 111000
            # Road parcels have aspect ratio > 10:1 and area < 0.5 acres
            short = min(w_m, h_m) + 1e-9
            long  = max(w_m, h_m)
            if long / short > 10 and (a.get(fld('area')) or 0) < 21780:  # 0.5 acres
                print(f"[arcgis-adj]   SKIP {owner!r} — likely road/corridor parcel", flush=True)
                continue

        results.append({
            "upc":         g('parcel_id'),
            "owner":       owner,
            "land_area":   a.get(fld('area')) or 0,
            "subdivision": g('subdivision'),
            "legal":       (str(legal_raw).strip())[:200],
            "address":     g('address_all'),
            "trs":         trs,
            "source":      "arcgis_spatial",
            "spatial_rel": spatial_rel_used,
        })

    print(f"[arcgis-adj] {len(results)} adjoiners after precision filtering", flush=True)
    return results


@app.route("/api/arcgis-adjoiners", methods=["POST"])
@require_auth
@require_pro
def api_arcgis_adjoiners():
    """Find adjacent parcels using ArcGIS spatial queries.

    Strategy:
      1. If UPC provided → fetch parcel geometry from ArcGIS
      2. If geometry provided directly (from KML index) → use that
      3. Query ArcGIS for all parcels touching that geometry
      4. Filter out the client's own parcel

    Body: { "upc": str, "geometry": { "rings": [...] } (optional),
            "client_name": str (optional, for filtering) }
    Returns: { success, adjoiners: [...], count, source }
    """
    try:
        data = request.get_json() or {}
        upc = (data.get("upc") or "").strip()
        geometry = data.get("geometry")  # Optional pre-supplied geometry
        client_name = (data.get("client_name") or "").strip().lower()

        # Step 1: Get geometry
        if not geometry and upc:
            print(f"[arcgis-adj] Fetching geometry for UPC {upc}...", flush=True)
            geometry = _arcgis_get_parcel_geometry(upc)

        if not geometry:
            # Try from local KML index as fallback
            survey = get_survey_data_path()
            if upc:
                polygon = xml_processor.extract_parcel_polygon(survey, upc)
                if polygon and polygon.get("coordinates"):
                    # Convert KML coords [[lng,lat], ...] to ArcGIS rings format
                    coords = polygon["coordinates"]
                    geometry = {
                        "rings": [[[c[0], c[1]] for c in coords]],
                        "spatialReference": {"wkid": 4326}
                    }
                    print(f"[arcgis-adj] Using KML geometry for UPC {upc}", flush=True)

        if not geometry:
            return jsonify({
                "success": False,
                "error": "Could not find parcel geometry. Try selecting the parcel on the map first.",
                "adjoiners": [], "count": 0,
            })

        # Step 2: Spatial query
        print("[arcgis-adj] Running spatial query for touching parcels...", flush=True)
        raw = _arcgis_find_touching_parcels(geometry)

        # Step 3: Filter out the client's own parcel
        adjoiners = []
        seen_upcs = set()
        for adj in raw:
            # Skip client's own parcel
            if upc and adj["upc"] == upc:
                continue
            if client_name and adj["owner"].lower() == client_name:
                continue
            # Deduplicate by UPC
            if adj["upc"] in seen_upcs:
                continue
            seen_upcs.add(adj["upc"])
            adjoiners.append(adj)

        print(f"[arcgis-adj] Found {len(adjoiners)} adjacent parcels", flush=True)

        return jsonify({
            "success":   True,
            "adjoiners": adjoiners,
            "count":     len(adjoiners),
            "source":    "arcgis_spatial",
            "client_upc": upc,
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e), "adjoiners": [], "count": 0})


# ══════════════════════════════════════════════════════════════════════════════
# DEEP DEED ANALYSIS & HEALTH CHECK
# ══════════════════════════════════════════════════════════════════════════════


def analyze_deed(detail: dict, pdf_path: str = "") -> dict:
    """Delegates to helpers.deed_analysis.analyze_deed."""
    return _analyze_deed_impl(detail, pdf_path, pdf_extractor=_extract_pdf_text)




@app.route("/api/extract-deed-description", methods=["POST"])
def api_extract_deed_description():
    """
    Extract the full property description from a deed PDF.

    Body: {
        pdf_path: str            – Path to the saved deed PDF
        detail:   dict (optional) – Scraped metadata from the website (fallback)
        doc_no:   str  (optional) – Document number for the PDF URL fallback
    }

    Returns: {
        success: bool,
        description: {
            full_text:        str   – Complete text extracted from the PDF
            legal_description: str  – Isolated legal description section
            source:           str   – 'text' or 'ocr'
            trs_refs:         list  – Township/Range/Section references
            calls_count:      int   – Number of metes-and-bounds calls found
            calls:            list  – Parsed bearing/distance calls
            desc_type:        str   – 'metes_and_bounds', 'lot_block', 'tract', 'trs_only', 'unknown'
            grantor:          str   – Extracted grantor name
            grantee:          str   – Extracted grantee name
            area_acres:       float – Computed area if metes-and-bounds
            perimeter_ft:     float – Computed perimeter if metes-and-bounds
            monuments:        list  – Monument types referenced
            adjoiners:        list  – Adjoiner names found in description
            pob_found:        bool  – Whether Point of Beginning was found
            cab_refs:         list  – Cabinet references found
        }
    }
    """
    try:
        data     = request.get_json(silent=True) or {}
        pdf_path = data.get("pdf_path", "")
        detail   = data.get("detail", {})
        doc_no   = data.get("doc_no", "")

        if isinstance(detail, str):
            try:
                detail = json.loads(detail)
            except Exception:
                detail = {}

        full_text = ""
        source    = "none"

        # ── 1. Extract text from saved PDF ────────────────────────────────
        if pdf_path and os.path.isfile(pdf_path):
            print(f"[extract-desc] Reading PDF: {pdf_path}", flush=True)
            full_text, source = _extract_pdf_text(pdf_path)
            print(f"[extract-desc] source={source} chars={len(full_text.strip())}", flush=True)

        # ── 2. Fallback: if no PDF, use URL or metadata only ──────────────
        if not full_text.strip() and doc_no:
            # Try to download the PDF from the online source into a temp file
            try:
                pdf_url  = f"{_get_portal_url()}/WebTemp/{doc_no}.pdf"
                pdf_resp = _session().get(pdf_url, stream=True, timeout=20)
                if pdf_resp.status_code == 200:
                    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tf:
                        for chunk in pdf_resp.iter_content(8192):
                            tf.write(chunk)
                        tmp_path = tf.name
                    full_text, source = _extract_pdf_text(tmp_path)
                    try:
                        os.unlink(tmp_path)
                    except Exception:
                        pass
                    print(f"[extract-desc] Downloaded PDF from URL, source={source} chars={len(full_text.strip())}", flush=True)
            except Exception as e:
                print(f"[extract-desc] PDF URL fallback failed: {e}", flush=True)

        # ── 3. Merge with scraped metadata fields ─────────────────────────
        metadata_text = ""
        legal_fields = [
            "Legal Description", "Legal", "Other Legal", "Other_Legal",
            "Subdivision Legal", "Subdivision_Legal", "Description",
            "Comments", "Remarks", "Reference",
        ]
        for field in legal_fields:
            val = detail.get(field, "")
            if val:
                metadata_text += val + "\n"

        combined = full_text + "\n" + metadata_text

        # ── 4. Isolate the legal description section ──────────────────────
        legal_desc = _isolate_legal_description(combined)

        # ── 5. Parse structured data from the text ────────────────────────
        trs_refs = extract_trs(combined)
        calls    = parse_metes_bounds(combined)

        # Description type
        desc_type = classify_description_type(combined, calls, trs_refs)

        # Area / perimeter if metes-and-bounds
        full      = calls_to_full_coords(calls) if calls else {}
        pts       = full.get("boundary_coords", []) if full else []
        tie_pts   = full.get("tie_coords", []) if full else []
        has_tie   = full.get("has_tie", False) if full else False
        bdy_calls = [c for c in calls if not c.get("tie_call")]
        perimeter = sum(c.get("distance", 0) for c in bdy_calls)
        area_sqft = shoelace_area(pts)

        # Monuments
        monuments = detect_monuments(combined)

        # Adjoiner names
        adjoiners = []
        if detail:
            adj_items = parse_adjoiner_names({"Legal": combined})
            adjoiners = [a["name"] for a in adj_items]

        # POB
        pob_found = has_pob(combined)

        # Cabinet refs
        cab_refs = parse_cabinet_refs({"Legal": combined, **detail})

        # Grantor / Grantee from detail or PDF text
        grantor = detail.get("Grantor", "") or detail.get("grantor", "")
        grantee = detail.get("Grantee", "") or detail.get("grantee", "")

        # Format calls for frontend
        calls_formatted = []
        for c in calls:
            if c.get("type") == "straight":
                calls_formatted.append({
                    "bearing": c.get("bearing_label", ""),
                    "distance": c.get("distance", 0),
                    "raw": c.get("bearing_raw", ""),
                    "azimuth": c.get("azimuth_deg", 0),
                    "unit": c.get("unit", "ft"),
                    "raw_distance": c.get("raw_distance", c.get("distance", 0)),
                })
            elif c.get("type") == "curve":
                calls_formatted.append({
                    "bearing": f"Curve {c.get('direction', '').title()} R={c.get('radius', 0):.1f}'",
                    "distance": round(c.get("arc_length", 0) or c.get("chord_length", 0), 2),
                    "raw": c.get("bearing_raw", ""),
                    "curve": True,
                })

        # ── 6. Closure error calculation ──────────────────────────────────
        closure_err = 0.0
        closure_ratio = ""
        if len(pts) >= 2:
            closure_err = round(
                math.hypot(pts[-1][0] - pts[0][0], pts[-1][1] - pts[0][1]), 4
            )
            if closure_err > 0.001 and perimeter > 0:
                ratio = perimeter / closure_err
                closure_ratio = f"1:{int(ratio)}"

        # Coordinates for the frontend boundary plotter (list of [x, y])
        coords     = [[round(p[0], 4), round(p[1], 4)] for p in pts]
        tie_coords = [[round(p[0], 4), round(p[1], 4)] for p in tie_pts]

        return jsonify({
            "success": True,
            "description": {
                "full_text":         full_text.strip()[:10000],  # Cap to avoid huge payloads
                "legal_description": legal_desc.strip()[:5000],
                "source":            source,
                "trs_refs":          [t["trs"] for t in trs_refs],
                "calls_count":       len(bdy_calls),
                "tie_calls_count":   len(calls) - len(bdy_calls),
                "has_tie":           has_tie,
                "calls":             calls_formatted[:100],
                "coords":            coords,
                "tie_coords":        tie_coords,
                "closure_err":       closure_err,
                "closure_ratio":     closure_ratio,
                "desc_type":         desc_type,
                "grantor":           grantor,
                "grantee":           grantee,
                "area_acres":        round(area_sqft / 43560.0, 3) if area_sqft else 0,
                "perimeter_ft":      round(perimeter, 2),
                "monuments":         monuments,
                "adjoiners":         adjoiners[:30],
                "pob_found":         pob_found,
                "cab_refs":          [{"cabinet": r["cabinet"], "doc": r["doc"]} for r in cab_refs],
            }
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


def _isolate_legal_description(text: str) -> str:
    """Delegates to helpers.deed_analysis.isolate_legal_description."""
    return _isolate_legal_description_impl(text)


@app.route("/api/analyze-deed", methods=["POST"])
@require_auth
@require_pro
def api_analyze_deed():
    """
    Deep deed analysis endpoint.
    Body: { detail: dict, pdf_path: str (optional) }
    Returns: { success, analysis: {...} }
    """
    try:
        data     = request.get_json(silent=True) or {}
        detail   = data.get("detail", {})
        pdf_path = data.get("pdf_path", "")

        if isinstance(detail, str):
            try:
                detail = json.loads(detail)
            except Exception:
                detail = {}

        result = analyze_deed(detail, pdf_path)
        return jsonify({"success": True, "analysis": result})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ══════════════════════════════════════════════════════════════════════════════
# ADJACENT PARCELS (standalone endpoint)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/adjacent-parcels", methods=["POST"])
def api_adjacent_parcels():
    """Find parcels geometrically adjacent to a given parcel.

    Body: { upc: str, max_results: int (default 20), threshold: float (default 0.0003) }
    Returns: { success, parcels: [...], count: int }
    """
    try:
        data = request.get_json(silent=True) or {}
        upc = (data.get("upc") or "").strip()
        if not upc:
            return jsonify({"success": False, "error": "upc is required"})

        survey = get_survey_data_path()
        if not survey:
            return jsonify({"success": False, "error": "Survey drive not found"})

        max_r = int(data.get("max_results", 20))
        threshold = float(data.get("threshold", 0.0003))

        results = xml_processor.find_adjacent_parcels(
            survey, upc, max_results=max_r, edge_threshold_deg=threshold
        )

        # Strip heavy polygon data from response to keep it lean
        clean = []
        for p in results:
            clean.append({
                "upc":            p.get("upc", ""),
                "owner":          p.get("owner", ""),
                "book":           p.get("book", ""),
                "page":           p.get("page", ""),
                "plat":           p.get("plat", ""),
                "trs":            p.get("trs", "") or (p.get("arcgis", {}) or {}).get("trs", ""),
                "cab_refs":       p.get("cab_refs", []),
                "centroid":       p.get("centroid"),
                "adjacency_dist": p.get("_adjacency_dist"),
                "adjacency_type": p.get("_adjacency_type", "edge"),
            })

        return jsonify({"success": True, "parcels": clean, "count": len(clean)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ══════════════════════════════════════════════════════════════════════════════
# LEGAL DESCRIPTION SIMILARITY SEARCH
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/similar-descriptions", methods=["POST"])
@app.route("/api/legal-similarity", methods=["POST"])
def api_legal_similarity():
    """Search for parcels with similar legal descriptions.

    Body: { text: str, min_score: float (default 20), limit: int (default 20) }
    Returns: { success, results: [...], count: int }
    """
    try:
        data = request.get_json(silent=True) or {}
        text = (data.get("text") or "").strip()
        if not text:
            return jsonify({"success": False, "error": "text is required"})

        survey = get_survey_data_path()
        if not survey:
            return jsonify({"success": False, "error": "Survey drive not found"})

        idx = xml_processor.load_index(survey)
        if not idx:
            return jsonify({"success": False, "error": "No parcel index found"})

        min_score = float(data.get("min_score", 20))
        limit = int(data.get("limit", 20))

        results = _search_similar_descriptions(
            text, idx.get("parcels", []),
            min_score=min_score, limit=limit
        )

        return jsonify({"success": True, "results": results, "count": len(results)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})



# ══════════════════════════════════════════════════════════════════════════════
# AUTONOMOUS JOB PIPELINE  (Website → Deed Helper → AI Surveyor)
# ══════════════════════════════════════════════════════════════════════════════
#
# Phase 1: POST /api/inquiry         — accept survey requests
# Phase 2: POST /api/auto-research   — autonomous 6-step research
#          GET  /api/inquiry/status   — track progress
# ══════════════════════════════════════════════════════════════════════════════

import threading
_INQUIRIES_PATH = Path("data/inquiries.json")

def _load_inquiries() -> list:
    if _INQUIRIES_PATH.exists():
        try:
            return json.loads(_INQUIRIES_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return []

def _save_inquiries(data: list):
    _INQUIRIES_PATH.parent.mkdir(parents=True, exist_ok=True)
    _INQUIRIES_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

def _find_inquiry(inquiry_id: str) -> tuple:
    """Return (inquiry_dict, index) or (None, -1)."""
    inquiries = _load_inquiries()
    for i, inq in enumerate(inquiries):
        if inq.get("id") == inquiry_id:
            return inq, i
    return None, -1

def _update_inquiry(inquiry_id: str, updates: dict):
    """Merge updates into an inquiry and persist."""
    inquiries = _load_inquiries()
    for inq in inquiries:
        if inq.get("id") == inquiry_id:
            inq.update(updates)
            break
    _save_inquiries(inquiries)


@app.route("/api/inquiry", methods=["POST"])
def api_inquiry():
    """Accept a survey inquiry from the website.

    Body: {
        client_name: str,     # "Garza, Veronica"
        address:     str,     # Optional street address
        job_type:    str,     # "BDY", "SUB", etc.
        email:       str,
        phone:       str,
        notes:       str,
        upc:         str,     # Optional parcel ID
    }

    Returns: { success, inquiry_id, estimated_complexity }
    """
    try:
        data = request.get_json(silent=True) or {}
        client_name = (data.get("client_name") or "").strip()
        if not client_name:
            return jsonify({"success": False, "error": "Client name is required"}), 400

        job_type = data.get("job_type", "BDY").upper()
        address  = (data.get("address") or "").strip()
        email    = (data.get("email") or "").strip()
        phone    = (data.get("phone") or "").strip()
        notes    = (data.get("notes") or "").strip()
        upc      = (data.get("upc") or "").strip()

        # Generate inquiry ID
        inquiries = _load_inquiries()
        seq = len(inquiries) + 3001
        inquiry_id = f"INQ-{seq}"

        # ── ML complexity prediction (optional — fails gracefully) ──
        complexity = {}
        try:
            from ai import get_predictor
            predictor = get_predictor()
            if predictor:
                complexity = predictor.predict_complexity(job_type, client_name)
        except Exception:
            pass

        # ── Knowledge Graph lookup (optional — fails gracefully) ──
        kg_info = {}
        try:
            from ai import get_knowledge_graph
            kg = get_knowledge_graph()
            if kg:
                adjoiners = kg.get_adjoiners(client_name)
                jobs = kg.get_person_jobs(client_name)
                kg_info = {
                    "known_adjoiners": [a.get("name", a) if isinstance(a, dict) else a for a in adjoiners[:10]],
                    "past_jobs": len(jobs),
                    "in_knowledge_graph": True,
                }
        except Exception:
            kg_info = {"in_knowledge_graph": False}

        # ── Build inquiry record ──
        inquiry = {
            "id":           inquiry_id,
            "client_name":  client_name,
            "address":      address,
            "job_type":     job_type,
            "email":        email,
            "phone":        phone,
            "notes":        notes,
            "upc":          upc,
            "status":       "pending",
            "step":         0,
            "created_at":   datetime.now().isoformat(),
            "updated_at":   datetime.now().isoformat(),
            "complexity":   complexity,
            "kg_info":      kg_info,
            "research":     {},  # populated by auto-research
            "log":          [f"Inquiry created at {datetime.now().isoformat()}"],
        }

        inquiries.append(inquiry)
        _save_inquiries(inquiries)

        print(f"[inquiry] Created {inquiry_id} for {client_name} ({job_type})", flush=True)

        return jsonify({
            "success":              True,
            "inquiry_id":           inquiry_id,
            "estimated_complexity": complexity,
            "kg_info":              kg_info,
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/inquiry/list", methods=["GET"])
def api_inquiry_list():
    """List all inquiries (for dashboard/admin)."""
    inquiries = _load_inquiries()
    # Return summaries, not full detail
    summaries = []
    for inq in reversed(inquiries):  # newest first
        summaries.append({
            "id":          inq["id"],
            "client_name": inq["client_name"],
            "job_type":    inq["job_type"],
            "status":      inq["status"],
            "step":        inq.get("step", 0),
            "created_at":  inq.get("created_at", ""),
            "updated_at":  inq.get("updated_at", ""),
        })
    return jsonify({"success": True, "inquiries": summaries, "count": len(summaries)})


@app.route("/api/inquiry/status/<inquiry_id>", methods=["GET"])
def api_inquiry_status(inquiry_id):
    """Get the current status and progress of an inquiry."""
    inq, _ = _find_inquiry(inquiry_id)
    if not inq:
        return jsonify({"success": False, "error": "Inquiry not found"}), 404

    # Build progress summary
    research = inq.get("research", {})
    subjects = research.get("subjects", [])
    deeds_saved = sum(1 for s in subjects if s.get("deed_saved"))
    plats_saved = sum(1 for s in subjects if s.get("plat_saved"))

    return jsonify({
        "success":         True,
        "inquiry_id":      inquiry_id,
        "status":          inq["status"],
        "step":            inq.get("step", 0),
        "client_name":     inq["client_name"],
        "job_type":        inq["job_type"],
        "subjects_count":  len(subjects),
        "deeds_saved":     deeds_saved,
        "plats_saved":     plats_saved,
        "job_number":      research.get("job_number", ""),
        "log":             inq.get("log", [])[-20:],  # last 20 log entries
        "complexity":      inq.get("complexity", {}),
        "created_at":      inq.get("created_at", ""),
        "updated_at":      inq.get("updated_at", ""),
    })


# ──────────────────────────────────────────────────────────────────────────────
# PHASE 2: AUTONOMOUS RESEARCH ENGINE
# ──────────────────────────────────────────────────────────────────────────────

@app.route("/api/auto-research", methods=["POST"])
def api_auto_research():
    """Trigger autonomous 6-step research for an inquiry or ad-hoc request.

    Body: {
        inquiry_id:  str (optional — links to existing inquiry),
        client_name: str (required if no inquiry_id),
        job_type:    str (default "BDY"),
        upc:         str (optional — enables spatial adjoiner discovery),
    }

    Runs the research in a background thread and returns immediately.
    Poll /api/inquiry/status/<id> for progress.
    """
    try:
        data = request.get_json(silent=True) or {}
        inquiry_id  = data.get("inquiry_id", "")
        client_name = (data.get("client_name") or "").strip()
        job_type    = data.get("job_type", "BDY").upper()
        upc         = (data.get("upc") or "").strip()

        # If linked to inquiry, use its data
        if inquiry_id:
            inq, _ = _find_inquiry(inquiry_id)
            if not inq:
                return jsonify({"success": False, "error": "Inquiry not found"}), 404
            client_name = client_name or inq["client_name"]
            job_type    = job_type or inq["job_type"]
            upc         = upc or inq.get("upc", "")

        if not client_name:
            return jsonify({"success": False, "error": "client_name is required"}), 400

        # Create inquiry if none exists
        if not inquiry_id:
            inquiries = _load_inquiries()
            seq = len(inquiries) + 3001
            inquiry_id = f"INQ-{seq}"
            inquiries.append({
                "id": inquiry_id, "client_name": client_name,
                "job_type": job_type, "upc": upc, "status": "starting",
                "step": 0, "address": "", "email": "", "phone": "",
                "notes": "Auto-research triggered",
                "created_at": datetime.now().isoformat(),
                "updated_at": datetime.now().isoformat(),
                "complexity": {}, "kg_info": {}, "research": {},
                "log": [f"Auto-research triggered at {datetime.now().isoformat()}"],
            })
            _save_inquiries(inquiries)

        _update_inquiry(inquiry_id, {
            "status": "starting",
            "updated_at": datetime.now().isoformat(),
        })

        # Launch in background thread
        t = threading.Thread(
            target=_auto_research_worker,
            args=(inquiry_id, client_name, job_type, upc),
            daemon=True,
        )
        t.start()

        print(f"[auto-research] Started background worker for {inquiry_id}: {client_name}", flush=True)
        return jsonify({
            "success":    True,
            "inquiry_id": inquiry_id,
            "message":    f"Auto-research started for {client_name}. Poll /api/inquiry/status/{inquiry_id} for progress.",
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# ── JOB TYPE WORKFLOW CONFIGURATIONS ──────────────────────────────────────────
# Each job type defines how the auto-research engine behaves.
# Team: edit these to match your actual processes.

_JOB_WORKFLOWS = {
    # ── Boundary Surveys ──────────────────────────────────────────────────
    "BDY": {
        "name": "Boundary Survey",
        "research_depth": "full",         # full | medium | minimal
        "needs_adjoiners": "all",         # all | perimeter | affected | none
        "needs_chain": False,             # deep chain-of-title search
        "auto_select_confidence": 25,     # relevance score threshold
        "search_as_grantee_too": True,    # also search grantee index
        "deliverables": ["boundary_plat", "legal_description", "reference_table", "dxf"],
        "notes": "Full research — client deed, plat, ALL adjoiners",
    },
    "BDY-E": {
        "name": "Boundary — Existing Corners",
        "research_depth": "full",
        "needs_adjoiners": "all",
        "needs_chain": False,
        "auto_select_confidence": 25,
        "search_as_grantee_too": True,
        "deliverables": ["boundary_plat", "legal_description", "reference_table", "dxf"],
        "notes": "Retracement — locate existing monuments. Check plats for monument descriptions.",
    },
    "BDY-N": {
        "name": "Boundary — New Corners",
        "research_depth": "full",
        "needs_adjoiners": "all",
        "needs_chain": True,              # deep chain to understand boundary history
        "auto_select_confidence": 20,     # more aggressive — want all deeds
        "search_as_grantee_too": True,
        "deliverables": ["boundary_plat", "legal_description", "reference_table", "dxf"],
        "notes": "Original survey — set new monuments. Need full chain and BLM/GLO check.",
    },

    # ── Subdivision ───────────────────────────────────────────────────────
    "SUB": {
        "name": "Subdivision",
        "research_depth": "full",
        "needs_adjoiners": "perimeter",   # only surrounding parcels
        "needs_chain": False,
        "auto_select_confidence": 25,
        "search_as_grantee_too": True,
        "deliverables": ["subdivision_plat", "legal_descriptions", "reference_table"],
        "notes": "Parent parcel being divided. Research perimeter adjoiners only.",
    },

    # ── Lot Line Adjustment ──────────────────────────────────────────────
    "LLA": {
        "name": "Lot Line Adjustment",
        "research_depth": "medium",
        "needs_adjoiners": "affected",    # only parcels touching the adjusted line
        "needs_chain": False,
        "auto_select_confidence": 25,
        "search_as_grantee_too": True,
        "deliverables": ["lla_plat", "legal_descriptions", "reference_table"],
        "notes": "Two parcels involved. Research both parcels + affected adjoiners.",
    },

    # ── Land Consolidation ───────────────────────────────────────────────
    "LC": {
        "name": "Land Consolidation",
        "research_depth": "medium",
        "needs_adjoiners": "perimeter",
        "needs_chain": False,
        "auto_select_confidence": 25,
        "search_as_grantee_too": True,
        "deliverables": ["consolidation_plat", "legal_description", "reference_table"],
        "notes": "Merging parcels. Perimeter adjoiners only.",
    },

    # ── FEMA Topo ────────────────────────────────────────────────────────
    "FT": {
        "name": "FEMA Topographic Survey",
        "research_depth": "minimal",
        "needs_adjoiners": "none",        # NO adjoiners for topo
        "needs_chain": False,
        "auto_select_confidence": 30,
        "search_as_grantee_too": False,
        "deliverables": ["elevation_certificate", "topo_dwg"],
        "notes": "Deed for legal description only. No adjoiners. No plat required.",
    },

    # ── Easement ─────────────────────────────────────────────────────────
    "EAS": {
        "name": "Easement Survey",
        "research_depth": "medium",
        "needs_adjoiners": "affected",    # parcels involved in easement
        "needs_chain": False,
        "auto_select_confidence": 25,
        "search_as_grantee_too": True,
        "deliverables": ["easement_exhibit", "legal_description"],
        "notes": "Search both servient and dominant parcels. Find easement document.",
    },

    # ── Permit Survey ────────────────────────────────────────────────────
    "P": {
        "name": "Permit Survey",
        "research_depth": "minimal",
        "needs_adjoiners": "none",        # NO adjoiners for permit
        "needs_chain": False,
        "auto_select_confidence": 30,
        "search_as_grantee_too": False,
        "deliverables": ["site_plan"],
        "notes": "Deed for legal desc only. No adjoiners. No plat needed.",
    },

    # ── Survey Exam (Desktop Review) ─────────────────────────────────────
    "SE": {
        "name": "Survey Exam",
        "research_depth": "maximum",      # MAXIMUM — this IS the job
        "needs_adjoiners": "all",
        "needs_chain": True,              # deep chain of title
        "auto_select_confidence": 15,     # very aggressive — grab everything
        "search_as_grantee_too": True,
        "deliverables": ["written_report", "reference_table"],
        "notes": "Desktop research is the entire job. Pull everything — deep chain, all adjoiners, similarity search.",
    },

    # ── Improvement Location Report ──────────────────────────────────────
    "ILR": {
        "name": "Improvement Location Report",
        "research_depth": "minimal",
        "needs_adjoiners": "none",
        "needs_chain": False,
        "auto_select_confidence": 30,
        "search_as_grantee_too": False,
        "deliverables": ["ilr_certificate"],
        "notes": "Deed for legal desc only. Measure structure setbacks.",
    },

    # ── Lot Survey ───────────────────────────────────────────────────────
    "LS": {
        "name": "Lot Survey",
        "research_depth": "medium",
        "needs_adjoiners": "affected",
        "needs_chain": False,
        "auto_select_confidence": 25,
        "search_as_grantee_too": True,
        "deliverables": ["lot_plat", "reference_table"],
        "notes": "Single lot in subdivision. Check subdivision plat + immediate adjoiners.",
    },

    # ── Water Rights ─────────────────────────────────────────────────────
    "WR": {
        "name": "Water Rights Survey",
        "research_depth": "minimal",
        "needs_adjoiners": "none",
        "needs_chain": False,
        "auto_select_confidence": 30,
        "search_as_grantee_too": False,
        "deliverables": ["ditch_exhibit", "legal_description"],
        "notes": "Map acequia/ditch. Deed for property description only.",
    },
}

def _get_workflow(job_type: str) -> dict:
    """Get workflow config for a job type. Falls back to BDY for unknown types."""
    # Try exact match, then base code (e.g., "BDY-E" → "BDY")
    wf = _JOB_WORKFLOWS.get(job_type.upper())
    if not wf:
        base = job_type.upper().split("-")[0]
        wf = _JOB_WORKFLOWS.get(base, _JOB_WORKFLOWS["BDY"])
    return wf


@app.route("/api/workflows", methods=["GET"])
def api_list_workflows():
    """Return all available job type workflows for UI display."""
    result = []
    for code, wf in _JOB_WORKFLOWS.items():
        result.append({
            "code": code,
            "name": wf["name"],
            "research_depth": wf["research_depth"],
            "needs_adjoiners": wf["needs_adjoiners"],
            "needs_chain": wf.get("needs_chain", False),
            "deliverables": wf.get("deliverables", []),
            "notes": wf.get("notes", ""),
        })
    return jsonify({"success": True, "workflows": result})

def _auto_research_worker(inquiry_id: str, client_name: str, job_type: str, upc: str):
    """Background worker: runs the autonomous research pipeline.

    Adapts behavior based on job type workflow configuration:
    - BDY/SUB/SE: full adjoiners + deep research
    - FT/P/ILR:   deed only, skip adjoiners
    - LLA/EAS:    medium depth, affected adjoiners only
    """
    def _log(msg):
        print(f"[auto-research:{inquiry_id}] {msg}", flush=True)
        inq, _ = _find_inquiry(inquiry_id)
        if inq:
            log_list = inq.get("log", [])
            log_list.append(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")
            _update_inquiry(inquiry_id, {"log": log_list, "updated_at": datetime.now().isoformat()})

    try:
        with app.app_context():
            wf = _get_workflow(job_type)
            _log(f"Starting autonomous research pipeline...")
            _log(f"  Workflow: {wf['name']} ({job_type})")
            _log(f"  Research depth: {wf['research_depth']}, Adjoiners: {wf['needs_adjoiners']}")

            # ── Auto-login to 1stNMTitle portal ─────────────────────────────
            _log("Step 0: Logging into 1stNMTitle portal...")
            login_ok = _auto_login()
            if login_ok:
                _log("  ✓ Portal session established")
            else:
                _log("  ⚠ Portal login failed — deed searches may not work")
                _log("    Check config.json: firstnm_user / firstnm_pass")

            # ── STEP 1: Setup — create project and session ──────────────────
            _update_inquiry(inquiry_id, {"status": "researching", "step": 1})
            _log("Step 1: Creating project and session...")

            next_num, _ = next_job_info()
            job_number = str(next_num)
            create_project_folders(job_number, client_name, job_type)
            _log(f"  Project #{job_number} created")

            # Create research session
            session_data = load_research(job_number, client_name, job_type)
            session_data["job_number"] = job_number
            session_data["client_name"] = client_name
            session_data["job_type"] = job_type
            session_data["client_upc"] = upc
            session_data["workflow"] = {
                "code": job_type,
                "name": wf["name"],
                "research_depth": wf["research_depth"],
                "needs_adjoiners": wf["needs_adjoiners"],
                "needs_chain": wf.get("needs_chain", False),
                "deliverables": wf.get("deliverables", []),
                "notes": wf.get("notes", ""),
            }
            save_research(job_number, client_name, job_type, session_data)

            _update_inquiry(inquiry_id, {
                "research": {"job_number": job_number, "subjects": session_data["subjects"]},
            })
            _log("  Session created")

            # ── STEP 2: Client Deed Search + Auto-Select ────────────────────
            _update_inquiry(inquiry_id, {"step": 2})
            _log("Step 2: Searching for client deed...")

            deed_result = _auto_search_deed(client_name, job_type)
            if deed_result:
                _log(f"  Found {deed_result['count']} results, best: {deed_result['best']['doc_no']} (score: {deed_result['best'].get('relevance_score', '?')})")

                # Auto-save if confidence is high enough
                best = deed_result["best"]
                score = best.get("relevance_score", 0)

                if score >= wf['auto_select_confidence'] or deed_result["count"] == 1:
                    save_result = _auto_save_deed(
                        best["doc_no"], best, job_number, client_name,
                        job_type, "client", is_adjoiner=False,
                    )
                    if save_result.get("success"):
                        # Update session
                        session_data = load_research(job_number, client_name, job_type)
                        for s in session_data["subjects"]:
                            if s["id"] == "client":
                                s["deed_saved"] = True
                                s["deed_path"] = save_result.get("saved_to", "")
                                s["doc_no"] = best["doc_no"]
                                s["detail"] = best
                                break
                        save_research(job_number, client_name, job_type, session_data)
                        _log(f"  ✓ Client deed saved: {best['doc_no']}")
                    else:
                        _log(f"  ✗ Deed save failed: {save_result.get('error', 'unknown')}")
                else:
                    _log(f"  ⚠ Low confidence ({score}) — skipping auto-select, needs human review")
            else:
                _log("  No deed results found")

            # ── STEP 3: Client Plat Search ──────────────────────────────────
            _update_inquiry(inquiry_id, {"step": 3})
            _log("Step 3: Searching for client plat...")

            if wf['research_depth'] != 'minimal':
                plat_result = _auto_search_plat(client_name, job_number, job_type)
                if plat_result:
                    _log(f"  Found plat: {plat_result.get('source', 'unknown')}")
                    session_data = load_research(job_number, client_name, job_type)
                    for s in session_data["subjects"]:
                        if s["id"] == "client":
                            s["plat_saved"] = True
                            s["plat_path"] = plat_result.get("saved_to", "")
                            break
                    save_research(job_number, client_name, job_type, session_data)
                    _log("  ✓ Client plat saved")
                else:
                    _log("  No plat found (may need manual search)")
            else:
                _log("  Skipped — not required for this job type")

            # ── STEP 4: Adjoiner Discovery ──────────────────────────────────
            _update_inquiry(inquiry_id, {"step": 4})

            adj_mode = wf['needs_adjoiners']
            if adj_mode == 'none':
                _log(f"Step 4: Adjoiners — SKIPPED ({wf['name']} does not require adjoiners)")
                adjoiners = []
            else:
                _log(f"Step 4: Discovering adjoiners (mode: {adj_mode})...")
                adjoiners = []
                if upc:
                    adjoiners = _auto_discover_adjoiners(upc, client_name)
                    _log(f"  Found {len(adjoiners)} adjoiners via ArcGIS")
                else:
                    _log("  No UPC — skipping spatial discovery (add adjoiners manually)")

            # Add adjoiners to session
            session_data = load_research(job_number, client_name, job_type)
            for adj in adjoiners:
                adj_name = adj.get("owner", "Unknown")
                if adj_name.upper() == client_name.upper():
                    continue
                adj_id = f"adj_{adj.get('upc', '')[:10]}_{len(session_data['subjects'])}"
                session_data["subjects"].append({
                    "id": adj_id, "type": "adjoiner", "name": adj_name,
                    "deed_saved": False, "plat_saved": False,
                    "status": "pending", "notes": "",
                    "deed_path": "", "plat_path": "",
                    "upc": adj.get("upc", ""),
                })
            save_research(job_number, client_name, job_type, session_data)

            _update_inquiry(inquiry_id, {
                "research": {"job_number": job_number, "subjects": session_data["subjects"]},
            })

            # ── STEP 5: Bulk Adjoiner Research ──────────────────────────────
            _update_inquiry(inquiry_id, {"step": 5})

            if adj_mode == 'none':
                _log(f"Step 5: Adjoiner research — SKIPPED")
            else:
                _log(f"Step 5: Researching {len(adjoiners)} adjoiners...")

            session_data = load_research(job_number, client_name, job_type)
            adj_subjects = [s for s in session_data["subjects"] if s["type"] == "adjoiner"]

            for i, subj in enumerate(adj_subjects):
                _log(f"  [{i+1}/{len(adj_subjects)}] Searching for {subj['name']}...")

                # Search deed
                deed_r = _auto_search_deed(subj["name"], job_type)
                if deed_r and deed_r["best"]:
                    best = deed_r["best"]
                    save_r = _auto_save_deed(
                        best["doc_no"], best, job_number, client_name,
                        job_type, subj["id"], is_adjoiner=True,
                        adjoiner_name=subj["name"],
                    )
                    if save_r.get("success"):
                        subj["deed_saved"] = True
                        subj["deed_path"] = save_r.get("saved_to", "")
                        subj["doc_no"] = best["doc_no"]
                        subj["detail"] = best
                        _log(f"    ✓ Deed saved: {best['doc_no']}")
                    else:
                        _log(f"    ✗ Deed save failed")
                else:
                    _log(f"    No deed found")

                # Search plat
                plat_r = _auto_search_plat(subj["name"], job_number, job_type, is_adjoiner=True)
                if plat_r:
                    subj["plat_saved"] = True
                    subj["plat_path"] = plat_r.get("saved_to", "")
                    _log(f"    ✓ Plat saved")

                # Brief pause to be polite to the portal
                import time
                time.sleep(0.5)

            save_research(job_number, client_name, job_type, session_data)
            _update_inquiry(inquiry_id, {
                "research": {"job_number": job_number, "subjects": session_data["subjects"]},
            })

            # ── STEP 6: Package ─────────────────────────────────────────────
            _update_inquiry(inquiry_id, {"step": 6})
            _log("Step 6: Generating reference table and package...")

            # Build reference summary
            session_data = load_research(job_number, client_name, job_type)
            ref_table = []
            deed_num = 0
            plat_num = 0
            for subj in session_data["subjects"]:
                if subj.get("deed_saved"):
                    deed_num += 1
                    ref_table.append({
                        "type": "Deed", "ref": f"D{deed_num}", "owner": subj["name"],
                        "doc_no": subj.get("doc_no", ""),
                        "book_page": subj.get("detail", {}).get("location", ""),
                        "grantor": subj.get("detail", {}).get("grantor", ""),
                        "grantee": subj.get("detail", {}).get("grantee", ""),
                        "date": subj.get("detail", {}).get("date", ""),
                        "relationship": "Client" if subj["type"] == "client" else "Adjoiner",
                    })
                if subj.get("plat_saved"):
                    plat_num += 1
                    plat_fname = os.path.basename(subj.get("plat_path", ""))
                    ref_table.append({
                        "type": "Plat", "ref": f"P{plat_num}", "owner": subj["name"],
                        "filename": plat_fname,
                        "path": subj.get("plat_path", ""),
                        "relationship": "Client" if subj["type"] == "client" else "Adjoiner",
                    })

            session_data["reference_table"] = ref_table
            session_data["auto_research_complete"] = True
            session_data["deliverables"] = wf.get("deliverables", [])

            # ── Generate Excel Reference Table (.xlsx) ─────────────────────
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Reference Table"

                # Title row
                ws.merge_cells("A1:G1")
                ws["A1"] = f"DOCUMENTS REFERENCED ON THIS PLAT — Job #{job_number} {client_name}"
                ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="2D8A6E")
                ws["A1"].alignment = Alignment(horizontal="center")

                # Deed header
                deed_header = ["Ref #", "Type", "Book/Page", "Grantor", "Grantee", "Date", "Relationship"]
                for i, h in enumerate(deed_header, 1):
                    c = ws.cell(row=3, column=i, value=h)
                    c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                    c.fill = PatternFill(start_color="2D8A6E", end_color="2D8A6E", fill_type="solid")
                    c.alignment = Alignment(horizontal="center")
                    c.border = Border(
                        bottom=Side(style="thin"),
                        top=Side(style="thin"),
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                    )

                # Deed rows
                row_idx = 4
                for entry in ref_table:
                    if entry["type"] == "Deed":
                        ws.cell(row=row_idx, column=1, value=entry.get("ref", ""))
                        ws.cell(row=row_idx, column=2, value="Deed")
                        ws.cell(row=row_idx, column=3, value=entry.get("book_page", ""))
                        ws.cell(row=row_idx, column=4, value=entry.get("grantor", ""))
                        ws.cell(row=row_idx, column=5, value=entry.get("grantee", ""))
                        ws.cell(row=row_idx, column=6, value=entry.get("date", ""))
                        ws.cell(row=row_idx, column=7, value=entry.get("relationship", ""))
                        for col in range(1, 8):
                            ws.cell(row=row_idx, column=col).font = Font(name="Calibri", size=10)
                            ws.cell(row=row_idx, column=col).border = Border(
                                bottom=Side(style="hair"), left=Side(style="hair"), right=Side(style="hair"),
                            )
                        row_idx += 1

                # Spacer
                row_idx += 1

                # Plat header
                plat_header = ["Ref #", "Type", "Cabinet/File", "Owner", "Surveyor", "Date", "Relationship"]
                for i, h in enumerate(plat_header, 1):
                    c = ws.cell(row=row_idx, column=i, value=h)
                    c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                    c.fill = PatternFill(start_color="4FACFE", end_color="4FACFE", fill_type="solid")
                    c.alignment = Alignment(horizontal="center")
                    c.border = Border(
                        bottom=Side(style="thin"), top=Side(style="thin"),
                        left=Side(style="thin"), right=Side(style="thin"),
                    )
                row_idx += 1

                # Plat rows
                for entry in ref_table:
                    if entry["type"] == "Plat":
                        ws.cell(row=row_idx, column=1, value=entry.get("ref", ""))
                        ws.cell(row=row_idx, column=2, value="Plat")
                        ws.cell(row=row_idx, column=3, value=entry.get("filename", ""))
                        ws.cell(row=row_idx, column=4, value=entry.get("owner", ""))
                        ws.cell(row=row_idx, column=5, value="")  # surveyor TBD
                        ws.cell(row=row_idx, column=6, value="")  # date TBD
                        ws.cell(row=row_idx, column=7, value=entry.get("relationship", ""))
                        for col in range(1, 8):
                            ws.cell(row=row_idx, column=col).font = Font(name="Calibri", size=10)
                            ws.cell(row=row_idx, column=col).border = Border(
                                bottom=Side(style="hair"), left=Side(style="hair"), right=Side(style="hair"),
                            )
                        row_idx += 1

                # Column widths
                ws.column_dimensions["A"].width = 8
                ws.column_dimensions["B"].width = 8
                ws.column_dimensions["C"].width = 25
                ws.column_dimensions["D"].width = 22
                ws.column_dimensions["E"].width = 22
                ws.column_dimensions["F"].width = 14
                ws.column_dimensions["G"].width = 14

                # Save to project E Research folder
                base = _job_base_path(job_number, client_name, job_type)
                ref_dir = base / "E Research"
                ref_dir.mkdir(parents=True, exist_ok=True)
                ref_xlsx = ref_dir / f"Reference_Table_{job_number}.xlsx"
                wb.save(str(ref_xlsx))
                session_data["reference_table_path"] = str(ref_xlsx)
                _log(f"  ✓ Excel reference table saved: {ref_xlsx.name}")

            except ImportError:
                _log("  ⚠ openpyxl not installed — skipped Excel ref table")
            except Exception as ex_ref:
                _log(f"  ⚠ Could not generate Excel ref table: {ex_ref}")

            # ── QA Checklist ───────────────────────────────────────────────
            qa = {
                "client_deed_found": any(s.get("deed_saved") for s in session_data["subjects"] if s["type"] == "client"),
                "client_plat_found": any(s.get("plat_saved") for s in session_data["subjects"] if s["type"] == "client"),
                "all_adjoiners_identified": len([s for s in session_data["subjects"] if s["type"] == "adjoiner"]) > 0 or wf["needs_adjoiners"] == "none",
                "adjoiners_with_deeds": sum(1 for s in session_data["subjects"] if s["type"] == "adjoiner" and s.get("deed_saved")),
                "adjoiners_total": sum(1 for s in session_data["subjects"] if s["type"] == "adjoiner"),
                "reference_table_generated": True,
                "needs_human_review": [],
            }
            # Flag items needing human attention
            if not qa["client_deed_found"]:
                qa["needs_human_review"].append("Client deed not found — manual search required")
            if not qa["client_plat_found"] and wf["research_depth"] != "minimal":
                qa["needs_human_review"].append("Client plat not found — check cabinet files manually")
            for s in session_data["subjects"]:
                if s["type"] == "adjoiner" and not s.get("deed_saved"):
                    qa["needs_human_review"].append(f"No deed for adjoiner: {s['name']}")

            session_data["qa_checklist"] = qa
            save_research(job_number, client_name, job_type, session_data)

            # Log QA summary
            _log(f"  QA: Client deed={'✓' if qa['client_deed_found'] else '✗'}, "
                 f"Client plat={'✓' if qa['client_plat_found'] else '✗'}, "
                 f"Adjoiners with deeds={qa['adjoiners_with_deeds']}/{qa['adjoiners_total']}")
            if qa["needs_human_review"]:
                _log(f"  ⚠ {len(qa['needs_human_review'])} items need human review")
                for item in qa["needs_human_review"][:5]:
                    _log(f"    → {item}")

            # ── Done ────────────────────────────────────────────────────────
            total_deeds = sum(1 for s in session_data["subjects"] if s.get("deed_saved"))
            total_plats = sum(1 for s in session_data["subjects"] if s.get("plat_saved"))
            total_subj  = len(session_data["subjects"])
            deliverables_str = ", ".join(wf.get("deliverables", []))

            _update_inquiry(inquiry_id, {
                "status": "complete",
                "step": 6,
                "research": {"job_number": job_number, "subjects": session_data["subjects"]},
                "updated_at": datetime.now().isoformat(),
            })
            _log(f"✅ COMPLETE — Job #{job_number}: {total_subj} subjects, {total_deeds} deeds, {total_plats} plats")
            _log(f"  Deliverables: {deliverables_str}")
            if wf.get('notes'):
                _log(f"  Workflow notes: {wf['notes']}")

    except Exception as e:
        traceback.print_exc()
        _update_inquiry(inquiry_id, {
            "status": "error",
            "error": str(e),
            "updated_at": datetime.now().isoformat(),
        })
        try:
            inq, _ = _find_inquiry(inquiry_id)
            if inq:
                log_list = inq.get("log", [])
                log_list.append(f"[ERROR] {str(e)}")
                _update_inquiry(inquiry_id, {"log": log_list})
        except Exception:
            pass


# ── AUTO-RESEARCH HELPER FUNCTIONS ────────────────────────────────────────────


@app.route("/api/workflows", methods=["GET"])
def api_workflows():
    """Return all available job type workflows with their configs.
    Used by the website inquiry form to show job type descriptions."""
    result = []
    for code, wf in _JOB_WORKFLOWS.items():
        result.append({
            "code": code,
            "name": wf["name"],
            "research_depth": wf["research_depth"],
            "needs_adjoiners": wf["needs_adjoiners"],
            "needs_chain": wf.get("needs_chain", False),
            "deliverables": wf.get("deliverables", []),
            "notes": wf.get("notes", ""),
        })
    return jsonify({"success": True, "workflows": result})

def _auto_search_deed(name: str, job_type: str = "BDY") -> dict | None:
    """Search 1stNMTitle for deeds matching `name`, return scored results.

    Uses the '__auto__' background session (logged in via _auto_login).
    Returns: { results: [...], best: {...}, count: int } or None on failure.
    """
    try:
        search_url = f"{_get_portal_url()}/scripts/hfweb.asp?Application=FNM&Database=TP"
        sess = _get_web_session("__auto__")  # use background session, not request session
        resp = sess.get(search_url, timeout=15)

        if 'CROSSNAMEFIELD' not in resp.text:
            print(f"[auto-search] Session expired — cannot search for {name}", flush=True)
            return None

        soup = BeautifulSoup(resp.text, "html.parser")
        form_data = _scrape_form_data(soup)
        if not form_data:
            return None

        # Search by grantor name
        form_data["CROSSNAMEFIELD"] = name
        form_data["CROSSNAMETYPE"] = "begin"
        form_data["CROSSTYPE"] = "GR"

        action = _get_portal_url() + "/scripts/hflook.asp"
        post_resp = sess.post(action, data=form_data, timeout=20)
        soup2 = BeautifulSoup(post_resp.text, "html.parser")

        results = []
        rows = soup2.find_all("tr")
        for row in rows:
            cells = row.find_all("td")
            if len(cells) < 9:
                continue
            doc_link = cells[1].find("a") if len(cells) > 1 else None
            if not doc_link:
                continue
            doc_no = doc_link.text.strip()
            if not doc_no or not re.match(r'^[A-Z0-9]+$', doc_no):
                continue
            r = {
                "doc_no":          doc_no,
                "location":        cells[2].text.strip() if len(cells) > 2 else "",
                "document_code":   cells[3].text.strip() if len(cells) > 3 else "",
                "instrument_type": cells[5].text.strip() if len(cells) > 5 else "",
                "recorded_date":   cells[7].text.strip() if len(cells) > 7 else "",
                "grantor":         cells[9].text.strip() if len(cells) > 9 else "",
                "grantee":         cells[10].text.strip() if len(cells) > 10 else "",
            }
            _score_search_result(r, client_name=name)
            results.append(r)

        if not results:
            return None

        # Sort: deed types first, then by relevance score descending, then recency
        def _rank(r):
            inst = (r.get("instrument_type") or "").lower()
            is_deed = any(kw in inst for kw in ['deed', 'warranty', 'quitclaim', 'grant', 'convey'])
            return (not is_deed, -r.get("relevance_score", 0), r.get("recorded_date", ""))

        results.sort(key=_rank)

        return {"results": results[:20], "best": results[0], "count": len(results)}

    except Exception as e:
        print(f"[auto-search] Error searching for {name}: {e}", flush=True)
        traceback.print_exc()
        return None


def _auto_save_deed(doc_no: str, detail: dict, job_number: str, client_name: str,
                    job_type: str, subject_id: str, is_adjoiner: bool = False,
                    adjoiner_name: str = "") -> dict:
    """Download and save a deed PDF to the project folder using RTSI naming (D1, D2...)."""
    try:
        grantor  = detail.get("grantor", "")
        grantee  = detail.get("grantee", "")
        location = detail.get("location", "")

        # Build save path first so we can count existing files for ref number
        base = _job_base_path(job_number, client_name, job_type)
        if is_adjoiner:
            dest_dir = base / "E Research" / "A Deeds" / "Adjoiners"
        else:
            dest_dir = base / "E Research" / "A Deeds"

        dest_dir.mkdir(parents=True, exist_ok=True)

        # Build RTSI-convention filename: D1 Grantor to Grantee.pdf
        loc_clean = re.sub(r'^[A-Z]', '', location.strip())
        def _cn(n):
            parts = n.split(",")
            return parts[0].strip().title() if parts else n.title()

        ref_num = _next_ref_number(dest_dir, prefix="D")

        if is_adjoiner and adjoiner_name:
            filename = f"D{ref_num} {_cn(adjoiner_name)} {doc_no}.pdf"
        else:
            filename = f"D{ref_num} {loc_clean} {_cn(grantor)} to {_cn(grantee)}.pdf"
        filename = re.sub(r'[<>:"/\\|?*]', '', filename).strip()
        if not filename.endswith(".pdf"):
            filename += ".pdf"

        save_path = dest_dir / filename


        if save_path.exists():
            return {"success": True, "skipped": True, "saved_to": str(save_path)}

        save_path.parent.mkdir(parents=True, exist_ok=True)

        # Fetch PDF from portal
        pdf_url = f"{_get_portal_url()}/WebTemp/{doc_no}.pdf"
        pdf_resp, pdf_err = _fetch_portal_pdf(doc_no, pdf_url)
        if pdf_err:
            return {"success": False, "error": pdf_err}

        with open(save_path, "wb") as f:
            for chunk in pdf_resp.iter_content(8192):
                f.write(chunk)

        return {"success": True, "skipped": False, "saved_to": str(save_path)}

    except Exception as e:
        return {"success": False, "error": str(e)}


def _auto_search_plat(name: str, job_number: str, job_type: str,
                      is_adjoiner: bool = False) -> dict | None:
    """Search for a plat in the local cabinet index via search_local_cabinet."""
    try:
        last_name = name.split(",")[0].strip()
        if len(last_name) < 2:
            return None

        cabinet_path = get_survey_data_path()
        if not cabinet_path:
            return None

        # Search all known cabinets for a name match
        src_path = ""
        cab_letter = ""
        for letter in CABINET_FOLDERS:
            results = _search_local_cabinet_impl(
                cabinet=letter,
                doc_num="",
                cabinet_path=cabinet_path,
                grantor=last_name,
            )
            if results:
                best_hit = results[0]
                src_path = best_hit.get("path", "")
                cab_letter = letter
                break

        if not src_path or not os.path.isfile(src_path):
            return None
        # Copy to project with RTSI P# naming convention
        base = _job_base_path(job_number, last_name, job_type)
        if is_adjoiner:
            dest_dir = base / "E Research" / "B Plats" / "Adjoiners"
        else:
            dest_dir = base / "E Research" / "B Plats"

        dest_dir.mkdir(parents=True, exist_ok=True)

        ref_num = _next_ref_number(dest_dir, prefix="P")
        orig_name = os.path.basename(src_path)
        dest = dest_dir / f"P{ref_num} {orig_name}"

        if not dest.exists():
            shutil.copy2(src_path, dest)

        return {"success": True, "saved_to": str(dest), "source": "cabinet", "cabinet": cab_letter}

    except Exception as e:
        print(f"[auto-plat] Error searching plat for {name}: {e}", flush=True)
        return None


def _auto_discover_adjoiners(upc: str, client_name: str) -> list:
    """Find adjacent parcels via ArcGIS spatial query."""
    try:
        # Get parcel geometry
        geometry = _arcgis_get_parcel_geometry(upc)
        if not geometry:
            # Fallback: KML index
            survey = get_survey_data_path()
            polygon = xml_processor.extract_parcel_polygon(survey, upc)
            if polygon and polygon.get("coordinates"):
                coords = polygon["coordinates"]
                geometry = {
                    "rings": [[[c[0], c[1]] for c in coords]],
                    "spatialReference": {"wkid": 4326}
                }

        if not geometry:
            return []

        raw = _arcgis_find_touching_parcels(geometry)

        # Filter out client's own parcel
        adjoiners = []
        seen = set()
        client_up = client_name.upper()
        for adj in raw:
            if adj["upc"] == upc:
                continue
            if adj["owner"].upper() == client_up:
                continue
            if adj["upc"] in seen:
                continue
            seen.add(adj["upc"])
            adjoiners.append(adj)

        return adjoiners

    except Exception as e:
        print(f"[auto-adjoiners] Error: {e}", flush=True)
        return []


# ── run ────────────────────────────────────────────────────────────────────────


if __name__ == "__main__":
    import socket as _sock
    def _get_lan_ip():
        try:
            s = _sock.socket(_sock.AF_INET, _sock.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except Exception:
            return "127.0.0.1"
    _lan = _get_lan_ip()
    print("=" * 60)
    print("  Deed & Plat Helper")
    print("  Local:   http://localhost:5000")
    print(f"  Network: http://{_lan}:5000")
    print("=" * 60)
    # Cabinet index already initialized at module-import time (see above).
    # No need to re-init here — this ensures desktop_app.py and direct
    # `python app.py` launches both have the cabinet index loaded.
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
